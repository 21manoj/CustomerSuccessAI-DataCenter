"""
Build CSPulse_FDE_Discovery.xlsx
6 tabs: README, CRO, CFO, CEO, VPCS, CSM, Consolidation
- Customer fills BLUE cells.
- BLACK formulas roll up.
- GREEN cells on Consolidation pull from persona tabs.
- YELLOW background = required.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ---------- styling tokens ----------
ARIAL = "Arial"
COLOR_INPUT   = "0000FF"  # blue
COLOR_FORMULA = "000000"  # black
COLOR_LINK    = "008000"  # green (cross-sheet)
COLOR_HEADER  = "FFFFFF"  # white text on dark fill

FILL_HEADER  = PatternFill("solid", start_color="1F3A5F")
FILL_SECTION = PatternFill("solid", start_color="2E5F8F")
FILL_REQUIRED= PatternFill("solid", start_color="FFF59D")  # light yellow
FILL_BAND    = PatternFill("solid", start_color="F2F2F2")
FILL_FLAG    = PatternFill("solid", start_color="FFCDD2")  # light red

THIN = Side(border_style="thin", color="B8B8B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE   = Font(name=ARIAL, size=18, bold=True, color="1F3A5F")
FONT_HEADER  = Font(name=ARIAL, size=11, bold=True, color=COLOR_HEADER)
FONT_SECTION = Font(name=ARIAL, size=11, bold=True, color=COLOR_HEADER)
FONT_LABEL   = Font(name=ARIAL, size=10, bold=True)
FONT_BODY    = Font(name=ARIAL, size=10)
FONT_INPUT   = Font(name=ARIAL, size=10, color=COLOR_INPUT, bold=True)
FONT_FORMULA = Font(name=ARIAL, size=10, color=COLOR_FORMULA)
FONT_LINK    = Font(name=ARIAL, size=10, color=COLOR_LINK)
FONT_NOTE    = Font(name=ARIAL, size=9, italic=True, color="666666")

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------- helpers ----------
def write_title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 26

def write_section_header(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 20

def write_col_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28

def style_input(ws, row, col, required=False):
    c = ws.cell(row=row, column=col)
    c.font = FONT_INPUT
    if required:
        c.fill = FILL_REQUIRED
    c.border = BORDER
    c.alignment = WRAP

def style_formula(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.font = FONT_FORMULA
    c.border = BORDER
    c.alignment = WRAP

def style_link(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.font = FONT_LINK
    c.border = BORDER
    c.alignment = WRAP

def style_label(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.font = FONT_LABEL
    c.border = BORDER
    c.alignment = WRAP
    c.fill = FILL_BAND

def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- README tab ----------
def build_readme(ws):
    ws.title = "README"
    set_column_widths(ws, [22, 18, 18, 18, 18, 18])
    write_title(ws, 1, "CS Pulse — FDE Discovery Workbook")

    rows = [
        ("Purpose", "Capture per-persona discovery data so we can configure the customer's tenant: KPI selections, pillar weights, signal channels, stakeholders, success outcomes, and decision cadence."),
        ("How to use", "FDE sends the relevant tab to each persona's stakeholder. Stakeholder fills blue cells. FDE reviews, then the Consolidation tab rolls everything into a config-export view that maps directly to CustomerConfig fields."),
        ("Color legend — blue", "Input cell. The stakeholder fills this in."),
        ("Color legend — yellow", "Required input. Do not skip."),
        ("Color legend — black", "Formula. Do not edit."),
        ("Color legend — green", "Cross-sheet link on Consolidation. Pulls from persona tabs."),
        ("Color legend — red", "Conflict flag. Two personas disagree by more than the tolerance band."),
        ("Tabs", "README · CRO · CFO · CEO · VPCS · CSM · Consolidation"),
        ("Persona owner mapping", "CRO = revenue. CFO = ROI / proof. CEO = portfolio rollup. VP CS = team + playbooks. CSM = daily account work."),
        ("Pillar legend", "P1 Deployment Velocity · P2 Operational Stability · P3 Workload Performance · P4 Channel & Partner Health · P5 Expansion Readiness. (Names vary by vertical — defaults shown are DC2_S.)"),
        ("KPI tier", "Default = Predictive 11 (11 KPIs across the 5 pillars). Pick Starter 9 for early-stage customers, or Full 43 for mature data programs."),
        ("Default rule", "If a stakeholder cannot answer, use the platform default and flag it in the Notes column. Do not invent."),
        ("Sign-off", "When all 5 tabs are filled and Consolidation has no red flags, the engagement lead countersigns and onboarding moves to the 4-CSV upload."),
    ]
    r = 3
    for label, body in rows:
        c = ws.cell(row=r, column=1, value=label)
        c.font = FONT_LABEL
        c.alignment = WRAP
        c.fill = FILL_BAND
        c.border = BORDER

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c2 = ws.cell(row=r, column=2, value=body)
        c2.font = FONT_BODY
        c2.alignment = WRAP
        c2.border = BORDER
        ws.row_dimensions[r].height = max(28, 14 * (len(body) // 70 + 1))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Version 1.0 · May 2026 · Internal — NDA covered").font = FONT_NOTE


# ---------- persona tab builder ----------
# Five personas share the same shape. Differences:
#  - title + intro
#  - the 10 "what good looks like" eval question prompts
#  - the suggested current-KPI candidates (advisory only)
PERSONAS = [
    {
        "key": "CRO",
        "title": "CRO — Chief Revenue Officer",
        "intro": "Owner of forward revenue. Cares about: NRR forecast, revenue at risk, expansion pipeline, account-level explainability.",
        "default_pillar_weights": [0.15, 0.20, 0.25, 0.15, 0.25],  # P1..P5
        "evals": [
            "Can I see total revenue at risk in the next quarter, by account, in under 10 seconds?",
            "When the NRR forecast moves, can I see which accounts moved it and why?",
            "Are the top 3 expansion opportunities ranked by dollars and confidence, not vibes?",
            "Can I trust the forecast — is the methodology defensible to my board?",
            "Does the dashboard tell me which CSM owns each at-risk account?",
            "When an account flips from healthy to at-risk, do I get an alert with the trigger?",
            "Can I compare this quarter's risk against last quarter and last year?",
            "Does the dashboard show the confidence interval on the NRR forecast?",
            "Can I drill from portfolio rollup down to a single account in two clicks?",
            "Is the revenue-at-risk number reconcile-able to the CRM?",
        ],
        "current_kpi_hints": [
            "NRR (gross / net)", "Logo retention", "Pipeline coverage on renewals",
            "Expansion bookings", "At-risk ARR", "Churn $", "Time to first value",
        ],
    },
    {
        "key": "CFO",
        "title": "CFO — Chief Financial Officer",
        "intro": "Owner of ROI proof. Cares about: every dollar of CS investment is justified, attribution is traceable, audit trail is defensible.",
        "default_pillar_weights": [0.20, 0.25, 0.20, 0.15, 0.20],
        "evals": [
            "Can I prove the ROI of CS investment with a number my auditor will accept?",
            "Is every Revenue Protected dollar traceable to a specific playbook that ran?",
            "Does CS investment scale with ARR — what's the budget rationale at 2x revenue?",
            "Are the headline numbers (Revenue Protected, Realized ROI) reconcile-able to GL?",
            "Does the audit trail show every assumption behind the ROI claim?",
            "Can I see CS investment broken down by playbook category?",
            "Is the Power-of-1 lift assumption disclosed and bounded?",
            "Does the dashboard distinguish realized vs forecasted dollars?",
            "Can I export the ROI calculation for board materials?",
            "Are model confidence intervals visible where they matter?",
        ],
        "current_kpi_hints": [
            "CS cost as % of ARR", "Revenue Protected $", "Realized ROI x",
            "CAC payback", "Gross margin on CS-touched accounts", "CS budget vs. plan",
        ],
    },
    {
        "key": "CEO",
        "title": "CEO — Chief Executive Officer",
        "intro": "Owner of the portfolio rollup. Cares about: one screen, every portfolio company, board-ready summary numbers.",
        "default_pillar_weights": [0.15, 0.20, 0.20, 0.20, 0.25],
        "evals": [
            "Can I see one screen with every business unit / portfolio company's health?",
            "Are the headline numbers on this screen the same numbers I will quote to the board?",
            "Can I sort the portfolio by NRR forecast, risk concentration, or growth?",
            "Does the dashboard flag the 2 or 3 things that need my attention this week?",
            "Can I see cross-customer / cross-BU comparison without exporting to Excel?",
            "Is the rollup methodology disclosed (revenue-weighted vs. simple avg)?",
            "Are the underlying assumptions (NRR ranges, risk thresholds) settable from this screen?",
            "Does each portfolio company link to its own full dashboard in one click?",
            "Is there a quarterly trend view I can take into the QBR?",
            "Are governance and AI-DD disclosures available on demand?",
        ],
        "current_kpi_hints": [
            "Portfolio-weighted NRR", "Top 5 at-risk accounts $",
            "ARR by segment", "Growth rate", "Customer concentration",
            "Board-ready exec summary", "Per-BU CS investment",
        ],
    },
    {
        "key": "VPCS",
        "title": "VP CS — VP of Customer Success",
        "intro": "Owner of the CS function. Cares about: team capacity, CSM performance, playbook execution rate, weekly business review readiness.",
        "default_pillar_weights": [0.20, 0.20, 0.20, 0.20, 0.20],
        "evals": [
            "Can I see CSM-by-CSM performance and book of business in one view?",
            "Does the dashboard tell me which CSMs are overloaded vs. under capacity?",
            "Are playbook execution rates trended week over week?",
            "Can I see which playbooks are succeeding vs. failing, with attribution?",
            "Does the dashboard surface accounts that have no recent activity?",
            "Are CSM rankings transparent and explainable?",
            "Can I run a weekly business review straight from this dashboard?",
            "Does the system flag CSM coaching opportunities (where playbooks failed)?",
            "Can I see team capacity vs. account count and adjust assignments?",
            "Are CSM success metrics tied to actual revenue outcomes, not activity vanity?",
        ],
        "current_kpi_hints": [
            "Accounts per CSM", "Playbook execution rate %",
            "Playbook success rate %", "QBR completion rate", "CSM-to-revenue ratio",
            "Average response time to risk signal", "Team capacity index",
        ],
    },
    {
        "key": "CSM",
        "title": "CSM — Customer Success Manager",
        "intro": "Owner of the daily account work. Cares about: today's Kanban, what to do next, signal triage, time-to-next-action.",
        "default_pillar_weights": [0.20, 0.25, 0.20, 0.15, 0.20],
        "evals": [
            "When I log in, does the dashboard tell me what to do today, in priority order?",
            "Is the Kanban split clearly into FIRE / THIS WEEK / OPPORTUNITIES?",
            "For each at-risk account, does the system suggest a specific playbook?",
            "Can I see signal triage (Slack mentions, CRM events) for my accounts in one feed?",
            "Does the system tell me which accounts I haven't touched in N days?",
            "Can I close a playbook with one click and have the outcome flow upstream?",
            "Are recommended actions specific (\"call sponsor\") not vague (\"engage\")?",
            "Can I drag accounts across Kanban columns and have the system remember?",
            "Does the system give me a per-account talk track for the next call?",
            "Can I see my own success metrics without my manager telling me?",
        ],
        "current_kpi_hints": [
            "Accounts with overdue follow-up", "Open playbook count",
            "Win rate on assigned playbooks", "Avg time to first response on risk signal",
            "QBR completion this quarter", "NPS / sentiment for owned accounts",
        ],
    },
]

# Column layout for persona tabs:
# A: section / field label
# B: input cell (blue)
# C: notes (blue, optional)
# D: default / hint (gray label)
# E: required flag (Y/N)
# F: source / system reference
PERSONA_COLS = ["Field", "Answer (your input)", "Notes", "Default / hint", "Required?", "Source / system"]
PERSONA_WIDTHS = [38, 30, 28, 24, 11, 22]


def build_persona(ws, persona):
    ws.title = persona["key"]
    set_column_widths(ws, PERSONA_WIDTHS)
    write_title(ws, 1, persona["title"])

    ws.cell(row=2, column=1, value=persona["intro"]).font = FONT_NOTE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 36

    r = 4

    # Section A: Stakeholder identity
    write_section_header(ws, r, "A. Stakeholder identity"); r += 1
    write_col_headers(ws, r, PERSONA_COLS); r += 1
    identity_rows = [
        ("Full name", "", "", "", "Y", ""),
        ("Title", "", "", persona["title"].split("—")[0].strip(), "Y", ""),
        ("Email", "", "", "", "Y", ""),
        ("Time zone", "", "", "PT / ET / CET / IST", "Y", ""),
        ("Years in role", "", "", "0-30", "N", ""),
        ("Reports to", "", "", "", "N", ""),
    ]
    for label, ans, note, default, req, src in identity_rows:
        style_label(ws, r, 1); ws.cell(row=r, column=1, value=label)
        style_input(ws, r, 2, required=(req == "Y"))
        style_input(ws, r, 3)
        ws.cell(row=r, column=4, value=default).font = FONT_NOTE; ws.cell(row=r, column=4).border = BORDER
        ws.cell(row=r, column=5, value=req).font = FONT_LABEL; ws.cell(row=r, column=5).border = BORDER; ws.cell(row=r, column=5).alignment = CENTER
        style_input(ws, r, 6)
        r += 1
    r += 1

    # Section B: Top 5 pain points
    write_section_header(ws, r, "B. Top 5 pain points (rank 1=worst, 5=mildly annoying)"); r += 1
    write_col_headers(ws, r, ["Rank", "Pain point", "Frequency (daily/weekly/monthly)", "Cost if unsolved", "Severity (1-5)", "Owner today"]); r += 1
    for i in range(1, 6):
        ws.cell(row=r, column=1, value=i).font = FONT_LABEL
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = BORDER
        for col in range(2, 7):
            style_input(ws, r, col, required=(i == 1))
        r += 1
    r += 1

    # Section C: Current KPIs they track today (~8 rows)
    write_section_header(ws, r, "C. Current KPIs you track today"); r += 1
    write_col_headers(ws, r, ["KPI name", "Source system", "Frequency", "Owner", "Is this in CS Pulse today? (Y/N)", "Notes"]); r += 1
    hints = persona["current_kpi_hints"]
    for i in range(8):
        hint_text = hints[i] if i < len(hints) else ""
        ws.cell(row=r, column=1, value=hint_text).font = FONT_INPUT
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = WRAP
        for col in range(2, 7):
            style_input(ws, r, col, required=(i < 3))
        r += 1
    r += 1

    # Section D: Pillar weight allocation (must sum to 100%)
    write_section_header(ws, r, "D. Pillar weight allocation (your view) — must sum to 100%"); r += 1
    write_col_headers(ws, r, ["Pillar", "Your weight %", "Platform default %", "Delta", "Justification", ""]); r += 1
    pillar_names = [
        "P1 Deployment Velocity",
        "P2 Operational Stability",
        "P3 Workload Performance",
        "P4 Channel & Partner Health",
        "P5 Expansion Readiness",
    ]
    pillar_start = r
    for i, pn in enumerate(pillar_names):
        ws.cell(row=r, column=1, value=pn).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = WRAP
        # input weight (blue, yellow bg)
        style_input(ws, r, 2, required=True)
        ws.cell(row=r, column=2).number_format = "0.0%"
        # default
        default_pct = persona["default_pillar_weights"][i]
        c_def = ws.cell(row=r, column=3, value=default_pct)
        c_def.font = FONT_NOTE; c_def.border = BORDER
        c_def.number_format = "0.0%"
        # delta = input - default
        f_delta = f"=IFERROR(B{r}-C{r},0)"
        c_delta = ws.cell(row=r, column=4, value=f_delta)
        c_delta.font = FONT_FORMULA; c_delta.border = BORDER
        c_delta.number_format = "+0.0%;-0.0%;-"
        style_input(ws, r, 5)
        ws.cell(row=r, column=6).border = BORDER
        r += 1
    # sum row
    ws.cell(row=r, column=1, value="Sum (must = 100%)").font = FONT_LABEL
    ws.cell(row=r, column=1).border = BORDER
    f_sum = f"=SUM(B{pillar_start}:B{r-1})"
    c_sum = ws.cell(row=r, column=2, value=f_sum)
    c_sum.font = FONT_FORMULA; c_sum.border = BORDER
    c_sum.number_format = "0.0%"
    # check formula in column D
    f_check = f'=IF(ABS(B{r}-1)<0.001,"OK","Adjust — must total 100%")'
    c_check = ws.cell(row=r, column=4, value=f_check)
    c_check.font = FONT_FORMULA; c_check.border = BORDER
    pillar_sum_row = r
    r += 2

    # Section E: Signal sources
    write_section_header(ws, r, "E. Signal sources (where does the customer hear about risk / expansion today?)"); r += 1
    write_col_headers(ws, r, ["Channel type", "Identifier (channel / alias / tool)", "Frequency", "Who monitors it?", "Wire to CS Pulse? (Y/N)", "Notes"]); r += 1
    signal_rows = [
        ("Slack channel(s)", "", "Daily / Weekly", "", "Y", "e.g. #cs-acme"),
        ("Shared email alias", "", "", "", "Y", "e.g. cs-team@..."),
        ("Call transcript tool", "", "", "", "Y", "Gong / Chorus / Zoom"),
        ("CRM events", "", "", "", "Y", "Salesforce / HubSpot"),
        ("Support tickets", "", "", "", "N", "Zendesk / ServiceNow"),
        ("Other", "", "", "", "N", ""),
    ]
    for ch, ident, freq, mon, wire, note in signal_rows:
        ws.cell(row=r, column=1, value=ch).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = WRAP
        # input cell holds identifier
        c = ws.cell(row=r, column=2, value=ident)
        c.font = FONT_INPUT; c.border = BORDER; c.alignment = WRAP
        for col in [3, 4, 5]:
            style_input(ws, r, col)
        ws.cell(row=r, column=6, value=note).font = FONT_NOTE
        ws.cell(row=r, column=6).border = BORDER
        r += 1
    r += 1

    # Section F: Persona-eval question set (10 questions)
    write_section_header(ws, r, "F. Persona-eval questions — how would you score the platform today? (0=no, 1=partial, 2=yes)"); r += 1
    write_col_headers(ws, r, ["#", "Question", "Score (0/1/2)", "Today's gap", "Priority (1=top)", "Owner to close"]); r += 1
    eval_start = r
    for i, q in enumerate(persona["evals"], 1):
        ws.cell(row=r, column=1, value=i).font = FONT_LABEL
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=q).font = FONT_BODY
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        style_input(ws, r, 3, required=True)
        ws.cell(row=r, column=3).number_format = "0"
        style_input(ws, r, 4)
        style_input(ws, r, 5)
        ws.cell(row=r, column=5).number_format = "0"
        style_input(ws, r, 6)
        r += 1

    # eval total row
    ws.cell(row=r, column=1, value="Total").font = FONT_LABEL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="Pass = 16 / 20").font = FONT_NOTE
    ws.cell(row=r, column=2).border = BORDER
    f_total = f"=SUM(C{eval_start}:C{r-1})"
    c_total = ws.cell(row=r, column=3, value=f_total)
    c_total.font = FONT_FORMULA; c_total.border = BORDER
    c_total.number_format = "0"
    # pass/fail
    f_pf = f'=IF(C{r}>=16,"PASS","BELOW THRESHOLD")'
    c_pf = ws.cell(row=r, column=4, value=f_pf)
    c_pf.font = FONT_FORMULA; c_pf.border = BORDER
    eval_total_row = r
    r += 2

    # Section G: Success outcomes
    write_section_header(ws, r, "G. Success outcomes — what does winning look like?"); r += 1
    write_col_headers(ws, r, ["Horizon", "Outcome (measurable)", "Target value", "Source of truth", "Who reports it?", "Notes"]); r += 1
    horizons = [("30 days", "Y"), ("90 days", "Y"), ("12 months", "Y")]
    for h, req in horizons:
        ws.cell(row=r, column=1, value=h).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = CENTER
        for col in range(2, 7):
            style_input(ws, r, col, required=(req == "Y"))
        r += 1
    r += 1

    # Section H: Decision cadence + cross-functional
    write_section_header(ws, r, "H. Decision cadence + cross-functional"); r += 1
    write_col_headers(ws, r, ["Question", "Answer", "Frequency", "Stakeholders", "Tool / forum", "Notes"]); r += 1
    cadence_rows = [
        "When do you make decisions based on this data?",
        "Who else sees these numbers with you?",
        "What is your current dashboard / tool of record?",
        "What format do you need the data in for executives?",
        "Where do KPI definitions get debated?",
    ]
    for q in cadence_rows:
        ws.cell(row=r, column=1, value=q).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = WRAP
        for col in range(2, 7):
            style_input(ws, r, col)
        r += 1
    r += 1

    # Footer: tab metadata for cross-sheet linking
    ws.cell(row=r, column=1, value="-- internal anchors (do not edit) --").font = FONT_NOTE
    r += 1
    ws.cell(row=r, column=1, value="pillar_sum_row").font = FONT_NOTE
    ws.cell(row=r, column=2, value=pillar_sum_row).font = FONT_NOTE
    r += 1
    ws.cell(row=r, column=1, value="eval_total_row").font = FONT_NOTE
    ws.cell(row=r, column=2, value=eval_total_row).font = FONT_NOTE

    # Return anchors so consolidation can reference them
    return {
        "pillar_first_row": pillar_start,     # P1 row
        "pillar_sum_row": pillar_sum_row,
        "eval_total_row": eval_total_row,
    }


# ---------- Consolidation tab ----------
def build_consolidation(ws, persona_anchors):
    ws.title = "Consolidation"
    set_column_widths(ws, [30, 14, 14, 14, 14, 14, 14, 18])
    write_title(ws, 1, "Consolidation — cross-persona rollup + CustomerConfig export")
    ws.cell(row=2, column=1, value="Green cells pull live from persona tabs. Black cells are formulas. Red flags = conflict beyond tolerance.").font = FONT_NOTE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    r = 4
    # ---- Pillar weight comparison ----
    write_section_header(ws, r, "1. Pillar weight comparison (per-persona)"); r += 1
    persona_keys = ["CRO", "CFO", "CEO", "VPCS", "CSM"]
    write_col_headers(ws, r, ["Pillar", "CRO", "CFO", "CEO", "VPCS", "CSM", "Mean (recommended)", "Range (max - min)"])
    r += 1
    pillar_names = [
        "P1 Deployment Velocity",
        "P2 Operational Stability",
        "P3 Workload Performance",
        "P4 Channel & Partner Health",
        "P5 Expansion Readiness",
    ]
    for i, pname in enumerate(pillar_names):
        ws.cell(row=r, column=1, value=pname).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = WRAP
        # link to each persona's row B[pillar_first_row + i]
        for j, pk in enumerate(persona_keys):
            anchor = persona_anchors[pk]["pillar_first_row"] + i
            f = f"='{pk}'!B{anchor}"
            c = ws.cell(row=r, column=2 + j, value=f)
            c.font = FONT_LINK; c.border = BORDER
            c.number_format = "0.0%"
        # mean (col G)
        first_col = get_column_letter(2)
        last_col = get_column_letter(6)
        f_mean = f"=IFERROR(AVERAGE({first_col}{r}:{last_col}{r}),0)"
        c_mean = ws.cell(row=r, column=7, value=f_mean)
        c_mean.font = FONT_FORMULA; c_mean.border = BORDER
        c_mean.number_format = "0.0%"
        # range
        f_range = f"=IFERROR(MAX({first_col}{r}:{last_col}{r})-MIN({first_col}{r}:{last_col}{r}),0)"
        c_range = ws.cell(row=r, column=8, value=f_range)
        c_range.font = FONT_FORMULA; c_range.border = BORDER
        c_range.number_format = "0.0%"
        r += 1
    # sum check row
    ws.cell(row=r, column=1, value="Sum check (each column should = 100%)").font = FONT_LABEL
    ws.cell(row=r, column=1).border = BORDER
    for j in range(5):
        col_letter = get_column_letter(2 + j)
        f_s = f"=SUM({col_letter}{r-5}:{col_letter}{r-1})"
        c_s = ws.cell(row=r, column=2 + j, value=f_s)
        c_s.font = FONT_FORMULA; c_s.border = BORDER
        c_s.number_format = "0.0%"
    r += 2

    # Conflict flags
    write_section_header(ws, r, "2. Conflict flags (range > 10pp on any pillar = red)"); r += 1
    write_col_headers(ws, r, ["Pillar", "Range", "Status", "Action", "", "", "", ""]); r += 1
    for i, pname in enumerate(pillar_names):
        ws.cell(row=r, column=1, value=pname).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        # range pulled from section 1 (offset back to first table)
        source_range_row = 6 + i  # section 1 header at row 5, first pillar row at 6
        f_rg = f"=H{source_range_row}"
        c_rg = ws.cell(row=r, column=2, value=f_rg)
        c_rg.font = FONT_FORMULA; c_rg.border = BORDER
        c_rg.number_format = "0.0%"
        f_st = f'=IF(B{r}>0.10,"CONFLICT","OK")'
        c_st = ws.cell(row=r, column=3, value=f_st)
        c_st.font = FONT_FORMULA; c_st.border = BORDER
        f_act = f'=IF(B{r}>0.10,"FDE to drive consensus before onboarding","No action — within tolerance")'
        c_act = ws.cell(row=r, column=4, value=f_act)
        c_act.font = FONT_FORMULA; c_act.border = BORDER
        r += 1
    r += 1

    # Persona eval score rollup
    write_section_header(ws, r, "3. Persona-eval scoreboard"); r += 1
    write_col_headers(ws, r, ["Persona", "Score (0-20)", "Threshold", "Pass?", "Gap to fix first", "", "", ""]); r += 1
    eval_start_row = r
    for pk in persona_keys:
        anchor_row = persona_anchors[pk]["eval_total_row"]
        ws.cell(row=r, column=1, value=pk).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        f_score = f"=IFERROR('{pk}'!C{anchor_row},0)"
        c_score = ws.cell(row=r, column=2, value=f_score)
        c_score.font = FONT_LINK; c_score.border = BORDER
        c_score.number_format = "0"
        ws.cell(row=r, column=3, value=16).font = FONT_FORMULA
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=3).number_format = "0"
        f_pass = f'=IF(B{r}>=C{r},"PASS","BELOW")'
        c_pass = ws.cell(row=r, column=4, value=f_pass)
        c_pass.font = FONT_FORMULA; c_pass.border = BORDER
        style_input(ws, r, 5)
        r += 1
    # portfolio rollup
    ws.cell(row=r, column=1, value="Portfolio total").font = FONT_LABEL
    ws.cell(row=r, column=1).border = BORDER
    f_tot = f"=SUM(B{eval_start_row}:B{r-1})"
    c_tot = ws.cell(row=r, column=2, value=f_tot)
    c_tot.font = FONT_FORMULA; c_tot.border = BORDER
    c_tot.number_format = "0"
    ws.cell(row=r, column=3, value=80).font = FONT_FORMULA
    ws.cell(row=r, column=3).border = BORDER
    f_p = f'=IF(B{r}>=80,"READY FOR HANDOVER","NEEDS MORE CALIBRATION")'
    c_p = ws.cell(row=r, column=4, value=f_p)
    c_p.font = FONT_FORMULA; c_p.border = BORDER
    r += 2

    # CustomerConfig export view
    write_section_header(ws, r, "4. CustomerConfig export — paste-ready for create_customer / configure_customer_kpis"); r += 1
    write_col_headers(ws, r, ["CustomerConfig field", "Value to write", "Source", "Notes", "", "", "", ""]); r += 1
    # Pillar weights → dc2s_pillar_weights (use mean from section 1)
    config_rows = [
        ("dc2s_pillar_weights.P1", "=G6",  "Mean of 5 personas (section 1)", "JSON key P1"),
        ("dc2s_pillar_weights.P2", "=G7",  "Mean of 5 personas (section 1)", "JSON key P2"),
        ("dc2s_pillar_weights.P3", "=G8",  "Mean of 5 personas (section 1)", "JSON key P3"),
        ("dc2s_pillar_weights.P4", "=G9",  "Mean of 5 personas (section 1)", "JSON key P4"),
        ("dc2s_pillar_weights.P5", "=G10", "Mean of 5 personas (section 1)", "JSON key P5"),
        ("kpi_tier", "Predictive 11", "Default", "Override to Starter 9 or Full 43 if customer asks"),
        ("vertical", "DC2_S", "Default", "Set per engagement"),
        ("signal_channels", "See section 5", "Per-persona signal sources rolled up", "Free-form list, deduped"),
        ("customer_playbooks", "Per-tenant playbook rows via /api/playbooks/library", "FDE via admin UI or MCP", "DB table customer_playbooks (customer_playbook_api.py)"),
    ]
    for field, val, src, note in config_rows:
        ws.cell(row=r, column=1, value=field).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        c_v = ws.cell(row=r, column=2, value=val)
        if str(val).startswith("="):
            c_v.font = FONT_FORMULA
            c_v.number_format = "0.0%"
        else:
            c_v.font = FONT_INPUT
        c_v.border = BORDER
        ws.cell(row=r, column=3, value=src).font = FONT_NOTE
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=4, value=note).font = FONT_NOTE
        ws.cell(row=r, column=4).border = BORDER
        r += 1
    r += 1

    # Signal channel rollup
    write_section_header(ws, r, "5. Signal channels rolled up — wire into signal_channels.json"); r += 1
    write_col_headers(ws, r, ["Channel type", "CRO", "CFO", "CEO", "VPCS", "CSM", "Wire? (Y/N)", "FDE notes"]); r += 1
    signal_rows = [
        ("Slack channel(s)",    32),  # ~row index of each persona's signal Slack row (will resolve per persona)
        ("Shared email alias",  33),
        ("Call transcript tool", 34),
        ("CRM events",          35),
        ("Support tickets",     36),
        ("Other",               37),
    ]
    # Build via dynamic references — persona signal section header is consistent because we built each tab the same way.
    # Approximate signal-row anchor: section E starts after sections A (10 rows) + B (8 rows) + C (10 rows) + D (8 rows) + spacers.
    # Simpler: pull from a known column position relative to the section header (we drew it in the build_persona loop).
    # For paste-readiness we just show free-form input; FDE consolidates manually after reading each tab.
    for ch, _ in signal_rows:
        ws.cell(row=r, column=1, value=ch).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        for col in range(2, 7):
            style_input(ws, r, col)
        style_input(ws, r, 7, required=True)
        style_input(ws, r, 8)
        r += 1
    r += 1

    # Sign-off
    write_section_header(ws, r, "6. Sign-off"); r += 1
    write_col_headers(ws, r, ["Role", "Name", "Date", "Signature / approval", "", "", "", ""]); r += 1
    signoff_rows = [
        "FDE (engagement lead)",
        "Customer executive sponsor",
        "CS Pulse customer engagement lead",
    ]
    for role in signoff_rows:
        ws.cell(row=r, column=1, value=role).font = FONT_LABEL
        ws.cell(row=r, column=1).border = BORDER
        for col in range(2, 5):
            style_input(ws, r, col, required=True)
        r += 1


# ---------- main ----------
def main():
    wb = Workbook()

    # README tab uses the default first sheet
    readme = wb.active
    build_readme(readme)

    # Persona tabs
    anchors = {}
    for p in PERSONAS:
        ws = wb.create_sheet(p["key"])
        anchors[p["key"]] = build_persona(ws, p)

    # Consolidation
    consolidation = wb.create_sheet("Consolidation")
    build_consolidation(consolidation, anchors)

    # Freeze top rows on persona tabs
    for p in PERSONAS:
        wb[p["key"]].freeze_panes = "A5"
    wb["Consolidation"].freeze_panes = "A5"

    out = os.path.join(os.path.dirname(__file__), "CSPulse_FDE_Discovery.xlsx")
    wb.save(out)
    print("Wrote:", out)


if __name__ == "__main__":
    main()
