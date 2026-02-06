#!/usr/bin/env python3
"""
Excel Import Service
====================

Parses Excel templates (real customer data) and imports into CS Pulse.

Handles:
- Customer Profile sheet
- KPI Coverage Detail sheet
- KPI History (if available)
- Contracts & Revenue
- NPS & CSAT
- Events

Integrates with:
- KPI Normalization Service (raw values → 0-100)
- Sparse KPI Handler (10-15 KPIs validation)
- Hierarchical Health Calculator (with weight renormalization)
- Qdrant (journey_weeks collection)
"""

import openpyxl
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

# Import our services
from kpi_normalization_service import KPINormalizationService
from sparse_kpi_handler import SparseKPIHandler, DataQualityLevel


@dataclass
class ImportValidationError:
    """Validation error during import"""
    sheet_name: str
    row_number: Optional[int]
    field_name: Optional[str]
    error_message: str
    severity: str  # 'error', 'warning', 'info'


@dataclass
class CustomerImportData:
    """Imported customer data"""
    
    # Customer Profile
    account_id: str
    account_name: str
    industry: str
    region: str
    tier: str
    arr: float
    starting_arr: float
    
    # KPI Coverage
    tracked_kpis: List[str]
    kpi_coverage_pct: float
    data_quality_level: str
    
    # KPI History (if available)
    kpi_history: List[Dict]  # [{week: 1, date: '2024-01-01', kpis: {...}}, ...]
    
    # Contracts
    contract_start: Optional[datetime]
    contract_end: Optional[datetime]
    contract_value: Optional[float]
    
    # NPS/CSAT (if available)
    nps_csat_data: List[Dict]
    
    # Events (if available)
    events: List[Dict]
    
    # Validation results
    validation_errors: List[ImportValidationError]
    is_valid: bool


class ExcelImportService:
    """
    Import customer data from Excel templates
    
    Process:
    1. Parse Excel file
    2. Validate structure
    3. Normalize raw KPI values → 0-100 scores
    4. Check 30% KPI coverage threshold
    5. Prepare data for Qdrant upload
    """
    
    def __init__(self):
        """Initialize import service with dependencies"""
        self.normalization_service = KPINormalizationService()
        self.sparse_kpi_handler = SparseKPIHandler()
        
        # Required sheets
        self.required_sheets = [
            'Customer Profile',
            'KPI Coverage Detail'
        ]
        
        # Optional sheets
        self.optional_sheets = [
            'KPI History',
            'Contracts & Revenue',
            'NPS & CSAT',
            'Events'
        ]
    
    def import_from_file(self, excel_path: str) -> List[CustomerImportData]:
        """
        Import customers from Excel file
        
        Args:
            excel_path: Path to Excel file
        
        Returns:
            List of CustomerImportData objects (one per customer)
        """
        
        if not Path(excel_path).exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Validate structure
        validation_errors = self._validate_structure(wb)
        if any(e.severity == 'error' for e in validation_errors):
            raise ValueError(f"Excel validation failed: {validation_errors}")
        
        # Parse customer profiles
        customers = self._parse_customer_profiles(wb)
        
        # Parse KPI coverage for each customer
        kpi_coverage = self._parse_kpi_coverage(wb)
        
        # Merge coverage data into customers
        for customer in customers:
            if customer.account_id in kpi_coverage:
                coverage = kpi_coverage[customer.account_id]
                customer.tracked_kpis = coverage['kpis']
                customer.kpi_coverage_pct = coverage['coverage_pct']
                customer.data_quality_level = coverage['quality_level']
                
                # Validate 30% threshold
                if coverage['coverage_pct'] < 30:
                    customer.validation_errors.append(
                        ImportValidationError(
                            sheet_name='KPI Coverage Detail',
                            row_number=None,
                            field_name='kpi_coverage_pct',
                            error_message=f"KPI coverage {coverage['coverage_pct']:.0f}% below 30% threshold",
                            severity='warning'
                        )
                    )
        
        # Parse optional sheets if available
        if 'KPI History' in wb.sheetnames:
            kpi_history = self._parse_kpi_history(wb)
            for customer in customers:
                if customer.account_id in kpi_history:
                    customer.kpi_history = kpi_history[customer.account_id]
        
        if 'Contracts & Revenue' in wb.sheetnames:
            contracts = self._parse_contracts(wb)
            for customer in customers:
                if customer.account_id in contracts:
                    contract = contracts[customer.account_id]
                    customer.contract_start = contract['start_date']
                    customer.contract_end = contract['end_date']
                    customer.contract_value = contract['value']
        
        if 'NPS & CSAT' in wb.sheetnames:
            nps_csat = self._parse_nps_csat(wb)
            for customer in customers:
                if customer.account_id in nps_csat:
                    customer.nps_csat_data = nps_csat[customer.account_id]
        
        if 'Events' in wb.sheetnames:
            events = self._parse_events(wb)
            for customer in customers:
                if customer.account_id in events:
                    customer.events = events[customer.account_id]
        
        # Mark as valid if no errors
        for customer in customers:
            customer.is_valid = not any(
                e.severity == 'error' for e in customer.validation_errors
            )
        
        return customers
    
    def _validate_structure(self, wb: openpyxl.Workbook) -> List[ImportValidationError]:
        """Validate Excel file structure"""
        errors = []
        
        # Check required sheets
        for sheet_name in self.required_sheets:
            if sheet_name not in wb.sheetnames:
                errors.append(
                    ImportValidationError(
                        sheet_name='Workbook',
                        row_number=None,
                        field_name=None,
                        error_message=f"Required sheet '{sheet_name}' not found",
                        severity='error'
                    )
                )
        
        # Check optional sheets (info only)
        for sheet_name in self.optional_sheets:
            if sheet_name not in wb.sheetnames:
                errors.append(
                    ImportValidationError(
                        sheet_name='Workbook',
                        row_number=None,
                        field_name=None,
                        error_message=f"Optional sheet '{sheet_name}' not found - skipping",
                        severity='info'
                    )
                )
        
        return errors
    
    def _parse_customer_profiles(self, wb: openpyxl.Workbook) -> List[CustomerImportData]:
        """Parse Customer Profile sheet"""
        ws = wb['Customer Profile']
        customers = []
        
        # Find header row (usually row 3 after title)
        header_row = None
        for row_idx in range(1, 10):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and 'Account ID' in str(cell_value):
                header_row = row_idx
                break
        
        if not header_row:
            raise ValueError("Could not find header row in Customer Profile sheet")
        
        # Get column indices
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            if header:
                headers[str(header).strip()] = col_idx
        
        # Required columns
        required_cols = ['Account ID*', 'Account Name*', 'Industry*', 'Region*', 'Tier*', 'ARR ($)*', 'Starting ARR ($)*']
        for col in required_cols:
            if col not in headers:
                # Try without asterisk
                col_base = col.replace('*', '').strip()
                if col_base not in headers:
                    raise ValueError(f"Required column '{col}' not found in Customer Profile")
        
        # Parse rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            account_id = ws.cell(row_idx, headers.get('Account ID*', headers.get('Account ID'))).value
            
            if not account_id:
                break  # End of data
            
            customer = CustomerImportData(
                account_id=str(account_id).strip(),
                account_name=str(ws.cell(row_idx, headers.get('Account Name*', headers.get('Account Name'))).value or ''),
                industry=str(ws.cell(row_idx, headers.get('Industry*', headers.get('Industry'))).value or ''),
                region=str(ws.cell(row_idx, headers.get('Region*', headers.get('Region'))).value or ''),
                tier=str(ws.cell(row_idx, headers.get('Tier*', headers.get('Tier'))).value or ''),
                arr=float(ws.cell(row_idx, headers.get('ARR ($)*', headers.get('ARR ($)'))).value or 0),
                starting_arr=float(ws.cell(row_idx, headers.get('Starting ARR ($)*', headers.get('Starting ARR ($)'))).value or 0),
                tracked_kpis=[],
                kpi_coverage_pct=0.0,
                data_quality_level='unknown',
                kpi_history=[],
                contract_start=None,
                contract_end=None,
                contract_value=None,
                nps_csat_data=[],
                events=[],
                validation_errors=[],
                is_valid=True
            )
            
            customers.append(customer)
        
        return customers
    
    def _parse_kpi_coverage(self, wb: openpyxl.Workbook) -> Dict[str, Dict]:
        """Parse KPI Coverage Detail sheet"""
        ws = wb['KPI Coverage Detail']
        coverage_by_customer = {}
        
        # Find header row
        header_row = None
        for row_idx in range(1, 10):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and 'Account ID' in str(cell_value):
                header_row = row_idx
                break
        
        if not header_row:
            raise ValueError("Could not find header row in KPI Coverage Detail")
        
        # Get column indices
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row_idx, col_idx).value
            if header:
                headers[str(header).strip()] = col_idx
        
        # Parse rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            account_id = ws.cell(row_idx, headers.get('Account ID', 1)).value
            kpi_name = ws.cell(row_idx, headers.get('KPI Name', 3)).value
            
            if not account_id or not kpi_name:
                continue
            
            account_id = str(account_id).strip()
            kpi_name = str(kpi_name).strip()
            
            if account_id not in coverage_by_customer:
                coverage_by_customer[account_id] = {'kpis': [], 'coverage_pct': 0, 'quality_level': ''}
            
            coverage_by_customer[account_id]['kpis'].append(kpi_name)
        
        # Calculate coverage percentage for each customer
        for account_id, coverage in coverage_by_customer.items():
            kpi_count = len(coverage['kpis'])
            coverage['coverage_pct'] = round((kpi_count / 35) * 100, 1)
            
            if coverage['coverage_pct'] >= 40:
                coverage['quality_level'] = 'excellent'
            elif coverage['coverage_pct'] >= 30:
                coverage['quality_level'] = 'good'
            else:
                coverage['quality_level'] = 'low'
        
        return coverage_by_customer
    
    def _parse_kpi_history(self, wb: openpyxl.Workbook) -> Dict[str, List[Dict]]:
        """
        Parse KPI History sheet (if available)
        
        Returns:
            Dict mapping account_id to list of weekly KPI data
            {
                'DC001': [
                    {
                        'week': 1,
                        'date': '2024-01-01',
                        'kpis': {
                            'mtbf': (8500, 'hours'),  # (raw_value, unit)
                            'gpu_utilization_rate': (65, '%'),
                            ...
                        },
                        'events': []  # Optional events for this week
                    },
                    ...
                ]
            }
        """
        
        if 'KPI History' not in wb.sheetnames:
            print("ℹ️  KPI History sheet not found - skipping historical data")
            return {}
        
        ws = wb['KPI History']
        history_by_customer = {}
        
        # Find header row
        header_row = None
        for row_idx in range(1, 15):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and ('Account ID' in str(cell_value) or 'account_id' in str(cell_value).lower()):
                header_row = row_idx
                break
        
        if not header_row:
            print("⚠️  Could not find header row in KPI History sheet")
            return {}
        
        # Get column mapping
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            if header:
                headers[str(header).strip()] = col_idx
        
        print(f"Found columns in KPI History: {list(headers.keys())[:10]}...")
        
        # Determine structure: Week-based or Date-based columns
        week_columns = []
        date_columns = []
        
        for header_name in headers.keys():
            if 'Week' in header_name and any(char.isdigit() for char in header_name):
                week_columns.append(header_name)
            elif '/' in header_name or '-' in header_name:
                # Might be date column
                date_columns.append(header_name)
        
        if not week_columns and not date_columns:
            print("⚠️  Could not find week or date columns in KPI History")
            return {}
        
        print(f"Found {len(week_columns)} week columns" if week_columns else f"Found {len(date_columns)} date columns")
        
        # Get KPI unit mappings from normalization service
        kpi_units = {}
        for kpi_name in self.normalization_service.get_all_kpis():
            ref_range = self.normalization_service.get_reference_range(kpi_name)
            if ref_range:
                kpi_units[kpi_name] = ref_range.unit
        
        # Parse rows (each row is one KPI for one customer)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            account_id_cell = ws.cell(row_idx, headers.get('Account ID', headers.get('account_id', 1))).value
            kpi_name_cell = ws.cell(row_idx, headers.get('KPI Name', headers.get('kpi_name', headers.get('KPI', 2)))).value
            
            if not account_id_cell or not kpi_name_cell:
                continue  # Empty row
            
            account_id = str(account_id_cell).strip()
            kpi_name = str(kpi_name_cell).strip()
            
            # Get unit for this KPI
            unit = kpi_units.get(kpi_name, 'unknown')
            
            # Initialize customer history if not exists
            if account_id not in history_by_customer:
                history_by_customer[account_id] = {}
            
            # Parse week values
            if week_columns:
                # Week-based structure
                for week_col in week_columns:
                    # Extract week number from column name (e.g., "Week 1" -> 1)
                    week_num_str = ''.join(filter(str.isdigit, week_col))
                    if not week_num_str:
                        continue
                    
                    week_num = int(week_num_str)
                    raw_value = ws.cell(row_idx, headers[week_col]).value
                    
                    if raw_value is not None and raw_value != '':
                        # Initialize week if not exists
                        if week_num not in history_by_customer[account_id]:
                            history_by_customer[account_id][week_num] = {
                                'week': week_num,
                                'date': None,  # Will calculate later
                                'kpis': {},
                                'events': []
                            }
                        
                        # Store raw value with unit
                        try:
                            history_by_customer[account_id][week_num]['kpis'][kpi_name] = (float(raw_value), unit)
                        except (ValueError, TypeError):
                            # Handle non-numeric values
                            print(f"⚠️  Non-numeric value for {account_id} Week {week_num} {kpi_name}: {raw_value}")
                            continue
        
        # Convert dict to list and calculate dates
        result = {}
        for account_id, weeks_dict in history_by_customer.items():
            weeks_list = []
            for week_num in sorted(weeks_dict.keys()):
                week_data = weeks_dict[week_num]
                
                # Calculate date (assuming week 1 starts on a Monday)
                # In production, use customer's actual start date
                from datetime import datetime, timedelta
                start_date = datetime(2024, 1, 1)  # Default start
                week_date = start_date + timedelta(weeks=week_num - 1)
                week_data['date'] = week_date.strftime('%Y-%m-%d')
                
                weeks_list.append(week_data)
            
            result[account_id] = weeks_list
        
        print(f"✅ Parsed KPI history for {len(result)} customers")
        for account_id, weeks in list(result.items())[:3]:
            print(f"   {account_id}: {len(weeks)} weeks, avg {sum(len(w['kpis']) for w in weeks) / len(weeks):.1f} KPIs/week")
        
        return result
    
    def _parse_contracts(self, wb: openpyxl.Workbook) -> Dict[str, Dict]:
        """Parse Contracts & Revenue sheet"""
        return {}
    
    def _parse_nps_csat(self, wb: openpyxl.Workbook) -> Dict[str, List[Dict]]:
        """Parse NPS & CSAT sheet"""
        return {}
    
    def _parse_events(self, wb: openpyxl.Workbook) -> Dict[str, List[Dict]]:
        """
        Parse Events sheet
        
        Returns:
            Dict mapping account_id to list of events by week
            {
                'DC001': [
                    {
                        'week': 5,
                        'event_type': 'outage',
                        'description': '...',
                        'sentiment': 'negative',
                        'sentiment_value': -0.7,
                        'csm_action': {
                            'action_type': 'war_room',
                            'description': '...',
                            'response_time_hours': 2,
                            'cost_estimate': 15000
                        },
                        'outcome': '...',
                        'kpi_impacts': {...}
                    },
                    ...
                ]
            }
        """
        
        if 'Events' not in wb.sheetnames:
            print("ℹ️  Events sheet not found - skipping event data")
            return {}
        
        ws = wb['Events']
        events_by_customer = {}
        
        # Find header row
        header_row = None
        for row_idx in range(1, 15):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and ('Account ID' in str(cell_value) or 'Week' in str(cell_value)):
                header_row = row_idx
                break
        
        if not header_row:
            print("⚠️  Could not find header row in Events sheet")
            return {}
        
        # Get column mapping
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            if header:
                headers[str(header).strip()] = col_idx
        
        print(f"Found columns in Events: {list(headers.keys())}")
        
        # Required columns
        required_cols = ['Account ID', 'Week', 'Event Type', 'Description']
        missing_cols = [col for col in required_cols if col not in headers and col.lower() not in [h.lower() for h in headers.keys()]]
        
        if missing_cols:
            print(f"⚠️  Missing required columns in Events sheet: {missing_cols}")
            return {}
        
        # Column name variations
        account_id_col = headers.get('Account ID', headers.get('account_id', headers.get('AccountID', 1)))
        week_col = headers.get('Week', headers.get('week', headers.get('Week Number', 2)))
        event_type_col = headers.get('Event Type', headers.get('event_type', headers.get('Type', 3)))
        description_col = headers.get('Description', headers.get('description', 4))
        
        # Optional columns
        sentiment_col = headers.get('Sentiment', headers.get('sentiment'))
        csm_action_col = headers.get('CSM Action', headers.get('csm_action', headers.get('Action')))
        outcome_col = headers.get('Outcome', headers.get('outcome'))
        cost_col = headers.get('Cost', headers.get('cost', headers.get('CSM Cost')))
        
        # Parse events
        for row_idx in range(header_row + 1, ws.max_row + 1):
            account_id = ws.cell(row_idx, account_id_col).value
            week = ws.cell(row_idx, week_col).value
            event_type = ws.cell(row_idx, event_type_col).value
            description = ws.cell(row_idx, description_col).value
            
            if not account_id or not week or not event_type:
                continue  # Skip incomplete rows
            
            account_id = str(account_id).strip()
            week = int(week) if isinstance(week, (int, float)) else int(''.join(filter(str.isdigit, str(week))))
            event_type = str(event_type).strip().lower().replace(' ', '_')
            description = str(description).strip() if description else ''
            
            # Parse optional fields
            sentiment_str = 'neutral'
            sentiment_value = 0.0
            if sentiment_col:
                sentiment_raw = ws.cell(row_idx, sentiment_col).value
                if sentiment_raw:
                    sentiment_str = str(sentiment_raw).strip().lower().replace(' ', '_')
                    # Map sentiment to value
                    sentiment_map = {
                        'very_positive': 0.9,
                        'positive': 0.5,
                        'neutral': 0.0,
                        'negative': -0.5,
                        'very_negative': -0.9
                    }
                    sentiment_value = sentiment_map.get(sentiment_str, 0.0)
            
            # Parse CSM action
            csm_action = None
            if csm_action_col:
                csm_action_raw = ws.cell(row_idx, csm_action_col).value
                if csm_action_raw:
                    csm_cost = 0
                    if cost_col:
                        cost_raw = ws.cell(row_idx, cost_col).value
                        if cost_raw:
                            try:
                                csm_cost = float(cost_raw)
                            except:
                                csm_cost = 0
                    
                    csm_action = {
                        'action_type': str(csm_action_raw).strip().lower().replace(' ', '_'),
                        'description': str(csm_action_raw).strip(),
                        'response_time_hours': 4,  # Default
                        'cost_estimate': csm_cost
                    }
            
            # Parse outcome
            outcome = ''
            if outcome_col:
                outcome_raw = ws.cell(row_idx, outcome_col).value
                if outcome_raw:
                    outcome = str(outcome_raw).strip()
            
            # Create event dict
            event = {
                'week': week,
                'event_type': event_type,
                'description': description,
                'sentiment': sentiment_str,
                'sentiment_value': sentiment_value,
                'csm_action': csm_action,
                'outcome': outcome,
                'kpi_impacts': {}  # Could be enhanced later
            }
            
            # Add to customer's events
            if account_id not in events_by_customer:
                events_by_customer[account_id] = []
            
            events_by_customer[account_id].append(event)
        
        # Sort events by week
        for account_id in events_by_customer:
            events_by_customer[account_id].sort(key=lambda x: x['week'])
        
        print(f"✅ Parsed events for {len(events_by_customer)} customers")
        for account_id, events in list(events_by_customer.items())[:3]:
            print(f"   {account_id}: {len(events)} events")
        
        return events_by_customer
    
    def normalize_customer_kpis(
        self, 
        customer: CustomerImportData,
        raw_kpi_values: Dict[str, Tuple[float, str]]
    ) -> Dict[str, float]:
        """
        Normalize raw KPI values for a customer
        
        Args:
            customer: Customer data
            raw_kpi_values: Dict of {kpi_name: (raw_value, unit)}
                           e.g., {'mtbf': (8500, 'hours'), 'uptime': (99.7, '% uptime')}
        
        Returns:
            Dict of {kpi_name: normalized_score} (0-100 scale)
        """
        
        normalized_kpis = {}
        
        for kpi_name, (raw_value, unit) in raw_kpi_values.items():
            try:
                normalized_score, status = self.normalization_service.normalize(
                    kpi_name, raw_value, unit
                )
                normalized_kpis[kpi_name] = normalized_score
                
            except Exception as e:
                customer.validation_errors.append(
                    ImportValidationError(
                        sheet_name='KPI Values',
                        row_number=None,
                        field_name=kpi_name,
                        error_message=f"Normalization failed: {str(e)}",
                        severity='error'
                    )
                )
        
        return normalized_kpis
    
    def prepare_for_qdrant(
        self, 
        customer: CustomerImportData,
        normalized_kpis: Dict[str, float],
        week_number: int,
        date: str
    ) -> Dict:
        """
        Prepare customer week data for Qdrant upload
        
        Args:
            customer: Customer data
            normalized_kpis: Dict of normalized KPI scores (0-100)
            week_number: Week number (1-52)
            date: Date string (YYYY-MM-DD)
        
        Returns:
            Dict ready for WeekData and Qdrant upload
        """
        
        # Analyze KPI coverage
        health_input = self.sparse_kpi_handler.prepare_for_health_calculation(normalized_kpis)
        
        # Return structure compatible with WeekData
        return {
            'account_id': customer.account_id,
            'week_number': week_number,
            'date': date,
            'kpis': normalized_kpis,  # Sparse dict (10-15 KPIs)
            'available_kpis': health_input['available_kpis'],
            'missing_kpis': health_input['missing_kpis'],
            'data_quality': health_input['data_quality'],
            'data_quality_level': health_input['data_quality_level'],
            'missing_kpi_count': health_input['missing_kpi_count'],
            'pillars_covered': health_input['pillars_covered'],
            
            # These will be calculated by health calculator:
            # - pillars: Dict[str, float]
            # - health_score: float
            # - phase: JourneyPhase
        }


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == '__main__':
    import sys
    
    print("="*70)
    print("EXCEL IMPORT SERVICE - EXAMPLE")
    print("="*70)
    print()
    
    # Example: Import from Excel template
    excel_path = '/mnt/user-data/outputs/DataCenter_Enterprise_FINAL.xlsx'
    
    if not Path(excel_path).exists():
        print(f"❌ Excel file not found: {excel_path}")
        print("   Please provide path to Excel template")
        sys.exit(1)
    
    service = ExcelImportService()
    
    print(f"Importing from: {excel_path}")
    print()
    
    try:
        customers = service.import_from_file(excel_path)
        
        print(f"✅ Successfully imported {len(customers)} customers")
        print()
        
        for customer in customers:
            print("-" * 70)
            print(f"Customer: {customer.account_name} ({customer.account_id})")
            print(f"  Industry: {customer.industry}")
            print(f"  Region: {customer.region}")
            print(f"  Tier: {customer.tier}")
            print(f"  ARR: ${customer.arr:,.0f}")
            print(f"  KPIs Tracked: {len(customer.tracked_kpis)}")
            print(f"  Coverage: {customer.kpi_coverage_pct}%")
            print(f"  Quality: {customer.data_quality_level}")
            
            if customer.validation_errors:
                print(f"  Validation Issues: {len(customer.validation_errors)}")
                for error in customer.validation_errors:
                    symbol = "⚠️ " if error.severity == 'warning' else "ℹ️ " if error.severity == 'info' else "❌"
                    print(f"    {symbol} {error.error_message}")
            else:
                print(f"  ✅ Valid")
            
            print()
        
        # Example: Normalize KPIs for first customer
        if customers:
            customer = customers[0]
            print("-" * 70)
            print(f"Example: Normalize KPIs for {customer.account_name}")
            print()
            
            # Simulated raw KPI values
            raw_kpis = {
                'mtbf': (8500, 'hours'),
                'unplanned_downtime': (99.7, '% uptime'),
                'critical_incident_rate': (1.5, 'P1 incidents/month'),
                'gpu_utilization_rate': (65, '%'),
            }
            
            normalized = service.normalize_customer_kpis(customer, raw_kpis)
            
            print("Raw → Normalized:")
            for kpi_name, (raw_val, unit) in raw_kpis.items():
                norm_val = normalized.get(kpi_name, 0)
                print(f"  {kpi_name}:")
                print(f"    Raw: {raw_val} {unit}")
                print(f"    Normalized: {norm_val}/100")
            
            print()
            
            # Prepare for Qdrant
            qdrant_data = service.prepare_for_qdrant(customer, normalized, week_number=1, date='2024-01-01')
            
            print("Qdrant Payload Preview:")
            print(f"  Account: {qdrant_data['account_id']}")
            print(f"  Week: {qdrant_data['week_number']}")
            print(f"  KPIs Available: {len(qdrant_data['kpis'])}")
            print(f"  Data Quality: {qdrant_data['data_quality']:.3f}")
            print(f"  Missing KPIs: {qdrant_data['missing_kpi_count']}")
            
    except Exception as e:
        print(f"❌ Import failed: {str(e)}")
        import traceback
        traceback.print_exc()
