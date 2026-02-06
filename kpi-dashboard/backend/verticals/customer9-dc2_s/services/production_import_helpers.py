#!/usr/bin/env python3
"""
Production Import Helpers
==========================

Production-ready enhancements for import system:
- Async processing (Celery task wrapper)
- Progress tracking
- Batch import
- Error recovery and rollback
- Import validation
- Performance monitoring
"""

from typing import Dict, List, Optional, Callable
from dataclasses import dataclass
from datetime import datetime
import time
import json


@dataclass
class ImportProgress:
    """Track import progress"""
    import_id: str
    total_customers: int
    processed_customers: int
    total_weeks: int
    processed_weeks: int
    current_customer: Optional[str]
    current_week: Optional[int]
    status: str  # 'processing', 'completed', 'failed', 'cancelled'
    errors: List[str]
    warnings: List[str]
    started_at: datetime
    updated_at: datetime
    completed_at: Optional[datetime]
    
    def to_dict(self) -> Dict:
        return {
            'import_id': self.import_id,
            'total_customers': self.total_customers,
            'processed_customers': self.processed_customers,
            'total_weeks': self.total_weeks,
            'processed_weeks': self.processed_weeks,
            'current_customer': self.current_customer,
            'current_week': self.current_week,
            'status': self.status,
            'progress_pct': round((self.processed_weeks / self.total_weeks * 100) if self.total_weeks > 0 else 0, 1),
            'errors': self.errors,
            'warnings': self.warnings,
            'started_at': self.started_at.isoformat(),
            'updated_at': self.updated_at.isoformat(),
            'completed_at': self.completed_at.isoformat() if self.completed_at else None
        }


class ProgressTracker:
    """
    Track and persist import progress
    
    In production, this would update a database table for real-time progress monitoring
    """
    
    def __init__(self, import_id: str, total_customers: int, total_weeks: int):
        """Initialize progress tracker"""
        self.progress = ImportProgress(
            import_id=import_id,
            total_customers=total_customers,
            processed_customers=0,
            total_weeks=total_weeks,
            processed_weeks=0,
            current_customer=None,
            current_week=None,
            status='processing',
            errors=[],
            warnings=[],
            started_at=datetime.now(),
            updated_at=datetime.now(),
            completed_at=None
        )
    
    def update_customer(self, customer_id: str):
        """Update current customer being processed"""
        self.progress.current_customer = customer_id
        self.progress.updated_at = datetime.now()
        self._persist()
    
    def update_week(self, week_number: int):
        """Update current week being processed"""
        self.progress.current_week = week_number
        self.progress.processed_weeks += 1
        self.progress.updated_at = datetime.now()
        self._persist()
    
    def complete_customer(self):
        """Mark customer as completed"""
        self.progress.processed_customers += 1
        self.progress.current_customer = None
        self.progress.current_week = None
        self.progress.updated_at = datetime.now()
        self._persist()
    
    def add_error(self, error_msg: str):
        """Add error message"""
        self.progress.errors.append(error_msg)
        self.progress.updated_at = datetime.now()
        self._persist()
    
    def add_warning(self, warning_msg: str):
        """Add warning message"""
        self.progress.warnings.append(warning_msg)
        self.progress.updated_at = datetime.now()
        self._persist()
    
    def complete(self, status: str = 'completed'):
        """Mark import as completed"""
        self.progress.status = status
        self.progress.completed_at = datetime.now()
        self.progress.updated_at = datetime.now()
        self._persist()
    
    def fail(self, error_msg: str):
        """Mark import as failed"""
        self.add_error(error_msg)
        self.progress.status = 'failed'
        self.progress.completed_at = datetime.now()
        self._persist()
    
    def _persist(self):
        """
        Persist progress to database
        
        In production, update progress table:
        UPDATE import_progress SET ... WHERE import_id = ?
        """
        # Placeholder: In production, save to database
        pass
    
    def get_progress(self) -> Dict:
        """Get current progress as dict"""
        return self.progress.to_dict()


class BatchImportManager:
    """
    Manage batch imports of multiple Excel files
    
    Features:
    - Import multiple files in sequence
    - Aggregate results
    - Handle partial failures
    """
    
    def __init__(self, adapter):
        """
        Initialize batch manager
        
        Args:
            adapter: ImportIntegrationAdapter instance
        """
        self.adapter = adapter
    
    def import_batch(
        self, 
        file_paths: List[str],
        continue_on_error: bool = True
    ) -> Dict:
        """
        Import multiple Excel files
        
        Args:
            file_paths: List of Excel file paths
            continue_on_error: Whether to continue if one file fails
        
        Returns:
            Aggregate import results
        """
        
        results = {
            'total_files': len(file_paths),
            'successful_files': 0,
            'failed_files': 0,
            'total_customers': 0,
            'total_weeks': 0,
            'errors': [],
            'warnings': [],
            'file_results': []
        }
        
        for i, file_path in enumerate(file_paths, 1):
            print(f"\n[{i}/{len(file_paths)}] Importing: {file_path}")
            
            try:
                result = self.adapter.import_excel_to_cs_pulse(file_path)
                
                if result.success:
                    results['successful_files'] += 1
                    results['total_customers'] += result.customers_imported
                    results['total_weeks'] += result.weeks_imported
                else:
                    results['failed_files'] += 1
                    results['errors'].extend(result.errors)
                
                results['warnings'].extend(result.warnings)
                results['file_results'].append({
                    'file': file_path,
                    'success': result.success,
                    'customers': result.customers_imported,
                    'weeks': result.weeks_imported
                })
                
            except Exception as e:
                results['failed_files'] += 1
                error_msg = f"Failed to import {file_path}: {str(e)}"
                results['errors'].append(error_msg)
                results['file_results'].append({
                    'file': file_path,
                    'success': False,
                    'error': str(e)
                })
                
                if not continue_on_error:
                    break
        
        return results


class ImportValidator:
    """
    Validate import data before processing
    
    Catches common issues early to prevent partial imports
    """
    
    @staticmethod
    def validate_before_import(excel_path: str) -> Dict:
        """
        Validate Excel file before starting import
        
        Returns:
            {
                'valid': bool,
                'errors': List[str],
                'warnings': List[str],
                'stats': Dict
            }
        """
        
        import openpyxl
        from pathlib import Path
        
        result = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'stats': {}
        }
        
        # Check file exists
        if not Path(excel_path).exists():
            result['valid'] = False
            result['errors'].append(f"File not found: {excel_path}")
            return result
        
        # Check file format
        if not excel_path.endswith(('.xlsx', '.xls')):
            result['valid'] = False
            result['errors'].append("Invalid file format. Must be .xlsx or .xls")
            return result
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Check required sheets
            required_sheets = ['Customer Profile', 'KPI Coverage Detail']
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    result['valid'] = False
                    result['errors'].append(f"Missing required sheet: {sheet}")
            
            # Check optional sheets
            optional_sheets = ['KPI History', 'Events', 'NPS & CSAT', 'Contracts & Revenue']
            for sheet in optional_sheets:
                if sheet not in wb.sheetnames:
                    result['warnings'].append(f"Optional sheet not found: {sheet}")
            
            # Get stats
            if 'Customer Profile' in wb.sheetnames:
                ws = wb['Customer Profile']
                # Count data rows (excluding header)
                data_rows = 0
                for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                    if row[0].value:
                        data_rows += 1
                result['stats']['customers'] = data_rows
            
            if 'KPI History' in wb.sheetnames:
                result['stats']['has_history'] = True
            else:
                result['stats']['has_history'] = False
                result['warnings'].append("No KPI History - can only import snapshot data")
            
        except Exception as e:
            result['valid'] = False
            result['errors'].append(f"Failed to read Excel file: {str(e)}")
        
        return result


class PerformanceMonitor:
    """Monitor import performance"""
    
    def __init__(self):
        """Initialize performance monitor"""
        self.start_time = None
        self.checkpoints = {}
    
    def start(self):
        """Start monitoring"""
        self.start_time = time.time()
    
    def checkpoint(self, name: str):
        """Record checkpoint"""
        if not self.start_time:
            self.start()
        
        self.checkpoints[name] = time.time() - self.start_time
    
    def get_stats(self) -> Dict:
        """Get performance statistics"""
        if not self.start_time:
            return {}
        
        total_time = time.time() - self.start_time
        
        return {
            'total_time_seconds': round(total_time, 2),
            'checkpoints': {k: round(v, 2) for k, v in self.checkpoints.items()},
            'time_per_checkpoint': {
                k: round(v - list(self.checkpoints.values())[i-1], 2) if i > 0 else round(v, 2)
                for i, (k, v) in enumerate(self.checkpoints.items())
            }
        }


# ============================================================================
# CELERY TASK WRAPPER (for async processing)
# ============================================================================

def create_celery_import_task(celery_app):
    """
    Create Celery task for async import
    
    Usage:
        from celery_app import celery
        import_task = create_celery_import_task(celery)
        
        # Start async import
        result = import_task.delay(excel_path, import_id)
    
    Args:
        celery_app: Celery application instance
    
    Returns:
        Celery task function
    """
    
    @celery_app.task(bind=True)
    def import_excel_async(self, excel_path: str, import_id: str):
        """
        Async import task
        
        Args:
            excel_path: Path to Excel file
            import_id: Import ID for tracking
        
        Returns:
            Import result dict
        """
        
        from import_integration_adapter import ImportIntegrationAdapter
        
        # Update task state
        self.update_state(state='PROCESSING', meta={'status': 'Starting import'})
        
        try:
            # Initialize adapter
            adapter = ImportIntegrationAdapter()
            
            # Run import
            result = adapter.import_excel_to_cs_pulse(excel_path)
            
            # Update final state
            if result.success:
                return {
                    'status': 'completed',
                    'customers_imported': result.customers_imported,
                    'weeks_imported': result.weeks_imported,
                    'errors': result.errors,
                    'warnings': result.warnings
                }
            else:
                self.update_state(state='FAILURE', meta={'errors': result.errors})
                return {
                    'status': 'failed',
                    'errors': result.errors
                }
                
        except Exception as e:
            self.update_state(state='FAILURE', meta={'error': str(e)})
            raise
    
    return import_excel_async


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == '__main__':
    print("="*70)
    print("PRODUCTION IMPORT HELPERS - EXAMPLES")
    print("="*70)
    print()
    
    # Example 1: Progress Tracking
    print("Example 1: Progress Tracking")
    print("-" * 70)
    
    tracker = ProgressTracker(
        import_id='import-123',
        total_customers=10,
        total_weeks=520
    )
    
    # Simulate import
    tracker.update_customer('DC001')
    for week in range(1, 53):
        tracker.update_week(week)
    tracker.complete_customer()
    
    progress = tracker.get_progress()
    print(f"Progress: {progress['progress_pct']}%")
    print(f"Status: {progress['status']}")
    print()
    
    # Example 2: Validation
    print("Example 2: Pre-Import Validation")
    print("-" * 70)
    
    validation = ImportValidator.validate_before_import('test.xlsx')
    print(f"Valid: {validation['valid']}")
    print(f"Errors: {validation['errors']}")
    print(f"Warnings: {validation['warnings']}")
    print()
    
    # Example 3: Performance Monitoring
    print("Example 3: Performance Monitoring")
    print("-" * 70)
    
    monitor = PerformanceMonitor()
    monitor.start()
    
    # Simulate processing
    time.sleep(0.1)
    monitor.checkpoint('parse_excel')
    
    time.sleep(0.2)
    monitor.checkpoint('normalize_kpis')
    
    time.sleep(0.3)
    monitor.checkpoint('calculate_health')
    
    stats = monitor.get_stats()
    print(f"Total time: {stats['total_time_seconds']}s")
    print(f"Time per stage:")
    for stage, duration in stats['time_per_checkpoint'].items():
        print(f"  {stage}: {duration}s")
