#!/usr/bin/env python3
"""
Generate CS Pulse DataCenter onboarding Excel template.

Creates a multi-tab workbook with:
  - README (corrected onboarding guide)
  - KPI Reference (38 KPIs with health ranges)
  - Weights (L1/L2 defaults + customization)
  - 11 CSV data-entry tabs (4 regular + 7 context graph)

All data is read programmatically from source-of-truth files:
  - config/csv_schemas.json
  - verticals/dc2_s/kpi_definitions.py
  - config/health_thresholds.json
  - verticals/dc2_s/vertical_config.py

Usage:
    python scripts/generate_onboarding_template.py [--output PATH]
"""

import argparse
import json
import os
import sys
from pathlib import Path

# Set up imports from backend root
BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling constants (match export_api.py patterns)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
REQUIRED_FONT = Font(bold=True, color="1F4E79", size=10)
OPTIONAL_FONT = Font(color="666666", size=10)
TITLE_FONT = Font(bold=True, size=13)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
NOTE_FONT = Font(italic=True, color="888888", size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center")

PILLAR_FILLS = {
    "P1": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    "P2": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "P3": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    "P4": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    "P5": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
}
PILLAR_FONT = Font(bold=True, color="FFFFFF", size=11)

HEALTHY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RISK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

INFERRED_CSVS = {"signal_edges.csv"}

SHEET_NAME_MAP = {
    "accounts.csv": "accounts",
    "kpi_measurements.csv": "kpi_measurements",
    "enhanced_qualitative_signals.csv": "enhanced_signals",
    "products.csv": "products",
    "stakeholders.csv": "CG-stakeholders",
    "engagement_events.csv": "CG-engagement_events",
    "account_business_profiles.csv": "CG-account_biz_profiles",
    "decisions.csv": "CG-decisions",
    "outcomes.csv": "CG-outcomes",
    "signal_edges.csv": "CG-signal_edges INFER",
    "industry_benchmarks.csv": "CG-industry_benchmarks",
}

# Regular CSV tab order (4 customer-provided files)
REGULAR_CSV_ORDER = [
    "accounts.csv", "kpi_measurements.csv", "enhanced_qualitative_signals.csv",
    "products.csv",
]
# Context graph CSV tab order (7 files: 4 customer-provided + 3 auto-generated)
CG_CSV_ORDER = [
    "stakeholders.csv", "engagement_events.csv", "account_business_profiles.csv",
    "decisions.csv", "outcomes.csv", "signal_edges.csv",
    "industry_benchmarks.csv",
]

# Sample values for example rows (all dates YYYY-MM-DD)
SAMPLE_VALUES = {
    # -- accounts --
    "account_id": 400001,
    "customer_id": 400,
    "account_name": "Nexus Cloud - Production",
    "industry": "Data Center Infrastructure",
    "region": "us-west-2",
    "vertical": "dc2_s",
    "tier": "Enterprise",
    "arr": 1250000,
    "revenue": 1250000,
    "mrr": 104167,
    "contract_start": "2025-01-15",
    "contract_end": "2026-12-31",
    "renewal_date": "2026-12-31",
    "csm_name": "Jordan Lee",
    "csm_email": "jordan.lee@example.com",
    "account_status": "active",
    "uuid": "dc2s_acct_nexus_prod",
    # -- kpi_measurements --
    "kpi_code": "P1-KPI1",
    "measured_at": "2026-03-01",
    "value": 12.5,
    "kpi_name": "Time-to-First-Workload",
    "pillar": "P1",
    "target": 14,
    "weight": 0.20,
    "unit": "days",
    "status": "healthy",
    # -- enhanced_qualitative_signals --
    "signal_id": "SIG-001",
    "signal_date": "2026-03-01",
    "content": "Deployment proceeding ahead of schedule with all racks commissioned.",
    "sentiment": "positive",
    "signal_type": "deployment_update",
    "sentiment_score": 0.82,
    "stakeholder_level": "executive",
    "stakeholder_title": "VP Infrastructure",
    "keywords": "deployment,on-schedule,commissioned",
    "source_platform": "salesforce",
    # -- products --
    "product_name": "DC2_S Platform",
    "product_category": "Infrastructure",
    "quantity": 1,
    "unit_price": 1250000,
    "deployment_date": "2025-01-15",
    # -- account_business_profiles (CSM/champion columns, consolidated from profiles) --
    "assigned_csm": "Jordan Lee",
    "csm_manager": "Sam Rivera",
    "executive_sponsor": "Jane Chen, CTO",
    "primary_champion_name": "Chris Martinez",
    "primary_champion_title": "VP Infrastructure",
    "primary_champion_email": "chris.martinez@nexuscloud.com",
    "primary_champion_engagement_score": 85,
    "last_updated": "2026-03-01",
    # -- stakeholders (CG) --
    "stakeholder_name": "Chris Martinez",
    "title": "VP Infrastructure",
    "role": "champion",
    "influence_score": 85,
    "engagement_frequency": "weekly",
    "department": "Engineering",
    "is_active": True,
    "source_platform": "salesforce",
    "first_observed_at": "2025-01-20",
    # -- engagement_events (CG) --
    "event_date": "2026-03-01",
    "event_type": "qbr",
    "description": "Quarterly business review with executive team",
    "sentiment_shift": 0.1,
    "channel": "video_call",
    "duration_minutes": 60,
    "outcome": "Positive — expansion discussed",
    # -- account_business_profiles (CG) --
    "employee_count": 2500,
    "fiscal_year_end": "December",
    "tech_stack": "NVIDIA DGX, Kubernetes, PyTorch",
    "cloud_provider": "On-premise + AWS hybrid",
    "competitive_landscape": "Dell, HPE",
    "strategic_initiatives": "AI infrastructure expansion",
    "budget_cycle": "Q4 planning",
    "profile_date": "2026-03-01",
    # -- decisions (CG) --
    "decision_date": "2026-02-15",
    "decision_maker_role": "VP Infrastructure",
    "chosen_option": "Expand GPU cluster by 40%",
    "decision_id": "DEC-001",
    "options_considered": "Expand 40% | Expand 20% | Hold",
    "revenue_impact": 480000,
    "risk_level": "medium",
    "evidence_refs": "SIG-001,SIG-003",
    "outcome_description": "Approved Phase 2 expansion",
    "confidence": 0.85,
    # -- outcomes (CG) --
    "outcome_date": "2026-03-01",
    "outcome_type": "expansion",
    "revenue_value": 480000,
    "outcome_id": "OUT-001",
    "evidence": "GPU utilization exceeded 80% for 3 months",
    "related_decision_id": "DEC-001",
    # -- signal_edges (CG) — also carries SOURCED_FROM decision-evidence columns --
    "from_signal_ref": "SIG-001",
    "to_signal_ref": "DEC-001",
    "edge_type": "CAUSED_BY",
    "lag_days": 30,
    "decision_ref": "DEC-001",
    "evidence_type": "kpi_trend",
    "evidence_description": "GPU utilization rose from 55% to 78% over 3 months",
    "kpi_value": 78.0,
    "signal_ref": "SIG-001",
    "timestamp": "2026-02-15",
    # -- industry_benchmarks (CG) --
    "benchmark_source": "Gartner DC Report 2025",
    "industry_p50": 72.5,
    "industry_p25": 55.0,
    "industry_p75": 88.0,
    "account_percentile": 65,
    "sample_size": 450,
    "benchmark_date": "2025-12-01",
    # -- enhanced_qualitative_signals (CG) --
    "signal_ref": "SIG-001",
    "causal_chain_ref": "SIG-001->DEC-001->OUT-001",
}


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_csv_schemas() -> dict:
    path = BACKEND_DIR / "config" / "csv_schemas.json"
    with open(path) as f:
        return json.load(f)


def load_kpi_definitions():
    from utils.vertical_registry import get_kpis, get_pillars
    return get_kpis('dc2_s'), get_pillars('dc2_s')


def load_health_thresholds() -> dict:
    path = BACKEND_DIR / "config" / "health_thresholds.json"
    with open(path) as f:
        return json.load(f)


def load_playbook_config() -> dict:
    from _ask_ai_helpers import _get_playbook_config
    PLAYBOOK_CONFIG, _ = _get_playbook_config('dc2_s')
    return PLAYBOOK_CONFIG


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def auto_size_columns(ws, max_width: int = 50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def get_sample(col_name: str) -> object:
    if col_name in SAMPLE_VALUES:
        return SAMPLE_VALUES[col_name]
    low = col_name.lower()
    if "date" in low or "at" == low[-2:]:
        return "2026-03-01"
    if "_id" in low:
        return 10001
    if "score" in low or "weight" in low:
        return 75
    if "name" in low:
        return "Sample Name"
    return ""


# ---------------------------------------------------------------------------
# Tab 1: README
# ---------------------------------------------------------------------------

def create_readme_tab(wb, kpi_defs, pillars, schemas, thresholds, playbooks):
    ws = wb.create_sheet("README")
    ws.sheet_properties.tabColor = "4472C4"

    rows = []

    def add(text, font=None):
        rows.append((text, font))

    add("CS Pulse DataCenter — Onboarding Template", TITLE_FONT)
    add("")
    add("PLATFORM OVERVIEW", SECTION_FONT)
    add(f"  38 KPIs across 5 pillars, 11 CSV file types (4 regular + 7 context graph)")
    add("")

    # Pillars
    for pid in ["P1", "P2", "P3", "P4", "P5"]:
        p = pillars[pid]
        kpi_count = sum(1 for k, v in kpi_defs.items() if v.get("pillar") == pid)
        add(f"  {pid}: {p['name']} — {int(p['weight_l2']*100)}% weight, {kpi_count} KPIs")

    add("")
    add("DATA FORMAT RULES", SECTION_FONT)
    add("  All dates: YYYY-MM-DD (e.g. 2026-03-01)")
    add("  KPI codes: P-format only (P1-KPI1 through P5-KPI8). NEVER use old letter-format (AI/CH/DV/EX/OS).")
    add("  Currency: USD, no currency symbols in CSV (e.g. 1250000 not $1,250,000)")
    add("")

    add("HEALTH SCORE BANDS", SECTION_FONT)
    th = thresholds.get("thresholds", {})
    healthy_min = th.get("healthy", {}).get("min", 70)
    atrisk_min = th.get("at_risk", {}).get("min", 50)
    add(f"  Healthy: {healthy_min}–100  |  At-Risk: {atrisk_min}–{healthy_min-1}  |  Critical: 0–{atrisk_min-1}")
    add("  See 'KPI Reference' tab for per-KPI health ranges.")
    add("")

    add("CSV FILES (11 total)", SECTION_FONT)
    add("  REGULAR MODEL (4 CSVs) — Required for all customers:")
    add("    1. accounts.csv                    — Account master data (REQUIRED)")
    add("    2. kpi_measurements.csv            — KPI metric values (REQUIRED)")
    add("    3. enhanced_qualitative_signals.csv — Customer signals/feedback (optional graph columns)")
    add("    4. products.csv                    — Product usage")
    add("")
    add("  CONTEXT GRAPH MODEL (7 additional CSVs) — When context graph feature is enabled:")
    add("    5. stakeholders.csv              — Key contacts & champions")
    add("    6. engagement_events.csv          — Meetings, calls, QBRs")
    add("    7. account_business_profiles.csv  — Business context + CSM/champion data")
    add("    8. decisions.csv                  — Strategic decisions")
    add("    9. outcomes.csv                   — Business outcomes (expansion, churn_averted)")
    add("   10. signal_edges.csv               [INFERRED] Causal edges + SOURCED_FROM evidence")
    add("   11. industry_benchmarks.csv        — External benchmarks")
    add("")
    add("  [INFERRED] = Platform derives these from node data. You may provide them, or leave blank.")
    add("")

    add("WEIGHT MANAGEMENT (L1-L2-L3-L4 Rollup)", SECTION_FONT)
    add("  L1: KPI scores (value-to-score via 4-band interpolation per KPI ranges)")
    add("  L2: Pillar scores = weighted avg of L1 KPIs (using kpi weight_l1)")
    add("  L3: Account health = weighted avg of L2 pillars (using pillar weight_l2)")
    add("  L4: Customer health = revenue-weighted avg of L3 across accounts")
    add("")
    add("  Weight lifecycle:")
    add("    1. Defaults from 'Weights' tab → passed to /complete endpoint")
    add("    2. Wizard C auto-calibrates from your data during /process-data")
    add("    3. Calibrated weights stored in CustomerConfig DB")
    add("    4. Score calculator reads from DB (falls back to defaults)")
    add("  See 'Weights' tab to customize starting L1 and L2 weights.")
    add("")

    add("PLAYBOOKS (PB-01 through PB-06)", SECTION_FONT)
    for pb_id in sorted(playbooks.keys()):
        pb = playbooks[pb_id]
        triggers = []
        for kpi, cond in pb.get("trigger_conditions", {}).items():
            triggers.append(f"{kpi} {cond['operator']} {cond['value']}")
        trigger_str = " AND ".join(triggers)
        add(f"  {pb_id}: {pb['name']} — Triggers: {trigger_str}")
        add(f"         Impact: {pb.get('estimated_impact', 'N/A')}  |  Duration: {pb.get('estimated_duration_display', 'N/A')}")
    add("")

    add("FLEXIBLE ONBOARDING (Start Small, Expand Later)", SECTION_FONT)
    add("  You do NOT need all 5 pillars or 38 KPIs on day 1. The /complete endpoint accepts:")
    add("    enabled_pillars: ['P1', 'P3']        — Pick a subset of pillars; weights auto-redistribute")
    add("    enabled_kpis: ['P1-KPI1', 'P3-KPI2'] — Pick exact KPIs (overrides enabled_pillars)")
    add("    weights: {'P1': 0.6, 'P3': 0.4}      — Custom pillar weights for your subset (sum to 1.0)")
    add("    kpi_weights: {'P1': {'P1-KPI1': 0.5, 'P1-KPI2': 0.5}} — Custom L1 KPI weights")
    add("  If neither is provided, all 38 KPIs across 5 pillars are enabled by default.")
    add("  You can expand your pillar/KPI set later by updating CustomerConfig.")
    add("")

    add("4-WEEK ONBOARDING SEQUENCE", SECTION_FONT)
    add("  Week 1: accounts.csv, products.csv (foundational data)")
    add("  Week 2: kpi_measurements.csv, enhanced_qualitative_signals.csv (health scores activate)")
    add("  Week 3: Context graph CSVs — stakeholders, engagement_events, account_business_profiles, decisions, outcomes")
    add("  Week 4: industry_benchmarks; run /process-data; verify health scores")
    add("")

    add("ACCOUNT NAMING CONVENTION (from Kacme template)", SECTION_FONT)
    add("  Production, Staging, Development, QA, UAT, DR, Sandbox, Integration, Performance, Lab")
    add("  For >10 accounts: Account-11, Account-12, ... Account-N")
    add("")

    add("API ENDPOINTS FOR ONBOARDING", SECTION_FONT)
    add("  POST /api/onboarding/complete    — Create customer, accounts, config")
    add("       Accepts: enabled_pillars, enabled_kpis, weights (L2), kpi_weights (L1)")
    add("  POST /api/onboarding/upload      — Upload CSV files (4 regular + 7 context graph)")
    add("  POST /api/onboarding/process-data — Run data ingestion + Wizard A/B/C")

    # Write rows
    for i, (text, font) in enumerate(rows, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = WRAP
        if font:
            cell.font = font

    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Tab 2: KPI Reference
# ---------------------------------------------------------------------------

def create_kpi_reference_tab(wb, kpi_defs, pillars):
    ws = wb.create_sheet("KPI Reference")
    ws.sheet_properties.tabColor = "70AD47"

    headers = [
        "Pillar", "KPI Code", "Name", "Unit", "Direction",
        "Weight L1", "Healthy Min", "Healthy Max",
        "Risk Min", "Risk Max", "Critical Min", "Critical Max", "Target",
    ]
    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER

    # Color the range header cells
    for c in (7, 8):
        ws.cell(row=1, column=c).fill = HEALTHY_FILL
    for c in (9, 10):
        ws.cell(row=1, column=c).fill = RISK_FILL
    for c in (11, 12):
        ws.cell(row=1, column=c).fill = CRITICAL_FILL

    row = 2
    for pid in ["P1", "P2", "P3", "P4", "P5"]:
        p = pillars[pid]
        # Pillar header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = ws.cell(row=row, column=1,
                       value=f"{pid}: {p['name']} (L2 weight: {int(p['weight_l2']*100)}%)")
        cell.fill = PILLAR_FILLS[pid]
        cell.font = PILLAR_FONT
        cell.alignment = WRAP_CENTER
        row += 1

        # KPI rows for this pillar
        pillar_kpis = sorted(
            [(code, defn) for code, defn in kpi_defs.items() if defn.get("pillar") == pid],
            key=lambda x: x[0],
        )
        for code, defn in pillar_kpis:
            ranges = defn.get("ranges", {})
            target_info = defn.get("target", {})
            target_str = f"{target_info.get('operator', '')} {target_info.get('value', '')}"
            direction = "Higher is better" if defn.get("higher_is_better") else "Lower is better"

            vals = [
                pid, code, defn.get("name", ""), defn.get("unit", ""),
                direction, defn.get("weight_l1", 0),
                ranges.get("healthy", {}).get("min"),
                ranges.get("healthy", {}).get("max"),
                ranges.get("risk", {}).get("min"),
                ranges.get("risk", {}).get("max"),
                ranges.get("critical", {}).get("min"),
                ranges.get("critical", {}).get("max"),
                target_str.strip(),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                # Color range cells
                if c in (7, 8):
                    cell.fill = HEALTHY_FILL
                elif c in (9, 10):
                    cell.fill = RISK_FILL
                elif c in (11, 12):
                    cell.fill = CRITICAL_FILL
            row += 1

    auto_size_columns(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Tab 3: Weights
# ---------------------------------------------------------------------------

def create_weights_tab(wb, kpi_defs, pillars):
    ws = wb.create_sheet("Weights")
    ws.sheet_properties.tabColor = "ED7D31"

    row = 1

    # --- Section A: L2 Pillar Weights ---
    cell = ws.cell(row=row, column=1, value="SECTION A: L2 PILLAR WEIGHTS")
    cell.font = SECTION_FONT
    row += 1

    l2_headers = ["Pillar Code", "Pillar Name", "Default Weight", "Your Weight"]
    for c, h in enumerate(l2_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for pid in ["P1", "P2", "P3", "P4", "P5"]:
        p = pillars[pid]
        ws.cell(row=row, column=1, value=pid)
        ws.cell(row=row, column=2, value=p["name"])
        ws.cell(row=row, column=3, value=p["weight_l2"])
        ws.cell(row=row, column=4, value="").font = Font(color="1F4E79")
        row += 1

    note = ws.cell(row=row, column=1,
                   value="Note: 'Your Weight' values are passed to /complete endpoint via 'weights' param. Sum must equal 1.0. Leave blank to use defaults.")
    note.font = NOTE_FONT
    row += 1
    flex_note = ws.cell(row=row, column=1,
                        value="FLEXIBILITY: You do NOT need all 5 pillars on day 1. Pass 'enabled_pillars': ['P1','P3'] to start with a subset. Weights auto-redistribute.")
    flex_note.font = Font(color="1F4E79", italic=True)
    row += 2

    # --- Section B: L1 KPI Weights ---
    cell = ws.cell(row=row, column=1, value="SECTION B: L1 KPI WEIGHTS (per-KPI within pillar)")
    cell.font = SECTION_FONT
    row += 1

    l1_headers = ["Pillar", "KPI Code", "KPI Name", "Default Weight L1", "Your Weight"]
    for c, h in enumerate(l1_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for pid in ["P1", "P2", "P3", "P4", "P5"]:
        pillar_kpis = sorted(
            [(code, defn) for code, defn in kpi_defs.items() if defn.get("pillar") == pid],
            key=lambda x: x[0],
        )
        for code, defn in pillar_kpis:
            ws.cell(row=row, column=1, value=pid)
            ws.cell(row=row, column=2, value=code)
            ws.cell(row=row, column=3, value=defn.get("name", ""))
            ws.cell(row=row, column=4, value=defn.get("weight_l1", 0))
            ws.cell(row=row, column=5, value="").font = Font(color="1F4E79")
            row += 1

    note = ws.cell(row=row, column=1,
                   value="Note: L1 weights within each pillar should sum to 1.0. Accepted by /complete endpoint via 'kpi_weights' param. Wizard C will auto-recalibrate.")
    note.font = NOTE_FONT
    row += 1
    flex_note = ws.cell(row=row, column=1,
                        value="FLEXIBILITY: Pass 'enabled_kpis': ['P1-KPI1','P1-KPI3','P3-KPI2'] to pick exactly which KPIs you want. You don't need all 38 on day 1.")
    flex_note.font = Font(color="1F4E79", italic=True)
    row += 2

    # --- Section C: Weight Lifecycle ---
    cell = ws.cell(row=row, column=1, value="SECTION C: WEIGHT LIFECYCLE")
    cell.font = SECTION_FONT
    row += 1
    lifecycle = [
        "1. Initial weights (this tab) → loaded at POST /complete (L2 pillar weights + L1 KPI weights)",
        "2. Data loaded via POST /upload + POST /process-data",
        "3. Wizard A generates journeys from KPI data",
        "4. Wizard B analyzes patterns and early warnings",
        "5. Wizard C auto-calibrates L1+L2 weights from your actual data → saves to CustomerConfig DB",
        "6. Score calculator reads calibrated weights from DB (falls back to defaults from kpi_definitions.py)",
        "",
        "bootstrap_weights_config.json is generated per-customer after Wizard C runs, at:",
        "  verticals/customer{id}-dc2_s/journey/config/bootstrap_weights_config.json",
    ]
    for line in lifecycle:
        ws.cell(row=row, column=1, value=line).alignment = WRAP
        row += 1

    auto_size_columns(ws)
    ws.freeze_panes = "A1"


# ---------------------------------------------------------------------------
# Tabs 4-14: CSV data-entry tabs
# ---------------------------------------------------------------------------

def create_csv_data_tab(wb, csv_filename, csv_schema, is_context_graph=False, is_inferred=False):
    sheet_name = SHEET_NAME_MAP.get(csv_filename, csv_filename.replace(".csv", "")[:31])
    ws = wb.create_sheet(sheet_name)

    if is_inferred:
        ws.sheet_properties.tabColor = "FFC7CE"
    elif is_context_graph:
        ws.sheet_properties.tabColor = "00B0F0"

    # Row 1: Title
    prefix = ""
    if is_inferred:
        prefix = "[INFERRED - Platform auto-generates] "
    elif is_context_graph:
        prefix = "[Context Graph] "

    db_table = csv_schema.get("db_table", "")
    title = f"{prefix}{csv_filename} — {db_table}"
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = WRAP

    required_cols = csv_schema.get("required_columns", [])
    optional_cols = csv_schema.get("optional_columns", [])
    all_cols = required_cols + optional_cols

    # Row 2: Column headers
    for c, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=2, column=c, value=col_name)
        cell.fill = HEADER_FILL
        if col_name in required_cols:
            cell.font = REQUIRED_FONT
        else:
            cell.font = OPTIONAL_FONT

    # Row 3: Example data
    for c, col_name in enumerate(all_cols, 1):
        ws.cell(row=3, column=c, value=get_sample(col_name))

    # Row 4: note for inferred tabs
    if is_inferred:
        note = ws.cell(row=4, column=1,
                       value="This CSV is auto-generated by the platform from node data. You may provide it to override, or leave empty.")
        note.font = NOTE_FONT

    # Special note for kpi_measurements
    schema_note = csv_schema.get("note")
    if schema_note:
        note_row = 5 if is_inferred else 4
        ws.cell(row=note_row, column=1, value=schema_note).font = NOTE_FONT

    if csv_filename == "kpi_measurements.csv":
        note_row = 5 if is_inferred else 4
        ws.cell(row=note_row, column=1,
                value="See 'KPI Reference' tab for valid kpi_codes (P1-KPI1 through P5-KPI8) and health ranges.").font = NOTE_FONT

    auto_size_columns(ws)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def build_workbook(output_path: str):
    schemas = load_csv_schemas()
    kpi_defs, pillars = load_kpi_definitions()
    thresholds = load_health_thresholds()
    playbooks = load_playbook_config()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Tab 1: README
    create_readme_tab(wb, kpi_defs, pillars, schemas, thresholds, playbooks)

    # Tab 2: KPI Reference
    create_kpi_reference_tab(wb, kpi_defs, pillars)

    # Tab 3: Weights
    create_weights_tab(wb, kpi_defs, pillars)

    # Tabs 4-7: Regular CSVs
    regular_files = schemas.get("regular_model", {}).get("files", {})
    for csv_name in REGULAR_CSV_ORDER:
        if csv_name in regular_files:
            create_csv_data_tab(wb, csv_name, regular_files[csv_name])

    # Tabs 8-14: Context Graph CSVs
    cg_files = schemas.get("context_graph_model", {}).get("files", {})
    for csv_name in CG_CSV_ORDER:
        if csv_name in cg_files:
            create_csv_data_tab(
                wb, csv_name, cg_files[csv_name],
                is_context_graph=True,
                is_inferred=(csv_name in INFERRED_CSVS),
            )

    wb.save(output_path)

    # Summary
    total_kpis = len(kpi_defs)
    total_pillars = len(pillars)
    print(f"""
CS Pulse Onboarding Template Generated
=======================================
Output: {output_path}
Tabs:   14 (README + KPI Reference + Weights + {len(REGULAR_CSV_ORDER)} Regular CSVs + {len(CG_CSV_ORDER)} Context Graph CSVs)
KPIs:   {total_kpis} across {total_pillars} pillars
CSVs:   11 total (4 regular + 7 context graph, 1 inferred)
""")


def main():
    parser = argparse.ArgumentParser(
        description="Generate CS Pulse DataCenter onboarding template workbook"
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output file path (default: config/CS_Pulse_Onboarding_Template.xlsx)",
    )
    args = parser.parse_args()

    if args.output is None:
        args.output = str(BACKEND_DIR / "config" / "CS_Pulse_Onboarding_Template.xlsx")

    build_workbook(args.output)


if __name__ == "__main__":
    main()
