#!/usr/bin/env python3
"""
Generate a sample .xlsx that combines Customer Profile Input + Champion & Stakeholder Input
for the data center vertical (10 accounts). Use as reference to align synthetic data gen.

Output: verticals/customer19-dc2_s/data/sample_account_profiles_dc2_s.xlsx
Run from backend: python scripts/generate_sample_account_profile_xlsx.py
"""

import os
from pathlib import Path
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

# Account IDs for customer 19 (10 accounts)
ACCOUNT_IDS = [19001 + i for i in range(10)]

# Data center vertical – 10 sample accounts
CUSTOMER_PROFILE_ROWS = [
    {
        "Account ID*": 19001,
        "Account Name*": "Nexus Data Centers - Production",
        "Industry*": "Data Center Infrastructure",
        "Region*": "us-west-2",
        "Status*": "active",
        "Account Tier*": "Enterprise",
        "Assigned CSM*": "Jordan Lee",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Alex Chen",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform, Monitoring Suite",
        "Contract Start Date*": "2024-01-15",
        "Contract End Date*": "2025-12-31",
        "Renewal Date": "2025-12-31",
        "ARR": 1250000,
        "MRR": 104167,
        "Last updated": "2025-01-24 10:15:00",
    },
    {
        "Account ID*": 19002,
        "Account Name*": "CoreSite West - Staging",
        "Industry*": "Colocation & Interconnection",
        "Region*": "us-west-2",
        "Status*": "active",
        "Account Tier*": "Enterprise",
        "Assigned CSM*": "Jordan Lee",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Morgan Taylor",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform",
        "Contract Start Date*": "2024-03-01",
        "Contract End Date*": "2026-02-28",
        "Renewal Date": "2026-02-28",
        "ARR": 980000,
        "MRR": 81667,
        "Last updated": "2025-01-24 10:22:00",
    },
    {
        "Account ID*": 19003,
        "Account Name*": "Equinix Enterprise - Development",
        "Industry*": "Data Center Infrastructure",
        "Region*": "us-east-1",
        "Status*": "active",
        "Account Tier*": "Strategic",
        "Assigned CSM*": "Casey Kim",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Jamie Foster",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform, Monitoring Suite",
        "Contract Start Date*": "2023-06-01",
        "Contract End Date*": "2025-05-31",
        "Renewal Date": "2025-05-31",
        "ARR": 2100000,
        "MRR": 175000,
        "Last updated": "2025-01-23 16:45:00",
    },
    {
        "Account ID*": 19004,
        "Account Name*": "Digital Realty Partner - Environment",
        "Industry*": "Colocation",
        "Region*": "us-east-1",
        "Status*": "active",
        "Account Tier*": "Professional",
        "Assigned CSM*": "Casey Kim",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "",
        "Strategic Account": "No",
        "Products Used*": "DC2_S Platform",
        "Contract Start Date*": "2024-09-01",
        "Contract End Date*": "2025-08-31",
        "Renewal Date": "2025-08-31",
        "ARR": 450000,
        "MRR": 37500,
        "Last updated": "2025-01-24 09:30:00",
    },
    {
        "Account ID*": 19005,
        "Account Name*": "Cyxtera Colo - Workspace",
        "Industry*": "Data Center Infrastructure",
        "Region*": "eu-west-1",
        "Status*": "active",
        "Account Tier*": "Enterprise",
        "Assigned CSM*": "Riley Adams",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Jordan Blake",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform, Monitoring Suite",
        "Contract Start Date*": "2024-01-01",
        "Contract End Date*": "2025-12-31",
        "Renewal Date": "2025-12-31",
        "ARR": 760000,
        "MRR": 63333,
        "Last updated": "2025-01-24 11:00:00",
    },
    {
        "Account ID*": 19006,
        "Account Name*": "Flexential Cloud - Cluster",
        "Industry*": "Cloud Infrastructure",
        "Region*": "us-west-2",
        "Status*": "active",
        "Account Tier*": "Professional",
        "Assigned CSM*": "Riley Adams",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "",
        "Strategic Account": "No",
        "Products Used*": "DC2_S Platform",
        "Contract Start Date*": "2024-05-15",
        "Contract End Date*": "2025-05-14",
        "Renewal Date": "2025-05-14",
        "ARR": 320000,
        "MRR": 26667,
        "Last updated": "2025-01-22 14:20:00",
    },
    {
        "Account ID*": 19007,
        "Account Name*": "QTS Hyperscale - Instance",
        "Industry*": "Data Center Infrastructure",
        "Region*": "us-east-1",
        "Status*": "active",
        "Account Tier*": "Strategic",
        "Assigned CSM*": "Jordan Lee",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Taylor Reed",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform, Monitoring Suite",
        "Contract Start Date*": "2023-09-01",
        "Contract End Date*": "2026-08-31",
        "Renewal Date": "2026-08-31",
        "ARR": 3100000,
        "MRR": 258333,
        "Last updated": "2025-01-24 08:45:00",
    },
    {
        "Account ID*": 19008,
        "Account Name*": "StackPath Edge - Node",
        "Industry*": "Edge & Colocation",
        "Region*": "us-west-2",
        "Status*": "active",
        "Account Tier*": "Professional",
        "Assigned CSM*": "Casey Kim",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "",
        "Strategic Account": "No",
        "Products Used*": "DC2_S Platform",
        "Contract Start Date*": "2024-07-01",
        "Contract End Date*": "2025-06-30",
        "Renewal Date": "2025-06-30",
        "ARR": 280000,
        "MRR": 23333,
        "Last updated": "2025-01-23 17:10:00",
    },
    {
        "Account ID*": 19009,
        "Account Name*": "Vantage DC - Server",
        "Industry*": "Data Center Infrastructure",
        "Region*": "us-west-2",
        "Status*": "active",
        "Account Tier*": "Enterprise",
        "Assigned CSM*": "Riley Adams",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Morgan Taylor",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform, Monitoring Suite",
        "Contract Start Date*": "2024-02-01",
        "Contract End Date*": "2026-01-31",
        "Renewal Date": "2026-01-31",
        "ARR": 890000,
        "MRR": 74167,
        "Last updated": "2025-01-24 12:30:00",
    },
    {
        "Account ID*": 19010,
        "Account Name*": "CyrusOne Global - System",
        "Industry*": "Colocation & Interconnection",
        "Region*": "eu-west-1",
        "Status*": "active",
        "Account Tier*": "Strategic",
        "Assigned CSM*": "Jordan Lee",
        "CSM Manager": "Sam Rivera",
        "Executive Sponsor": "Alex Chen",
        "Strategic Account": "Yes",
        "Products Used*": "DC2_S Platform, Monitoring Suite",
        "Contract Start Date*": "2023-12-01",
        "Contract End Date*": "2025-11-30",
        "Renewal Date": "2025-11-30",
        "ARR": 1850000,
        "MRR": 154167,
        "Last updated": "2025-01-24 13:00:00",
    },
]

# Champion & Stakeholder – one row per account (can expand to multiple champions per account later)
CHAMPION_COLUMNS = [
    "Account ID*", "Account Name", "Primary Champion Name*", "Champion Title", "Champion Email",
    "Champion Status", "Champion Tenure (months)", "Champion Influence Level",
    "Economic Buyer Name", "Economic Buyer Title", "Executive Sponsor Name", "Executive Sponsor Title",
    "Technical Champion Name", "Number of Active Champions",
]
CHAMPION_ROWS = [
    (19001, "Nexus Data Centers - Production", "Chris Martinez", "VP Infrastructure", "cmartinez@nexusdc.example.com", "Active", 18, "High", "Dana Wright", "CFO", "Alex Chen", "VP Sales", "Jordan Blake", 2),
    (19002, "CoreSite West - Staging", "Sam Patel", "Director of Cloud Ops", "spatel@coresite.example.com", "Active", 12, "High", "Morgan Taylor", "CTO", "Morgan Taylor", "CTO", "Sam Patel", 1),
    (19003, "Equinix Enterprise - Development", "Jamie Foster", "CTO", "jfoster@equinix.example.com", "Active", 24, "Executive", "Jamie Foster", "CTO", "Jamie Foster", "CTO", "Riley Adams", 3),
    (19004, "Digital Realty Partner - Environment", "Taylor Reed", "Infrastructure Manager", "treed@digrealty.example.com", "Active", 6, "Medium", "Casey Kim", "Director IT", "", "", "Taylor Reed", 1),
    (19005, "Cyxtera Colo - Workspace", "Jordan Blake", "VP Engineering", "jblake@cyxtera.example.com", "Active", 14, "High", "Jordan Blake", "VP Engineering", "Jordan Blake", "VP Engineering", "Alex Chen", 2),
    (19006, "Flexential Cloud - Cluster", "Riley Adams", "Cloud Architect", "radams@flexential.example.com", "Active", 8, "Medium", "Dana Wright", "VP Operations", "", "", "Riley Adams", 1),
    (19007, "QTS Hyperscale - Instance", "Morgan Taylor", "SVP Data Center Ops", "mtaylor@qts.example.com", "Active", 30, "Executive", "Morgan Taylor", "SVP Data Center Ops", "Taylor Reed", "EVP", "Casey Kim", 4),
    (19008, "StackPath Edge - Node", "Casey Kim", "Director of Infrastructure", "ckim@stackpath.example.com", "Active", 10, "Medium", "Sam Rivera", "Director Procurement", "", "", "Casey Kim", 1),
    (19009, "Vantage DC - Server", "Alex Chen", "VP Technology", "achen@vantagedc.example.com", "Active", 16, "High", "Alex Chen", "VP Technology", "Morgan Taylor", "CTO", "Jordan Lee", 2),
    (19010, "CyrusOne Global - System", "Dana Wright", "Director Global Infrastructure", "dwright@cyrusone.example.com", "Active", 22, "Executive", "Dana Wright", "Director Global Infrastructure", "Alex Chen", "VP Sales", "Jamie Foster", 3),
]


def main():
    backend_dir = Path(__file__).resolve().parent.parent
    out_dir = backend_dir / "verticals" / "customer19-dc2_s" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "sample_account_profiles_dc2_s.xlsx"

    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # ---- Sheet 1: Customer Profile Input ----
    ws1 = wb.active
    ws1.title = "Customer Profile Input"
    profile_headers = list(CUSTOMER_PROFILE_ROWS[0].keys())
    for col, h in enumerate(profile_headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, row_data in enumerate(CUSTOMER_PROFILE_ROWS, 2):
        for col_idx, h in enumerate(profile_headers, 1):
            ws1.cell(row=row_idx, column=col_idx, value=row_data.get(h, ""))
    for col in range(1, len(profile_headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 14
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 36

    # ---- Sheet 2: Champion & Stakeholder Input ----
    ws2 = wb.create_sheet("Champion & Stakeholder Input", 1)
    for col, h in enumerate(CHAMPION_COLUMNS, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, row_tuple in enumerate(CHAMPION_ROWS, 2):
        for col_idx, val in enumerate(row_tuple, 1):
            ws2.cell(row=row_idx, column=col_idx, value=val)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 22

    wb.save(out_path)
    print(f"Created: {out_path}")
    print("Sheets: 'Customer Profile Input' (10 rows), 'Champion & Stakeholder Input' (10 rows)")
    return str(out_path)


if __name__ == "__main__":
    main()
