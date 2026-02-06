#!/usr/bin/env python3
"""
Onboarding Wizard API
=====================

Comprehensive API for onboarding wizard CSV uploads.
Handles all 5 CSV file types and populates ALL database fields:
- accounts.csv → accounts table (with all fields)
- kpis.csv → kpi_measurements table
- signals.csv → qualitative_signals table
- products.csv → account_products table
- profiles.csv → account_profiles table (all 121 columns)

This fixes the limitations of the old onboarding APIs that only updated
a few fields and stored data in JSON metadata.
"""

from flask import Blueprint, request, jsonify, current_app, send_file
from auth_middleware import get_current_customer_id, get_current_user_id
from werkzeug.utils import secure_filename
import pandas as pd
from extensions import db
from models import Account, Customer, User, CustomerConfig
from werkzeug.security import generate_password_hash
import json
from sqlalchemy import text
import io
import os
import sys
import subprocess
import shutil
import time
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from onboarding_logger import get_onboarding_logger, close_onboarding_session
import random

# Import Excel import services from _template (source)
# These services are in verticals/_template/services/
TEMPLATE_SERVICES_DIR = Path(__file__).parent.parent / "verticals" / "_template" / "services"
EXCEL_IMPORT_AVAILABLE = False
ExcelImportService = None
ImportIntegrationAdapter = None

def _load_excel_import_services():
    """Lazy load Excel import services from _template/services"""
    global EXCEL_IMPORT_AVAILABLE, ExcelImportService, ImportIntegrationAdapter
    
    if EXCEL_IMPORT_AVAILABLE:
        return True
    
    if not TEMPLATE_SERVICES_DIR.exists():
        import logging
        logging.warning(f"Template services directory not found: {TEMPLATE_SERVICES_DIR}")
        return False
    
    try:
        # Add to path for imports
        if str(TEMPLATE_SERVICES_DIR) not in sys.path:
            sys.path.insert(0, str(TEMPLATE_SERVICES_DIR))
        
        # Import services (they import each other)
        # These services are in _template/services/ and use relative imports
        from excel_import_service import ExcelImportService as _ExcelImportService
        from import_integration_adapter import ImportIntegrationAdapter as _ImportIntegrationAdapter
        
        ExcelImportService = _ExcelImportService
        ImportIntegrationAdapter = _ImportIntegrationAdapter
        EXCEL_IMPORT_AVAILABLE = True
        import logging
        logging.info(f"✅ Loaded Excel import services from {TEMPLATE_SERVICES_DIR}")
        return True
    except ImportError as e:
        import logging
        import traceback
        logging.warning(f"Excel import services not available from {TEMPLATE_SERVICES_DIR}: {e}")
        logging.debug(traceback.format_exc())
        return False
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Error loading Excel import services: {e}")
        logging.debug(traceback.format_exc())
        return False

onboarding_api = Blueprint('onboarding_api', __name__)

# File type mapping
FILE_TYPES = {
    'accounts': {
        'table': 'accounts',
        'description': 'Account list',
        'required': True
    },
    'kpis': {
        'table': 'kpi_measurements',
        'description': 'Historical KPI measurements',
        'required': False
    },
    'signals': {
        'table': 'qualitative_signals',
        'description': 'Qualitative signals',
        'required': False
    },
    'products': {
        'table': 'account_products',
        'description': 'Account-product associations',
        'required': False
    },
    'profiles': {
        'table': 'account_profiles',
        'description': 'Detailed account profiles',
        'required': False
    }
}


def prepare_accounts_dataframe(df, customer_id):
    """Prepare accounts dataframe with all required fields"""
    # Set defaults for missing columns
    if 'account_status' not in df.columns:
        df['account_status'] = 'active'
    if 'external_account_id' not in df.columns:
        df['external_account_id'] = df['account_id'].astype(str)
    if 'revenue' not in df.columns:
        # Use final_arr if available, otherwise initial_arr
        if 'final_arr' in df.columns:
            df['revenue'] = df['final_arr']
        elif 'initial_arr' in df.columns:
            df['revenue'] = df['initial_arr']
        else:
            df['revenue'] = 0
    if 'region' not in df.columns:
        # Extract region from datacenter_location if available
        if 'datacenter_location' in df.columns:
            df['region'] = df['datacenter_location'].apply(
                lambda x: x.split('(')[0].strip() if pd.notna(x) and '(' in str(x) else 'Global'
            )
        else:
            df['region'] = 'Global'
    if 'vertical' not in df.columns:
        # Try to infer from customer or set default
        df['vertical'] = 'DC2_S'  # Default for data center customers
    if 'created_at' not in df.columns:
        df['created_at'] = datetime.now()
    if 'updated_at' not in df.columns:
        df['updated_at'] = datetime.now()
    
    # Ensure customer_id is set
    df['customer_id'] = customer_id
    
    return df


def prepare_profiles_dataframe(df):
    """Prepare account_profiles dataframe with boolean conversions"""
    # Convert boolean columns - most are boolean, but co_marketing_opportunities is integer
    bool_columns = ['strategic_account', 'reference_customer', 'case_study_approved', 
                  'logo_usage_approved', 'innovation_partner', 'advisory_board_member',
                  'narrative_available', 'budget_approved_next_year']
    for col in bool_columns:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: True if (x == True or str(x).lower() == 'true' or x == 1 or str(x) == '1') else False
            )
    
    # Integer column (co_marketing_opportunities)
    if 'co_marketing_opportunities' in df.columns:
        df['co_marketing_opportunities'] = df['co_marketing_opportunities'].apply(
            lambda x: 1 if (x == True or str(x).lower() == 'true' or x == 1 or str(x) == '1') else 0
        )
    
    return df


def prepare_kpis_dataframe(df):
    """Prepare kpi_measurements dataframe"""
    # Rename 'date' to 'measurement_month' if needed
    if 'date' in df.columns and 'measurement_month' not in df.columns:
        df = df.rename(columns={'date': 'measurement_month'})
    
    # Convert boolean columns
    if 'threshold_breached' in df.columns:
        df['threshold_breached'] = df['threshold_breached'].apply(
            lambda x: True if (x == True or str(x).lower() == 'true' or x == 1 or str(x) == '1') else False
        )
    
    return df


def prepare_signals_dataframe(df):
    """Prepare qualitative_signals dataframe"""
    # Convert boolean columns
    if 'is_narrative_signal' in df.columns:
        df['is_narrative_signal'] = df['is_narrative_signal'].apply(
            lambda x: True if (x == True or str(x).lower() == 'true' or x == 1 or str(x) == '1') else False
        )
    
    return df


def prepare_products_dataframe(df):
    """Prepare account_products dataframe"""
    # Convert boolean columns
    if 'primary_product' in df.columns:
        df['primary_product'] = df['primary_product'].apply(
            lambda x: True if (x == True or str(x).lower() == 'true' or x == 1 or str(x) == '1') else False
        )
    
    return df


# Helper function to get customer directory path
def get_customer_directory(customer_id: int, vertical_slug: str = 'dc2_s') -> Path:
    """Get path to customer directory"""
    # Path(__file__) is onboarding_api.py in backend/, so parent is backend/, parent.parent is kpi-dashboard/
    # But verticals/ is actually in backend/, so we need backend/verticals/
    verticals_dir = Path(__file__).parent / "verticals"
    return verticals_dir / f"customer{customer_id}-{vertical_slug}"


# Helper function to calculate expected account ID range
def calculate_account_id_range(customer_id: int) -> Tuple[int, int]:
    """
    Calculate expected account ID range for a customer
    
    Formula: account_id_start = customer_id * 1000
    Range: account_id_start to account_id_start + 999
    
    Returns:
        (start_id, end_id) tuple
    """
    start_id = customer_id * 1000
    end_id = start_id + 999
    return (start_id, end_id)


# Helper function to validate account IDs in uploaded file
def validate_account_ids_in_file(file_path: Path, customer_id: int, file_type: str) -> Tuple[bool, List[str]]:
    """
    Validate that account IDs in uploaded file match expected range
    
    Args:
        file_path: Path to uploaded file
        customer_id: Customer ID
        file_type: Type of file ('accounts', 'kpis', 'signals', 'products', 'profiles')
    
    Returns:
        (is_valid, errors) tuple
    """
    errors = []
    expected_start, expected_end = calculate_account_id_range(customer_id)
    
    try:
        # Read file based on type
        if file_path.suffix.lower() == '.csv':
            df = pd.read_csv(file_path, nrows=1000)  # Sample first 1000 rows for validation
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            # For Excel, read first sheet
            try:
                df = pd.read_excel(file_path, nrows=1000, engine='openpyxl' if file_path.suffix.lower() == '.xlsx' else None)
            except ImportError:
                # Fallback if openpyxl not available
                df = pd.read_excel(file_path, nrows=1000)
        else:
            return (True, [])  # Skip validation for unknown file types
        
        if df.empty:
            return (True, [])  # Empty file, no validation needed
        
        # Determine account_id column name
        account_id_col = None
        possible_names = ['account_id', 'Account ID', 'AccountID', 'accountId', 'ACCOUNT_ID']
        for col in df.columns:
            if col.strip() in possible_names:
                account_id_col = col
                break
        
        if not account_id_col:
            # If no account_id column, skip validation (might be profiles or products)
            if file_type in ['profiles', 'products']:
                return (True, [])
            # For other types, warn but don't fail
            return (True, [f"Warning: No account_id column found in {file_type} file"])
        
        # Validate account IDs
        account_ids = df[account_id_col].dropna().unique()
        invalid_ids = []
        
        for account_id in account_ids:
            try:
                account_id_int = int(float(account_id))  # Handle float account IDs
                if account_id_int < expected_start or account_id_int > expected_end:
                    invalid_ids.append(account_id_int)
            except (ValueError, TypeError):
                # Skip non-numeric account IDs (might be strings like "NEW_ACCOUNT")
                continue
        
        if invalid_ids:
            errors.append(
                f"Invalid account IDs found: {sorted(set(invalid_ids))[:10]}. "
                f"Expected range: {expected_start}-{expected_end} for Customer {customer_id}. "
                f"Formula: customer_id * 1000"
            )
            return (False, errors)
        
        return (True, [])
        
    except Exception as e:
        # If validation fails due to file reading error, log but don't block upload
        current_app.logger.warning(f"Account ID validation error for {file_path}: {e}")
        return (True, [f"Warning: Could not validate account IDs: {str(e)}"])


# Helper function to execute Python script
def execute_script(script_path: Path, customer_id: int, timeout: int = 300, 
                   additional_args: List[str] = None, env: dict = None) -> Tuple[bool, str, str]:
    """
    Execute a Python script synchronously
    
    Args:
        script_path: Path to script
        customer_id: Customer ID to pass as environment variable
        timeout: Execution timeout in seconds
        additional_args: Additional command-line arguments to pass to script
        env: Additional environment variables (merged with CUSTOMER_ID)
    
    Returns:
        (success, stdout, stderr)
    """
    try:
        # Change to script directory
        script_dir = script_path.parent
        script_name = script_path.name
        
        # Build command
        cmd = [sys.executable, script_name]
        if additional_args:
            cmd.extend(additional_args)
        
        # Build environment
        script_env = {**os.environ, 'CUSTOMER_ID': str(customer_id)}
        if env:
            script_env.update(env)
        
        # Run script
        result = subprocess.run(
            cmd,
            cwd=str(script_dir),
            capture_output=True,
            text=True,
            timeout=timeout,
            env=script_env
        )
        
        return (result.returncode == 0, result.stdout, result.stderr)
    except subprocess.TimeoutExpired:
        return (False, "", f"Script execution timed out after {timeout} seconds")
    except Exception as e:
        return (False, "", str(e))


def _generate_journey_api_template(customer_id: int, vertical_slug: str = 'dc2_s') -> str:
    """
    Generate journey API template file for a customer.
    Based on customer113_journey_api.py pattern.
    """
    template = f'''"""
Customer {customer_id} Journey API Adapter
{"=" * (len(str(customer_id)) + 30)}

Serves journey data from JSON files to the frontend JourneyVisualizerV2 component.
Reads from wizard_a outputs directory and converts to the expected API format.

Endpoint: GET /api/journey/<account_id>
"""

from flask import Blueprint, jsonify
from pathlib import Path
import json
import os
import csv
from collections import defaultdict
from datetime import datetime, timedelta

journey_api_c{customer_id} = Blueprint('journey_api_c{customer_id}', __name__, url_prefix='/api/journey')

# Base path to customer {customer_id} journey data
BASE_DIR = Path(__file__).parent.parent.parent.parent
C{customer_id}_JOURNEY_BASE = BASE_DIR / "verticals/customer{customer_id}-{vertical_slug}/journey/wizard_a"
C{customer_id}_DATA_DIR = BASE_DIR / "verticals/customer{customer_id}-{vertical_slug}/data"

def load_journey_json(account_id):
    """
    Load journey JSON file for an account.
    Customer {customer_id} uses direct account IDs ({customer_id}001, {customer_id}002, etc.)
    """
    # Check outputs directory first (most recent)
    outputs_dir = C{customer_id}_JOURNEY_BASE / "outputs"
    journey_file = outputs_dir / f"account_{{account_id}}_journey.json"
    
    if journey_file.exists():
        with open(journey_file, 'r') as f:
            journey_data = json.load(f)
            return journey_data
    
    # Fallback: check test_run directories
    if not C{customer_id}_JOURNEY_BASE.exists():
        return None
    
    test_runs = []
    for item in C{customer_id}_JOURNEY_BASE.iterdir():
        if item.is_dir() and (item.name.startswith('test_run_') or item.name == 'outputs'):
            test_runs.append((item.stat().st_mtime, item))
    
    if not test_runs:
        return None
    
    # Try most recent first
    test_runs.sort(reverse=True)
    for _, test_dir in test_runs:
        # Check data subdirectory if it exists
        data_dir = test_dir / "data"
        if data_dir.exists():
            journey_file = data_dir / f"account_{{account_id}}_journey.json"
        else:
            journey_file = test_dir / f"account_{{account_id}}_journey.json"
        
        if journey_file.exists():
            with open(journey_file, 'r') as f:
                journey_data = json.load(f)
                return journey_data
    
    return None


def date_to_week_number(date_str, start_date_str):
    """Convert a date string to week number based on journey start date"""
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        days_diff = (date - start_date).days
        week_num = (days_diff // 7) + 1
        return max(1, week_num)  # Ensure at least week 1
    except:
        return None


def convert_to_weekly_format(journey_data):
    """
    Convert customer {customer_id} journey JSON format to frontend visualizer format.
    """
    if not journey_data:
        return None
    
    # Group events by week
    events_by_week = defaultdict(list)
    health_by_week = {{}}
    phases_by_week = {{}}
    dates_by_week = {{}}
    
    for event in journey_data.get('events', []):
        week_num = event.get('week_number')
        if week_num:
            events_by_week[week_num].append({{
                'event_type': event.get('event_type'),
                'description': event.get('description'),
                'sentiment': event.get('sentiment'),
                'date': event.get('date'),
                'health_impact': event.get('health_score_after', 0) - event.get('health_score_before', 0)
            }})
            
            # Use the health_score_after as the health for that week
            if week_num not in health_by_week:
                health_by_week[week_num] = event.get('health_score_after', 0)
                phases_by_week[week_num] = event.get('phase', 'unknown')
                dates_by_week[week_num] = event.get('date', '')
    
    # Build weekly_data array
    weekly_data = []
    
    # Get all weeks from start to end
    total_weeks = journey_data.get('total_weeks', 52)
    start_date = journey_data.get('start_date', '2024-01-01')
    account_id = journey_data.get('account_id', '')
    
    # Get starting health for interpolation
    starting_health = journey_data.get('starting_health', 0)
    ending_health = journey_data.get('ending_health', 0)
    
    for week_num in range(1, total_weeks + 1):
        # Get health score for this week (use after score from events, or interpolate)
        health_score = health_by_week.get(week_num)
        
        # If no event for this week, interpolate from surrounding weeks
        if health_score is None:
            # Find the most recent week with health data before or at this week
            previous_weeks = [w for w in sorted(health_by_week.keys()) if w <= week_num]
            if previous_weeks:
                # Use the most recent week's health
                health_score = health_by_week[previous_weeks[-1]]
            else:
                # No events before this week, use starting health
                health_score = starting_health
            
            # If still None (shouldn't happen), use ending health
            if health_score is None:
                health_score = ending_health
        
        # Calculate data quality metrics
        available_kpis = []  # Could load from KPI data if available
        coverage_pct = 0.0
        data_quality = 0.0
        
        # Calculate pillar scores from journey data if available
        pillar_scores = journey_data.get('pillar_scores', {{}})
        if not pillar_scores:
            # Default pillar scores
            pillar_scores = {{
                'Infrastructure Health': 0.0,
                'Operational Efficiency': 0.0,
                'Revenue Protection': 0.0,
                'Customer Success': 0.0,
                'Strategic Growth': 0.0
            }}
        
        week_entry = {{
            'week_number': week_num,
            'date': dates_by_week.get(week_num, start_date),
            'health_score': float(health_score) if health_score else 0,
            'phase': phases_by_week.get(week_num, 'unknown'),
            'events': events_by_week.get(week_num, []),
            'raw_kpis': {{}},
            'normalized_kpis': {{}},
            'kpis': {{}},
            'pillar_scores': pillar_scores,
            'pillars': pillar_scores,
            'data_quality': data_quality,
            'coverage_pct': coverage_pct,
            'available_kpis': available_kpis,
            'missing_kpis': [],
            'signals': []  # Could load from signals data if available
        }}
        
        weekly_data.append(week_entry)
    
    return weekly_data


@journey_api_c{customer_id}.route('/<account_id>', methods=['GET'])
def get_account_journey(account_id):
    """
    Get full journey data for an account.
    This is the main endpoint used by JourneyVisualizerV2.
    Only handles Customer {customer_id} accounts ({customer_id}001-{customer_id}999).
    """
    try:
        # Only handle Customer {customer_id} accounts (start with {customer_id})
        account_id_str = str(account_id)
        if not account_id_str.startswith('{customer_id}'):
            # This account doesn't belong to Customer {customer_id}, let other journey APIs handle it
            return None  # Return None to let Flask try next route handler
        
        # Load journey JSON
        journey_data = load_journey_json(account_id_str)
        
        if not journey_data:
            return jsonify({{
                'error': f'Account {{account_id}} journey data not found',
                'message': 'No journey JSON file found for this account. This account may not have journey data generated yet.'
            }}), 404
        
        # Convert to weekly format
        weekly_data = convert_to_weekly_format(journey_data)
        
        if not weekly_data:
            return jsonify({{
                'error': 'Failed to convert journey data',
                'message': 'Could not process journey JSON format'
            }}), 500
        
        # Collect all signals from weekly data for separate signals array
        all_signals = []
        for week in weekly_data:
            for signal in week.get('signals', []):
                all_signals.append(signal)
        
        # Count total KPI readings
        total_kpi_readings = sum(len(w.get('available_kpis', [])) for w in weekly_data)
        
        # Get real account name from journey data or accounts.csv
        account_name = journey_data.get('account_name', f'Account {{account_id}}')
        accounts_file = C{customer_id}_DATA_DIR / "accounts.csv"
        if accounts_file.exists():
            try:
                with open(accounts_file, 'r') as af:
                    reader = csv.DictReader(af)
                    for row in reader:
                        if row.get('account_id') == account_id_str:
                            account_name = row.get('account_name', account_name)
                            break
            except Exception as e:
                print(f"Warning: Could not load account name from CSV: {{e}}")
        
        # Build response
        response = {{
            'account_id': account_id,
            'account_name': account_name,
            'pattern_type': journey_data.get('pattern_type', 'unknown'),
            'total_weeks': journey_data.get('total_weeks', len(weekly_data)),
            'starting_health': float(journey_data.get('starting_health', 0)),
            'ending_health': float(journey_data.get('ending_health', 0)),
            'weekly_data': weekly_data,
            'signals': all_signals,
            'summary': {{
                'total_events': len(journey_data.get('events', [])),
                'total_kpi_readings': total_kpi_readings,
                'total_signals': len(all_signals),
                'weeks_with_data': len([w for w in weekly_data if w.get('events') or w.get('available_kpis')]),
                'lowest_health': float(journey_data.get('lowest_health', 0)),
                'highest_health': float(journey_data.get('highest_health', 0))
            }}
        }}
        
        # Add summary data if available
        if 'summary' in journey_data:
            response['journey_summary'] = journey_data['summary']
        
        return jsonify(response)
        
    except Exception as e:
        import traceback
        return jsonify({{
            'error': str(e),
            'traceback': traceback.format_exc()
        }}), 500


@journey_api_c{customer_id}.route('/accounts', methods=['GET'])
def list_journey_accounts():
    """
    List all accounts that have journey data.
    """
    try:
        outputs_dir = C{customer_id}_JOURNEY_BASE / "outputs"
        accounts = []
        
        if outputs_dir.exists():
            for journey_file in outputs_dir.glob('account_*_journey.json'):
                account_id = journey_file.stem.replace('account_', '').replace('_journey', '')
                
                try:
                    with open(journey_file, 'r') as f:
                        journey_data = json.load(f)
                    
                    accounts.append({{
                        'account_id': journey_data.get('account_id', account_id),
                        'account_name': journey_data.get('account_name', f'Account {{account_id}}'),
                        'pattern_type': journey_data.get('pattern_type', 'unknown'),
                        'total_weeks': journey_data.get('total_weeks', 0),
                        'starting_health': float(journey_data.get('starting_health', 0)),
                        'ending_health': float(journey_data.get('ending_health', 0)),
                        'final_outcome': journey_data.get('summary', {{}}).get('final_outcome')
                    }})
                except Exception as e:
                    print(f"Error loading {{journey_file}}: {{e}}")
                    continue
        
        return jsonify({{
            'accounts': accounts,
            'total': len(accounts)
        }})
        
    except Exception as e:
        import traceback
        return jsonify({{
            'error': str(e),
            'traceback': traceback.format_exc()
        }}), 500
'''
    return template


def get_default_pillar_weights(vertical: str) -> dict:
    """
    Get default pillar weights for a vertical.
    
    Args:
        vertical: Vertical type ('dc2_s', 'saas', etc.)
    
    Returns:
        Dictionary of pillar weights (normalized to sum to 1.0)
    """
    if vertical == 'dc2_s':
        # DC2_S 5-Pillar Supermicro Model (FIXED IDs and weights)
        return {
            'P1_deployment_velocity': 0.15,
            'P2_operational_stability': 0.20,
            'P3_ai_workload_performance': 0.25,
            'P4_channel_partner_health': 0.15,
            'P5_expansion_readiness': 0.25
        }
    elif vertical == 'saas':
        # Generic SaaS model
        return {
            'P1': 0.20,
            'P2': 0.25,
            'P3': 0.20,
            'P4': 0.15,
            'P5': 0.20
        }
    else:
        # Default equal weights
        return {
            'P1': 0.20,
            'P2': 0.20,
            'P3': 0.20,
            'P4': 0.20,
            'P5': 0.20
        }


def handle_excel_upload(file, customer_id, user_id):
    """
    Handle Excel file upload using full import pipeline:
    Excel → excel_import_service → import_integration_adapter → kpi_normalization_service → PostgreSQL + Qdrant
    """
    try:
        # Save uploaded file temporarily
        upload_dir = Path(current_app.config.get('UPLOAD_FOLDER', 'uploads'))
        upload_dir.mkdir(exist_ok=True)
        
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_file_path = upload_dir / f"onboarding_{customer_id}_{timestamp}_{filename}"
        
        file.save(str(temp_file_path))
        current_app.logger.info(f"Saved Excel file to: {temp_file_path}")
        
        # Initialize import adapter (uses services from _template)
        try:
            adapter = ImportIntegrationAdapter()
        except Exception as e:
            current_app.logger.error(f"Failed to initialize ImportIntegrationAdapter: {e}")
            return jsonify({
                'status': 'error',
                'message': f'Excel import service initialization failed: {str(e)}'
            }), 500
        
        # Run full import pipeline
        try:
            result = adapter.import_excel_to_cs_pulse(str(temp_file_path))
            
            # Clean up temp file
            try:
                temp_file_path.unlink()
            except:
                pass
            
            if result.success:
                return jsonify({
                    'status': 'success',
                    'message': f'Successfully imported Excel file via import pipeline',
                    'file_type': 'excel',
                    'customers_imported': result.customers_imported,
                    'weeks_imported': result.weeks_imported,
                    'qdrant_points': len(result.qdrant_point_ids),
                    'warnings': result.warnings[:10] if result.warnings else []  # Limit warnings
                })
            else:
                return jsonify({
                    'status': 'error',
                    'message': 'Excel import completed with errors',
                    'errors': result.errors[:10],  # Limit errors
                    'warnings': result.warnings[:10] if result.warnings else []
                }), 400
                
        except Exception as e:
            # Clean up temp file on error
            try:
                temp_file_path.unlink()
            except:
                pass
            
            current_app.logger.error(f"Error in Excel import pipeline: {e}", exc_info=True)
            return jsonify({
                'status': 'error',
                'message': f'Excel import pipeline failed: {str(e)}'
            }), 500
            
    except Exception as e:
        current_app.logger.error(f"Error handling Excel upload: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Failed to process Excel file: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/upload', methods=['POST'])
def upload_onboarding_csv():
    """
    Upload CSV or Excel file to customer directory (scripts handle DB loading)
    
    Files are saved to customer{N}-dc2_s/data/ directory.
    Scripts (02_load_customer{N}_data_SMART.py) will handle loading to PostgreSQL.
    
    Expected form data:
    - file: CSV or Excel file (.csv, .xlsx, .xls)
    - file_type: One of 'accounts', 'kpis', 'signals', 'products', 'profiles'
    
    Returns:
    - status: success/error
    - file_path: Path where file was saved
    - message: Status message
    """
    try:
        customer_id = get_current_customer_id()
        user_id = get_current_user_id()
        
        if not customer_id:
            return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
        
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        file_type = request.form.get('file_type', 'accounts')
        upload_mode = request.form.get('upload_mode', 'incremental')  # Default to incremental
        
        # Validate upload_mode
        valid_modes = ['full_refresh', 'incremental', 'upsert', 'merge']
        if upload_mode not in valid_modes:
            return jsonify({
                'status': 'error',
                'message': f'Invalid upload_mode: {upload_mode}. Valid modes: {", ".join(valid_modes)}'
            }), 400
        
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'Empty filename'}), 400
        
        customer_id = int(customer_id)
        user_id = int(user_id) if user_id else None
        
        # Get customer directory
        customer_dir = get_customer_directory(customer_id)
        data_dir = customer_dir / "data"
        
        # Check if customer directory exists
        if not customer_dir.exists():
            return jsonify({
                'status': 'error',
                'message': f'Customer directory not found. Please provision customer first: POST /api/onboarding/provision'
            }), 404
        
        # Create data directory if it doesn't exist
        data_dir.mkdir(parents=True, exist_ok=True)
        
        # Detect file type from extension
        filename = secure_filename(file.filename)
        file_ext = Path(filename).suffix.lower()
        is_excel = file_ext in ['.xlsx', '.xls']
        is_csv = file_ext == '.csv'
        
        if not (is_csv or is_excel):
            return jsonify({
                'status': 'error',
                'message': f'Unsupported file format: {file_ext}. Use .csv or .xlsx'
            }), 400
        
        # Determine target filename based on file_type
        if file_type in FILE_TYPES:
            # Map file_type to expected filename
            filename_map = {
                'accounts': 'accounts.csv',
                'kpis': 'kpi_measurements.csv',
                'signals': 'qualitative_signals.csv',
                'products': 'products.csv',
                'profiles': 'profiles.csv'
            }
            target_filename = filename_map.get(file_type, filename)
        else:
            target_filename = filename
        
        # Save file to customer data directory
        file_path = data_dir / target_filename
        
        # Initialize onboarding logger
        onboarding_logger = get_onboarding_logger(customer_id)
        
        try:
            # Save file temporarily first for validation
            temp_file_path = data_dir / f"temp_{target_filename}"
            file.save(str(temp_file_path))
            
            # Validate account IDs before final save
            is_valid, validation_errors = validate_account_ids_in_file(temp_file_path, customer_id, file_type)
            
            if not is_valid:
                # Remove temp file
                temp_file_path.unlink(missing_ok=True)
                onboarding_logger.log_step(
                    f'UPLOAD_{file_type.upper()}_VALIDATION_ERROR',
                    f'Account ID validation failed',
                    'ERROR',
                    {'errors': validation_errors, 'customer_id': customer_id}
                )
                return jsonify({
                    'status': 'error',
                    'message': 'Account ID validation failed',
                    'errors': validation_errors,
                    'expected_range': f"{calculate_account_id_range(customer_id)[0]}-{calculate_account_id_range(customer_id)[1]}",
                    'log_file': onboarding_logger.get_log_file_path()
                }), 400
            
            # Move temp file to final location
            if temp_file_path.exists():
                temp_file_path.rename(file_path)
            
            file_size = file_path.stat().st_size
            current_app.logger.info(f"✅ Saved file to {file_path}")
            
            # Log file upload
            onboarding_logger.log_upload_file(
                file_type,
                target_filename,
                str(file_path.relative_to(Path(__file__).parent.parent)),
                file_size
            )
            
            # Store upload_mode in a metadata file for the data loading script
            upload_metadata = {
                'upload_mode': upload_mode,
                'uploaded_at': datetime.now().isoformat(),
                'file_type': file_type,
                'filename': target_filename
            }
            metadata_file = data_dir / f".upload_metadata_{file_type}.json"
            with open(metadata_file, 'w') as f:
                json.dump(upload_metadata, f, indent=2)
            
            # Log validation success if warnings were present
            if validation_errors:
                onboarding_logger.log_step(
                    f'UPLOAD_{file_type.upper()}_VALIDATION_WARNING',
                    f'Account ID validation completed with warnings',
                    'WARNING',
                    {'warnings': validation_errors}
                )
        except Exception as e:
            current_app.logger.error(f"Error saving file: {e}")
            onboarding_logger = get_onboarding_logger(customer_id)
            onboarding_logger.log_step(
                f'UPLOAD_{file_type.upper()}_ERROR',
                f'Failed to save file: {str(e)}',
                'ERROR',
                {'error': str(e), 'filename': target_filename}
            )
            return jsonify({
                'status': 'error',
                'message': f'Failed to save file: {str(e)}',
                'log_file': onboarding_logger.get_log_file_path()
            }), 500
        
        # For Excel files, also validate structure (optional)
        if is_excel and _load_excel_import_services():
            try:
                # Quick validation (don't process, just check structure)
                import openpyxl
                wb = openpyxl.load_workbook(str(file_path), data_only=True)
                sheet_count = len(wb.sheetnames)
                current_app.logger.info(f"Excel file has {sheet_count} sheets")
            except Exception as e:
                current_app.logger.warning(f"Could not validate Excel structure: {e}")
        
        return jsonify({
            'status': 'success',
            'message': f'File saved to customer directory',
            'file_type': file_type,
            'filename': target_filename,
            'file_path': str(file_path.relative_to(Path(__file__).parent.parent)),
            'customer_id': customer_id,
            'upload_mode': upload_mode,  # Return upload_mode in response
            'next_step': 'Run POST /api/onboarding/process-data to load data to database',
            'log_file': onboarding_logger.get_log_file_path()
        })
            
    except Exception as e:
        current_app.logger.error(f"Error in onboarding upload: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Unexpected error: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/upload-status', methods=['GET'])
def get_upload_status():
    """
    Get upload status for onboarding wizard
    
    Returns which files have been uploaded to customer directory
    (Checks files in customer{N}-dc2_s/data/ directory)
    """
    try:
        customer_id = get_current_customer_id()
        
        if not customer_id:
            return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
        
        customer_id = int(customer_id)
        customer_dir = get_customer_directory(customer_id)
        data_dir = customer_dir / "data"
        
        status = {}
        
        # Check for files in data directory
        if data_dir.exists():
            # Map file types to expected filenames
            file_mapping = {
                'accounts': 'accounts.csv',
                'kpis': 'kpi_measurements.csv',
                'signals': 'qualitative_signals.csv',
                'products': 'products.csv',
                'profiles': 'profiles.csv'
            }
            
            for file_type, filename in file_mapping.items():
                file_path = data_dir / filename
                if file_path.exists():
                    # Get file size and modification time
                    stat = file_path.stat()
                    status[file_type] = {
                        'uploaded': True,
                        'filename': filename,
                        'size_bytes': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    }
                else:
                    status[file_type] = {
                        'uploaded': False,
                        'filename': filename
                    }
        else:
            # No data directory - no files uploaded
            for file_type in ['accounts', 'kpis', 'signals', 'products', 'profiles']:
                status[file_type] = {
                    'uploaded': False
                }
        
        return jsonify({
            'status': 'success',
            'upload_status': status,
            'data_directory': str(data_dir) if data_dir.exists() else None
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting upload status: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Unexpected error: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/validate-excel', methods=['POST'])
def validate_excel_file():
    """
    Validate Excel file before upload (checks structure, sheets, columns)
    
    Returns validation results with issues and warnings
    """
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        filename = secure_filename(file.filename)
        file_ext = Path(filename).suffix.lower()
        
        if file_ext not in ['.xlsx', '.xls']:
            return jsonify({
                'status': 'error',
                'message': f'Invalid file type: {file_ext}. Expected .xlsx or .xls'
            }), 400
        
        # Save temporarily
        upload_dir = Path(current_app.config.get('UPLOAD_FOLDER', 'uploads'))
        upload_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_file_path = upload_dir / f"validate_{timestamp}_{filename}"
        
        file.save(str(temp_file_path))
        
        # Load Excel import service
        if not _load_excel_import_services():
            return jsonify({
                'status': 'error',
                'message': 'Excel import services not available'
            }), 500
        
        try:
            # Use ExcelImportService to validate
            import openpyxl
            wb = openpyxl.load_workbook(temp_file_path, data_only=True)
            
            # Check required sheets
            required_sheets = ['Customer Profile', 'KPI Coverage Detail']
            optional_sheets = ['KPI History', 'Contracts & Revenue', 'NPS & CSAT', 'Events']
            
            issues = []
            warnings = []
            sheet_info = {}
            
            # Check required sheets
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    issues.append(f"Missing required sheet: '{sheet_name}'")
                else:
                    ws = wb[sheet_name]
                    sheet_info[sheet_name] = {
                        'rows': ws.max_row,
                        'columns': ws.max_column,
                        'exists': True
                    }
            
            # Check optional sheets
            for sheet_name in optional_sheets:
                if sheet_name not in wb.sheetnames:
                    warnings.append(f"Optional sheet '{sheet_name}' not found")
                else:
                    ws = wb[sheet_name]
                    sheet_info[sheet_name] = {
                        'rows': ws.max_row,
                        'columns': ws.max_column,
                        'exists': True
                    }
            
            # Validate Customer Profile sheet columns if it exists
            if 'Customer Profile' in wb.sheetnames:
                ws = wb['Customer Profile']
                # Find header row
                header_row = None
                for row_idx in range(1, min(10, ws.max_row + 1)):
                    cell_value = ws.cell(row_idx, 1).value
                    if cell_value and 'Account ID' in str(cell_value):
                        header_row = row_idx
                        break
                
                if header_row:
                    headers = []
                    for col_idx in range(1, min(ws.max_column + 1, 50)):
                        header = ws.cell(header_row, col_idx).value
                        if header:
                            headers.append(str(header).strip())
                    
                    required_cols = ['Account ID', 'Account Name', 'Industry', 'Region', 'Tier', 'ARR ($)']
                    for col in required_cols:
                        # Check with and without asterisk
                        if col not in headers and col + '*' not in headers:
                            issues.append(f"Customer Profile sheet missing column: '{col}'")
                else:
                    issues.append("Customer Profile sheet: Could not find header row")
            
            # Clean up
            try:
                temp_file_path.unlink()
            except:
                pass
            
            return jsonify({
                'status': 'success' if len(issues) == 0 else 'error',
                'valid': len(issues) == 0,
                'issues': issues,
                'warnings': warnings,
                'sheet_info': sheet_info,
                'filename': filename
            })
            
        except Exception as e:
            # Clean up on error
            try:
                temp_file_path.unlink()
            except:
                pass
            
            current_app.logger.error(f"Error validating Excel: {e}", exc_info=True)
            return jsonify({
                'status': 'error',
                'message': f'Failed to validate Excel file: {str(e)}'
            }), 500
            
    except Exception as e:
        current_app.logger.error(f"Error in Excel validation: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Unexpected error: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/validate', methods=['POST'])
def validate_onboarding_csv():
    """
    Validate CSV file before upload
    
    Checks:
    - File format (CSV)
    - Required columns
    - Data types
    - Basic data integrity
    """
    try:
        customer_id = get_current_customer_id()
        
        if not customer_id:
            return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
        
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        file_type = request.form.get('file_type', 'accounts')
        
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'Empty filename'}), 400
        
        if file_type not in FILE_TYPES:
            return jsonify({
                'status': 'error',
                'message': f'Invalid file_type. Must be one of: {", ".join(FILE_TYPES.keys())}'
            }), 400
        
        # Read and validate CSV
        try:
            df = pd.read_csv(io.BytesIO(file.read()))
        except Exception as e:
            return jsonify({
                'status': 'error',
                'message': f'Failed to parse CSV: {str(e)}'
            }), 400
        
        if df.empty:
            return jsonify({
                'status': 'error',
                'message': 'CSV file is empty'
            }), 400
        
        # Basic validation
        errors = []
        warnings = []
        
        # Check required columns based on file type
        if file_type == 'accounts':
            required_cols = ['account_id', 'customer_id', 'account_name']
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                errors.append(f'Missing required columns: {", ".join(missing)}')
        
        # Check for customer_id consistency
        if 'customer_id' in df.columns:
            customer_id_int = int(customer_id)
            mismatched = df[df['customer_id'] != customer_id_int]
            if len(mismatched) > 0:
                errors.append(f'{len(mismatched)} rows have mismatched customer_id')
        
        # Check for duplicate account_ids (for accounts file)
        if file_type == 'accounts' and 'account_id' in df.columns:
            duplicates = df[df.duplicated(subset=['account_id'], keep=False)]
            if len(duplicates) > 0:
                errors.append(f'Found {len(duplicates)} duplicate account_ids')
        
        return jsonify({
            'status': 'success' if len(errors) == 0 else 'error',
            'valid': len(errors) == 0,
            'rows': len(df),
            'columns': len(df.columns),
            'errors': errors,
            'warnings': warnings
        })
        
    except Exception as e:
        current_app.logger.error(f"Error validating CSV: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Unexpected error: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/provision', methods=['POST'])
def provision_customer_directory():
    """
    Provision customer directory structure from _template
    
    Expected JSON body:
    {
        "customer_id": 18,
        "customer_name": "Acme Corp",  # Optional
        "vertical_slug": "dc2_s"  # Optional, defaults to dc2_s
    }
    
    This endpoint:
    1. Calls provision_dc_customer.py programmatically
    2. Creates customer{N}-dc2_s/ directory structure
    3. Returns provisioning status
    """
    onboarding_logger = None
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'status': 'error', 'message': 'No data provided'}), 400
        
        customer_id = data.get('customer_id')
        if not customer_id:
            return jsonify({
                'status': 'error',
                'message': 'Missing required field: customer_id'
            }), 400
        
        customer_id = int(customer_id)
        customer_name = data.get('customer_name')
        vertical_slug = data.get('vertical_slug', 'dc2_s')
        
        # Initialize onboarding logger
        onboarding_logger = get_onboarding_logger(customer_id, customer_name)
        onboarding_logger.log_provision_start(customer_name or f"Customer {customer_id}", vertical_slug)
        
        # Import provisioning function
        try:
            # Add verticals directory to path (provision_dc_customer.py is in backend/verticals/)
            verticals_path = Path(__file__).parent / "verticals"
            if str(verticals_path.resolve()) not in sys.path:
                sys.path.insert(0, str(verticals_path.resolve()))
            from provision_dc_customer import provision_customer
        except ImportError as e:
            error_msg = f'Failed to import provisioning script: {str(e)}'
            current_app.logger.error(error_msg)
            onboarding_logger.log_provision_error(error_msg)
            return jsonify({
                'status': 'error',
                'message': error_msg
            }), 500
        
        # Check if directory already exists
        customer_dir = get_customer_directory(customer_id, vertical_slug)
        if customer_dir.exists():
            onboarding_logger.log_step(
                'PROVISION_SKIP',
                f'Customer directory already exists',
                'WARNING',
                {'directory': str(customer_dir)}
            )
            return jsonify({
                'status': 'warning',
                'message': f'Customer directory already exists: customer{customer_id}-{vertical_slug}',
                'customer_id': customer_id,
                'directory': str(customer_dir),
                'exists': True,
                'log_file': onboarding_logger.get_log_file_path()
            }), 200
        
        # Call provisioning function
        try:
            success = provision_customer(
                customer_id=customer_id,
                customer_name=customer_name,
                vertical_slug=vertical_slug,
                dry_run=False,
                force=True  # Auto-mode, skip confirmations
            )
            
            if success:
                onboarding_logger.log_provision_complete(str(customer_dir))
                return jsonify({
                    'status': 'success',
                    'message': f'Customer directory provisioned successfully',
                    'customer_id': customer_id,
                    'directory': str(customer_dir),
                    'vertical': vertical_slug,
                    'log_file': onboarding_logger.get_log_file_path()
                })
            else:
                error_msg = 'Provisioning failed - check logs for details'
                onboarding_logger.log_provision_error(error_msg)
                return jsonify({
                    'status': 'error',
                    'message': error_msg,
                    'log_file': onboarding_logger.get_log_file_path()
                }), 500
                
        except Exception as e:
            error_msg = f'Provisioning failed: {str(e)}'
            current_app.logger.error(error_msg, exc_info=True)
            onboarding_logger.log_provision_error(error_msg)
            return jsonify({
                'status': 'error',
                'message': error_msg,
                'log_file': onboarding_logger.get_log_file_path()
            }), 500
            
    except Exception as e:
        error_msg = f'Unexpected error: {str(e)}'
        current_app.logger.error(error_msg, exc_info=True)
        if onboarding_logger:
            onboarding_logger.log_provision_error(error_msg)
        return jsonify({
            'status': 'error',
            'message': error_msg,
            'log_file': onboarding_logger.get_log_file_path() if onboarding_logger else None
        }), 500


@onboarding_api.route('/api/onboarding/complete', methods=['POST'])
def complete_onboarding():
    """
    Complete onboarding wizard - creates customer, user, and config records
    
    Expected JSON body:
    {
        "company_name": "Acme Corp",
        "company_email": "admin@acme.com",
        "admin_name": "John Doe",
        "admin_email": "john@acme.com",
        "admin_password": "secure_password",
        "pillars": {...},  # Optional: pillar configuration
        "weights": {...},  # Optional: category weights
        "vertical": "dc2_s"  # Optional: vertical type
    }
    
    This endpoint:
    1. Creates Customer record in DB
    2. Creates User record (admin)
    3. Saves CustomerConfig (pillars, weights, etc.)
    4. Returns customer_id for subsequent file uploads
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'status': 'error', 'message': 'No data provided'}), 400
        
        # Required fields
        company_name = data.get('company_name')
        company_email = data.get('company_email')
        admin_name = data.get('admin_name')
        admin_email = data.get('admin_email')
        admin_password = data.get('admin_password')
        
        if not all([company_name, company_email, admin_name, admin_email, admin_password]):
            return jsonify({
                'status': 'error',
                'message': 'Missing required fields: company_name, company_email, admin_name, admin_email, admin_password'
            }), 400
        
        # Extract domain from email
        email_domain = admin_email.split('@')[1] if '@' in admin_email else None
        
        # Check if customer_id is provided (from provisioning step)
        provided_customer_id = data.get('customer_id')
        
        # Check if customer already exists (by customer_id, domain, or email)
        existing_customer = None
        if provided_customer_id:
            existing_customer = Customer.query.filter_by(customer_id=provided_customer_id).first()
        if not existing_customer and email_domain:
            existing_customer = Customer.query.filter_by(domain=email_domain).first()
        if not existing_customer:
            existing_customer = Customer.query.filter_by(email=company_email).first()
        
        if existing_customer:
            # Use existing customer
            customer_id = existing_customer.customer_id
            current_app.logger.info(f"Using existing customer: {customer_id} ({company_name})")
        else:
            # Create new Customer
            # If customer_id provided, try to use it (may fail if DB doesn't allow manual ID assignment)
            customer = Customer(
                customer_name=company_name,
                email=company_email,
                domain=email_domain,
                phone=data.get('phone')
            )
            
            # Try to set customer_id if provided (PostgreSQL auto-increment)
            if provided_customer_id:
                try:
                    # Use raw SQL to set sequence so next value is the provided ID
                    from sqlalchemy import text
                    # Set sequence to provided_id - 1, so next value will be provided_id
                    # Use true parameter to actually use the value (not skip it)
                    db.session.execute(
                        text("SELECT setval('customers_customer_id_seq', :max_id, true)"),
                        {"max_id": provided_customer_id - 1}
                    )
                    current_app.logger.info(f"Set sequence to {provided_customer_id - 1}, next customer_id will be {provided_customer_id}")
                except Exception as e:
                    current_app.logger.warning(f"Could not set customer_id sequence: {e}. Will use auto-increment.")
            
            db.session.add(customer)
            db.session.flush()
            customer_id = customer.customer_id
            
            # If customer_id doesn't match, log warning
            if provided_customer_id and customer_id != provided_customer_id:
                current_app.logger.warning(
                    f"Customer ID mismatch: requested {provided_customer_id}, got {customer_id}. "
                    f"Database auto-increment overrode provided ID."
                )
        
        # Initialize onboarding logger
        onboarding_logger = get_onboarding_logger(customer_id, company_name)
        onboarding_logger.log_complete_start(company_name, admin_email)
        
        # Check if user already exists for this customer
        existing_user = User.query.filter_by(
            customer_id=customer_id,
            user_name=admin_name
        ).first()
        
        if existing_user:
            # User already exists - update password and reactivate if needed
            existing_user.password_hash = generate_password_hash(admin_password)
            existing_user.email = admin_email
            existing_user.active = True
            existing_user.vertical = data.get('vertical', 'dc2_s')
            existing_user.role = 'admin'
            db.session.commit()
            user_id = existing_user.user_id
            current_app.logger.info(f"Using existing user: {user_id} ({admin_name}) for customer {customer_id}")
        else:
            # Create User (admin)
            user = User(
                customer_id=customer_id,
                user_name=admin_name,
                email=admin_email,
                password_hash=generate_password_hash(admin_password),
                vertical=data.get('vertical', 'dc2_s'),
                role='admin',
                active=True
            )
            db.session.add(user)
            db.session.flush()
            user_id = user.user_id
            current_app.logger.info(f"Created new user: {user_id} ({admin_name}) for customer {customer_id}")
        
        # Create team members (if provided)
        team_members = data.get('team', [])
        team_user_ids = []
        if team_members:
            for member in team_members:
                if not member.get('email') or not member.get('name'):
                    continue
                
                # Check if email already exists
                existing_team_user = User.query.filter_by(email=member['email']).first()
                if existing_team_user:
                    current_app.logger.warning(f"Team member email {member['email']} already exists, skipping")
                    continue
                
                # Generate password if not provided
                team_password = member.get('password') or f"TempPass{member['email'].split('@')[0]}123!"
                
                team_user = User(
                    customer_id=customer_id,
                    user_name=member.get('name', member['email'].split('@')[0]),
                    email=member['email'],
                    password_hash=generate_password_hash(team_password),
                    vertical=data.get('vertical', 'dc2_s'),
                    role=member.get('role', 'user'),
                    active=True
                )
                db.session.add(team_user)
                db.session.flush()
                team_user_ids.append(team_user.user_id)
        
        # Create CustomerConfig
        # Use provided weights or default DC2_S pillar weights (5-Pillar Supermicro Model)
        # Weights should be normalized (sum to 1.0) and stored as percentages
        provided_weights = data.get('weights') or {}
        vertical = data.get('vertical', 'dc2_s')
        
        # Get default weights for vertical
        default_weights = get_default_pillar_weights(vertical)
        
        # Merge provided weights with defaults (provided weights take precedence)
        final_weights = {**default_weights, **provided_weights}
        
        # Normalize weights to ensure they sum to 1.0
        total = sum(final_weights.values())
        if total > 0:
            final_weights = {k: round(v / total, 4) for k, v in final_weights.items()}
        
        # Check if CustomerConfig already exists
        existing_config = CustomerConfig.query.filter_by(customer_id=customer_id).first()
        
        if existing_config:
            # Update existing config
            existing_config.kpi_upload_mode = data.get('kpi_upload_mode', 'account_rollup')
            existing_config.category_weights = json.dumps(final_weights)
            config = existing_config
            current_app.logger.info(f"Updated existing CustomerConfig for customer {customer_id}")
        else:
            # Create new config
            config = CustomerConfig(
                customer_id=customer_id,
                kpi_upload_mode=data.get('kpi_upload_mode', 'account_rollup'),
                category_weights=json.dumps(final_weights),
                master_file_name=None
            )
            db.session.add(config)
            current_app.logger.info(f"Created new CustomerConfig for customer {customer_id}")
        
        # Commit all changes
        db.session.commit()
        
        current_app.logger.info(f"✅ Onboarding complete: Customer {customer_id} ({company_name}), User {user_id} ({admin_email})")
        
        # Log success
        onboarding_logger.log_complete_success(customer_id, user_id, len(team_user_ids))
        onboarding_logger.log_step(
            'COMPLETE_CONFIG',
            f'CustomerConfig created with pillar weights',
            'SUCCESS',
            {'weights': final_weights, 'vertical': vertical}
        )
        
        return jsonify({
            'status': 'success',
            'message': 'Onboarding completed successfully',
            'customer_id': customer_id,
            'user_id': user_id,
            'customer_name': company_name,
            'admin_email': admin_email,
            'team_members_created': len(team_user_ids),
            'team_user_ids': team_user_ids,
            'log_file': onboarding_logger.get_log_file_path()
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error completing onboarding: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Failed to complete onboarding: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/process-data', methods=['POST'])
def process_customer_data():
    """
    Process customer data: Execute all scripts synchronously
    
    Expected JSON body:
    {
        "customer_id": 18,
        "skip_validation": false,  # Optional: skip validation script
        "skip_wizard_b": true,  # Optional: skip Wizard B (pattern analysis)
        "upload_mode": "incremental"  # Optional: upload mode (full_refresh, incremental, upsert, merge)
    }
    
    Executes in order:
    1. 02_load_customer{N}_data_SMART.py (loads CSV → PostgreSQL)
    2. 03_embed_customer{N}_OPENAI.py (creates embeddings → Qdrant)
    3. 04_validate_data_integrity.py (validates data)
    4. wizard_a_journey_generator.py (creates journey JSON files)
    5. wizard_b_pattern_analyzer.py (optional - pattern analysis)
    
    Returns:
    - status: success/error
    - steps_completed: List of completed steps
    - errors: List of errors (if any)
    """
    try:
        data = request.get_json() or {}
        customer_id = data.get('customer_id') or get_current_customer_id()
        
        if not customer_id:
            return jsonify({'status': 'error', 'message': 'customer_id required'}), 400
        
        customer_id = int(customer_id)
        skip_validation = data.get('skip_validation', False)
        skip_wizard_b = data.get('skip_wizard_b', False)  # Run Wizard B by default
        
        # Get customer directory
        customer_dir = get_customer_directory(customer_id)
        if not customer_dir.exists():
            return jsonify({
                'status': 'error',
                'message': f'Customer directory not found. Please provision first: POST /api/onboarding/provision'
            }), 404
        
        # Check if data directory exists and has files
        data_dir = customer_dir / "data"
        if not data_dir.exists():
            return jsonify({
                'status': 'error',
                'message': 'Data directory not found. Please upload files first.'
            }), 404
        
        # Check for required files
        required_files = ['accounts.csv', 'kpi_measurements.csv']
        missing_files = [f for f in required_files if not (data_dir / f).exists()]
        if missing_files:
            return jsonify({
                'status': 'error',
                'message': f'Missing required files: {", ".join(missing_files)}'
            }), 400
        
        # Initialize onboarding logger
        onboarding_logger = get_onboarding_logger(customer_id)
        onboarding_logger.log_process_data_start(skip_validation, skip_wizard_b)
        
        # Track execution state for rollback
        execution_state = {
            'customer_id': customer_id,
            'steps_completed': [],
            'errors': [],
            'rollback_needed': False,
            'upload_mode': upload_mode  # Store upload mode for rollback logic
        }
        
        # Step 1: Data Loading Script
        current_app.logger.info(f"Step 1: Executing data loading script for customer {customer_id}")
        load_script = customer_dir / "scripts" / f"02_load_customer{customer_id}_data_SMART.py"
        onboarding_logger.log_script_start('02_load_data', str(load_script))
        
        if not load_script.exists():
            execution_state['errors'].append(f"Data loading script not found: {load_script}")
            return jsonify({
                'status': 'error',
                'message': 'Data loading script not found. Customer directory may not be properly provisioned.',
                **execution_state
            }), 404
        
        import time
        script_start_time = time.time()
        # Pass upload_mode to the script via environment variable
        env_vars = {'UPLOAD_MODE': upload_mode}
        success, stdout, stderr = execute_script(load_script, customer_id, timeout=600, env=env_vars)
        script_duration = time.time() - script_start_time
        
        if not success:
            execution_state['errors'].append(f"Data loading failed: {stderr}")
            execution_state['rollback_needed'] = True
            onboarding_logger.log_script_error('02_load_data', stderr, script_duration)
            _rollback_operations(execution_state, onboarding_logger)
            return jsonify({
                'status': 'error',
                'message': 'Data loading script failed',
                'error': stderr,
                'log_file': onboarding_logger.get_log_file_path(),
                **execution_state
            }), 500
        
        execution_state['steps_completed'].append('data_loading')
        current_app.logger.info(f"✅ Data loading completed")
        onboarding_logger.log_script_success('02_load_data', stdout, script_duration)
        
        # Step 2: Embedding Script
        current_app.logger.info(f"Step 2: Executing embedding script for customer {customer_id}")
        embed_script = customer_dir / "scripts" / f"03_embed_customer{customer_id}_OPENAI.py"
        onboarding_logger.log_script_start('03_embed_data', str(embed_script))
        
        if not embed_script.exists():
            execution_state['errors'].append(f"Embedding script not found: {embed_script}")
            execution_state['rollback_needed'] = True
            onboarding_logger.log_script_error('03_embed_data', f"Script not found: {embed_script}", None)
            _rollback_operations(execution_state, onboarding_logger)
            return jsonify({
                'status': 'error',
                'message': 'Embedding script not found',
                'log_file': onboarding_logger.get_log_file_path(),
                **execution_state
            }), 404
        
        script_start_time = time.time()
        success, stdout, stderr = execute_script(embed_script, customer_id, timeout=600)
        script_duration = time.time() - script_start_time
        
        if not success:
            execution_state['errors'].append(f"Embedding failed: {stderr}")
            execution_state['rollback_needed'] = True
            onboarding_logger.log_script_error('03_embed_data', stderr, script_duration)
            _rollback_operations(execution_state, onboarding_logger)
            return jsonify({
                'status': 'error',
                'message': 'Embedding script failed',
                'error': stderr,
                'log_file': onboarding_logger.get_log_file_path(),
                **execution_state
            }), 500
        
        execution_state['steps_completed'].append('embeddings')
        current_app.logger.info(f"✅ Embeddings created")
        onboarding_logger.log_script_success('03_embed_data', stdout, script_duration)
        
        # Step 3: Validation Script (optional)
        if not skip_validation:
            current_app.logger.info(f"Step 3: Executing validation script for customer {customer_id}")
            validate_script = customer_dir / "scripts" / f"04_validate_data_integrity.py"
            onboarding_logger.log_script_start('04_validate_data', str(validate_script))
            
            if validate_script.exists():
                script_start_time = time.time()
                success, stdout, stderr = execute_script(validate_script, customer_id, timeout=300)
                script_duration = time.time() - script_start_time
                
                if not success:
                    execution_state['errors'].append(f"Validation warnings: {stderr}")
                    # Don't fail on validation, just warn
                    current_app.logger.warning(f"Validation script returned warnings: {stderr}")
                    onboarding_logger.log_script_error('04_validate_data', stderr, script_duration)
                else:
                    execution_state['steps_completed'].append('validation')
                    current_app.logger.info(f"✅ Validation completed")
                    onboarding_logger.log_script_success('04_validate_data', stdout, script_duration)
        
        # Step 4: Journey Generator (Wizard A)
        current_app.logger.info(f"Step 4: Executing journey generator for customer {customer_id}")
        journey_script = customer_dir / "journey" / "wizard_a" / "wizard_journey_generator.py"
        
        if not journey_script.exists():
            # Try alternative name
            journey_script = customer_dir / "journey" / "wizard_a" / "wizard_a_journey_generator.py"
        
        onboarding_logger.log_script_start('wizard_a_journey', str(journey_script))
        
        if not journey_script.exists():
            execution_state['errors'].append(f"Journey generator script not found")
            execution_state['rollback_needed'] = True
            onboarding_logger.log_script_error('wizard_a_journey', "Script not found", None)
            _rollback_operations(execution_state, onboarding_logger)
            return jsonify({
                'status': 'error',
                'message': 'Journey generator script not found',
                'log_file': onboarding_logger.get_log_file_path(),
                **execution_state
            }), 404
        
        # Get account count and start ID for Wizard A arguments
        accounts = Account.query.filter_by(customer_id=customer_id).all()
        account_count = len(accounts)
        
        if account_count == 0:
            execution_state['errors'].append(f"No accounts found for customer {customer_id}. Cannot run Wizard A.")
            execution_state['rollback_needed'] = True
            onboarding_logger.log_script_error('wizard_a_journey', "No accounts found", None)
            _rollback_operations(execution_state, onboarding_logger)
            return jsonify({
                'status': 'error',
                'message': 'No accounts found. Data loading may have failed.',
                'log_file': onboarding_logger.get_log_file_path(),
                **execution_state
            }), 500
        
        # Get first account ID as start_id
        start_id = accounts[0].account_id if accounts else customer_id * 1000 + 1
        
        # Default pattern mix
        pattern_mix = '{"crisis":0.2,"churn":0.15,"stable":0.4,"expansion":0.25}'
        
        # Output directory
        output_dir = str(journey_script.parent / "outputs")
        
        # Build additional arguments for Wizard A
        additional_args = [
            '--accounts', str(account_count),
            '--start-id', str(start_id),
            '--pattern-mix', pattern_mix,
            '--output-dir', output_dir
        ]
        
        script_start_time = time.time()
        success, stdout, stderr = execute_script(journey_script, customer_id, timeout=600, additional_args=additional_args)
        script_duration = time.time() - script_start_time
        
        if not success:
            execution_state['errors'].append(f"Journey generation failed: {stderr}")
            execution_state['rollback_needed'] = True
            onboarding_logger.log_script_error('wizard_a_journey', stderr, script_duration)
            _rollback_operations(execution_state, onboarding_logger)
            return jsonify({
                'status': 'error',
                'message': 'Journey generation script failed',
                'error': stderr,
                'log_file': onboarding_logger.get_log_file_path(),
                **execution_state
            }), 500
        
        execution_state['steps_completed'].append('journey_generation')
        current_app.logger.info(f"✅ Journey data generated")
        onboarding_logger.log_script_success('wizard_a_journey', stdout, script_duration)
        
        # Step 5: Wizard B (Pattern Analysis) - Optional
        if not skip_wizard_b:
            current_app.logger.info(f"Step 5: Executing pattern analyzer (Wizard B) for customer {customer_id}")
            wizard_b_script = customer_dir / "journey" / "wizard_b" / "wizard_b_pattern_analyzer.py"
            
            if wizard_b_script.exists():
                success, stdout, stderr = execute_script(wizard_b_script, customer_id, timeout=300)
                if success:
                    execution_state['steps_completed'].append('pattern_analysis')
                    current_app.logger.info(f"✅ Pattern analysis completed")
                else:
                    execution_state['errors'].append(f"Pattern analysis warnings: {stderr}")
                    current_app.logger.warning(f"Pattern analysis returned warnings: {stderr}")
        
        # Step 6: Wizard C (Weight Calibration) - Optional but recommended
        skip_wizard_c = data.get('skip_wizard_c', False)  # Default: run it!
        
        if not skip_wizard_c:
            current_app.logger.info(f"Step 6: Executing weight calibrator (Wizard C) for customer {customer_id}")
            wizard_c_script = customer_dir / "journey" / "wizard_c" / "wizard_c_weight_calibrator.py"
            
            if not wizard_c_script.exists():
                # Try alternative names
                wizard_c_script = customer_dir / "journey" / "wizard_c" / "weight_calibrator.py"
            
            onboarding_logger.log_script_start('wizard_c_weights', str(wizard_c_script))
            
            if wizard_c_script.exists():
                script_start_time = time.time()
                success, stdout, stderr = execute_script(wizard_c_script, customer_id, timeout=300)
                script_duration = time.time() - script_start_time
                
                if success:
                    execution_state['steps_completed'].append('weight_calibration')
                    current_app.logger.info(f"✅ Weight calibration completed")
                    onboarding_logger.log_script_success('wizard_c_weights', stdout, script_duration)
                    
                    # Parse calibrated weights from stdout and update CustomerConfig
                    try:
                        # Wizard C should output JSON with calibrated weights
                        # Example: {"P1_deployment_velocity": 0.18, "P2_operational_stability": 0.22, ...}
                        import re
                        import json
                        
                        # Look for JSON in stdout
                        json_match = re.search(r'\{[^}]+\}', stdout)
                        if json_match:
                            calibrated_weights = json.loads(json_match.group(0))
                            
                            # Update CustomerConfig
                            config = CustomerConfig.query.filter_by(customer_id=customer_id).first()
                            if config:
                                config.category_weights = json.dumps(calibrated_weights)
                                db.session.commit()
                                current_app.logger.info(f"✅ Updated CustomerConfig with calibrated weights")
                                onboarding_logger.log_step(
                                    'WIZARD_C_WEIGHTS_UPDATE',
                                    'CustomerConfig updated with calibrated weights',
                                    'SUCCESS',
                                    {'weights': calibrated_weights}
                                )
                    except Exception as e:
                        current_app.logger.warning(f"Could not parse calibrated weights: {e}")
                else:
                    execution_state['errors'].append(f"Weight calibration warnings: {stderr}")
                    current_app.logger.warning(f"Weight calibration returned warnings: {stderr}")
                    onboarding_logger.log_script_error('wizard_c_weights', stderr, script_duration)
            else:
                current_app.logger.warning(f"Wizard C script not found at {wizard_c_script}")
                onboarding_logger.log_step(
                    'WIZARD_C_SKIP',
                    f'Wizard C script not found, using default weights',
                    'WARNING',
                    {'path': str(wizard_c_script)}
                )
        
        # Step 7: Journey API Ready (Dynamic API handles all customers automatically)
        # No need to create/register customer-specific APIs - the dynamic journey API
        # in journey_api_dynamic.py automatically discovers journey files for any customer
        current_app.logger.info(f"✅ Journey API ready (dynamic discovery enabled for customer {customer_id})")
        onboarding_logger.log_step('JOURNEY_API_READY', 
            f'Dynamic journey API will automatically discover journey files for customer {customer_id}', 
            'SUCCESS')
        execution_state['steps_completed'].append('journey_api_ready')
        
        # Log completion
        onboarding_logger.log_process_data_complete(execution_state['steps_completed'], execution_state['errors'])
        
        # Close session with summary
        overall_status = 'SUCCESS' if not execution_state['errors'] else 'WARNING'
        close_onboarding_session(
            customer_id,
            overall_status,
            {
                'steps_completed': execution_state['steps_completed'],
                'errors': execution_state['errors'],
                'total_steps': len(execution_state['steps_completed'])
            }
        )
        
        return jsonify({
            'status': 'success',
            'message': 'Data processing completed successfully',
            'customer_id': customer_id,
            'log_file': onboarding_logger.get_log_file_path(),
            **execution_state
        })
        
    except Exception as e:
        current_app.logger.error(f"Error in process-data endpoint: {e}", exc_info=True)
        # Attempt rollback if we have execution state
        if 'execution_state' in locals():
            _rollback_operations(execution_state)
        return jsonify({
            'status': 'error',
            'message': f'Data processing failed: {str(e)}'
        }), 500


def _rollback_operations(execution_state: Dict, onboarding_logger=None):
    """
    Rollback partial operations if processing fails
    
    Currently logs what would be rolled back.
    In production, could delete created records, collections, etc.
    """
    customer_id = execution_state.get('customer_id')
    steps_completed = execution_state.get('steps_completed', [])
    
    if not steps_completed:
        return
    
    current_app.logger.warning(f"⚠️  Rollback needed for customer {customer_id}")
    current_app.logger.warning(f"   Steps completed: {', '.join(steps_completed)}")
    
    # Rollback logic (simplified - in production would delete created data)
    rollback_actions = []
    
    if 'embeddings' in steps_completed:
        rollback_actions.append("Delete Qdrant collection")
    
    if 'data_loading' in steps_completed:
        rollback_actions.append("Delete PostgreSQL data for customer")
    
    if 'journey_generation' in steps_completed:
        rollback_actions.append("Delete journey JSON files")
    
    current_app.logger.warning(f"   Rollback actions needed: {', '.join(rollback_actions)}")
    current_app.logger.warning(f"   ⚠️  Manual cleanup may be required for customer {customer_id}")
    
    # Log to onboarding logger if provided
    if onboarding_logger:
        onboarding_logger.log_rollback(steps_completed, rollback_actions)


@onboarding_api.route('/api/onboarding/templates/<file_type>', methods=['GET'])
def download_template(file_type):
    """
    Download CSV template file for onboarding
    
    Supported file types:
    - accounts
    - kpis (maps to kpi_measurements.csv)
    - signals (maps to qualitative_signals.csv)
    - products
    - profiles (maps to account_profiles.csv)
    - customers
    
    Returns the template CSV file for download
    """
    # Map file_type to actual filename
    template_map = {
        'accounts': 'accounts.csv',
        'kpis': 'kpi_measurements.csv',
        'signals': 'qualitative_signals.csv',
        'products': 'products.csv',
        'profiles': 'account_profiles.csv',
        'customers': 'customers.csv'
    }
    
    if file_type not in template_map:
        return jsonify({
            'status': 'error',
            'message': f'Invalid file type: {file_type}. Supported: {", ".join(template_map.keys())}'
        }), 400
    
    template_filename = template_map[file_type]
    # Path calculation: onboarding_api.py is in backend/, so parent is backend/, parent.parent is kpi-dashboard/
    # But we need: backend/verticals/_template/templates/
    template_path = Path(__file__).parent / "verticals" / "_template" / "templates" / template_filename
    
    if not template_path.exists():
        current_app.logger.error(f"Template file not found: {template_path}")
        return jsonify({
            'status': 'error',
            'message': f'Template file not found: {template_filename}'
        }), 404
    
    try:
        return send_file(
            str(template_path),
            as_attachment=True,
            download_name=template_filename,
            mimetype='text/csv'
        )
    except Exception as e:
        current_app.logger.error(f"Error serving template file: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to serve template file: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/templates', methods=['GET'])
def list_templates():
    """
    List all available template files
    
    Returns list of available templates with descriptions
    """
    templates = [
        {
            'file_type': 'accounts',
            'filename': 'accounts.csv',
            'description': 'Account master data with profile metadata',
            'required': True
        },
        {
            'file_type': 'kpis',
            'filename': 'kpi_measurements.csv',
            'description': 'KPI time-series measurements',
            'required': True
        },
        {
            'file_type': 'signals',
            'filename': 'qualitative_signals.csv',
            'description': 'Qualitative signals (emails, meetings, escalations)',
            'required': False
        },
        {
            'file_type': 'products',
            'filename': 'products.csv',
            'description': 'Product catalog',
            'required': False
        },
        {
            'file_type': 'profiles',
            'filename': 'account_profiles.csv',
            'description': 'Extended account profile attributes',
            'required': False
        },
        {
            'file_type': 'customers',
            'filename': 'customers.csv',
            'description': 'Customer/tenant-level data',
            'required': False
        }
    ]
    
    return jsonify({
        'status': 'success',
        'templates': templates,
        'download_url': '/api/onboarding/templates/{file_type}'
    })


@onboarding_api.route('/api/onboarding/processing-status', methods=['GET'])
def get_processing_status():
    """
    Get processing status for a customer
    
    Returns which scripts have been executed and their status
    """
    try:
        customer_id = get_current_customer_id() or request.args.get('customer_id')
        
        if not customer_id:
            return jsonify({'status': 'error', 'message': 'customer_id required'}), 400
        
        customer_id = int(customer_id)
        customer_dir = get_customer_directory(customer_id)
        
        if not customer_dir.exists():
            return jsonify({
                'status': 'error',
                'message': 'Customer directory not found'
            }), 404
        
        status = {
            'customer_id': customer_id,
            'directory_exists': customer_dir.exists(),
            'data_directory_exists': (customer_dir / "data").exists(),
            'scripts': {},
            'journey_api_exists': False
        }
        
        # Check script files
        scripts = {
            'data_loading': customer_dir / "scripts" / f"02_load_customer{customer_id}_data_SMART.py",
            'embeddings': customer_dir / "scripts" / f"03_embed_customer{customer_id}_OPENAI.py",
            'validation': customer_dir / "scripts" / f"04_validate_data_integrity.py",
            'journey_generator': customer_dir / "journey" / "wizard_a" / "wizard_journey_generator.py"
        }
        
        for step, script_path in scripts.items():
            status['scripts'][step] = {
                'exists': script_path.exists(),
                'path': str(script_path.relative_to(Path(__file__).parent.parent)) if script_path.exists() else None
            }
        
        # Check journey API
        journey_api = customer_dir / "journey" / f"customer{customer_id}_journey_api.py"
        status['journey_api_exists'] = journey_api.exists()
        
        # Check for journey data
        wizard_a_dir = customer_dir / "journey" / "wizard_a"
        journey_data_exists = False
        if wizard_a_dir.exists():
            # Look for test_run directories
            test_runs = [d for d in wizard_a_dir.iterdir() if d.is_dir() and d.name.startswith('test_run')]
            if test_runs:
                journey_data_exists = True
                status['latest_journey_run'] = max(test_runs, key=lambda x: x.stat().st_mtime).name
        
        status['journey_data_exists'] = journey_data_exists
        
        return jsonify({
            'status': 'success',
            **status
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting processing status: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Failed to get status: {str(e)}'
        }), 500


@onboarding_api.route('/api/onboarding/next-customer-id', methods=['GET'])
def get_next_customer_id():
    """Get next available customer ID"""
    try:
        from sqlalchemy import func
        max_id = db.session.query(func.max(Customer.customer_id)).scalar() or 0
        next_id = max_id + 1
        account_id_start = next_id * 1000
        return jsonify({
            'next_customer_id': next_id,
            'account_id_range': {
                'start': account_id_start + 1,
                'end': account_id_start + 10
            }
        })
    except Exception as e:
        current_app.logger.error(f"Error getting next customer ID: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@onboarding_api.route('/api/onboarding/validate-customer-id/<int:customer_id>', methods=['GET'])
def validate_customer_id(customer_id):
    """Validate if customer ID exists"""
    try:
        exists = Customer.query.filter_by(customer_id=customer_id).first() is not None
        if exists:
            # Get next available ID
            from sqlalchemy import func
            max_id = db.session.query(func.max(Customer.customer_id)).scalar() or 0
            next_id = max_id + 1
            return jsonify({
                'exists': True,
                'valid': False,
                'message': f'Customer ID {customer_id} already exists',
                'suggested_id': next_id
            })
        else:
            account_id_start = customer_id * 1000
            return jsonify({
                'exists': False,
                'valid': True,
                'message': 'Available',
                'account_id_range': {
                    'start': account_id_start + 1,
                    'end': account_id_start + 10
                }
            })
    except Exception as e:
        current_app.logger.error(f"Error validating customer ID: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@onboarding_api.route('/api/onboarding/generate-sample-files', methods=['POST'])
def generate_sample_files():
    """
    Generate synthetic sample CSV files for onboarding wizard.
    Uses generate_synthetic_customer_data.py which aligns with provision_dc_customer.py.
    Returns ZIP file with:
    - accounts.csv (account master data)
    - kpi_measurements.csv (monthly KPI time-series)
    - qualitative_signals.csv (emails, meetings, escalations)
    - products.csv (product catalog)
    - profiles.csv (extended account attributes)
    - customers.csv (tenant-level data)
    - DEMO_MANIFEST.md (markdown guide for demo scenarios)
    
    Account IDs use formula: 10000 + customer_id * 1000
    """
    try:
        data = request.json or {}
        
        # Get parameters from request
        customer_id = data.get('customer_id')  # Required, validated by frontend
        industry = data.get('industry', 'Technology')  # From Step 1
        company_name = data.get('company_name', 'Demo Customer')  # From Step 1
        num_accounts = data.get('num_accounts', 10)  # Default 10 accounts
        num_months = 12  # Fixed to 12 months (per requirements)
        
        # Customer ID is required and should be system-generated (validated by frontend)
        if not customer_id:
            return jsonify({
                'status': 'error',
                'message': 'Customer ID is required'
            }), 400
        
        # Validate customer_id doesn't exist (safety check)
        existing = Customer.query.filter_by(customer_id=customer_id).first()
        if existing:
            from sqlalchemy import func
            max_id = db.session.query(func.max(Customer.customer_id)).scalar() or 0
            next_id = max_id + 1
            return jsonify({
                'status': 'error',
                'message': f'Customer ID {customer_id} already exists',
                'suggested_id': next_id
            }), 400
        
        current_app.logger.info(f"Generating sample files: customer_id={customer_id}, industry={industry}, company={company_name}, accounts={num_accounts}")
        
        # Import generation function
        import tempfile
        import zipfile
        from io import BytesIO
        
        # Create temporary directory for generated files
        temp_dir = Path(tempfile.mkdtemp())
        
        try:
            # Import the synthetic customer data generator module
            import sys
            import importlib.util
            
            # Load generate_synthetic_customer_data.py as a module
            synth_gen_path = Path(__file__).parent / 'generate_synthetic_customer_data.py'
            if not synth_gen_path.exists():
                raise ImportError(f"Synthetic data generator not found at {synth_gen_path}")
            
            spec = importlib.util.spec_from_file_location("synthetic_data_generator", synth_gen_path)
            synth_gen = importlib.util.module_from_spec(spec)
            sys.modules['synthetic_data_generator'] = synth_gen
            spec.loader.exec_module(synth_gen)
            
            # Override MONTHS constant to match request (fixed to 12 per requirements)
            original_months = synth_gen.MONTHS
            synth_gen.MONTHS = num_months
            
            try:
                # Generate data using the synthetic generator functions
                # Pass industry and company_name as parameters (not overriding constants)
                accounts = synth_gen.generate_accounts(customer_id, num_accounts, industry=industry, company_name=company_name)
                kpi_measurements = synth_gen.generate_kpi_measurements(accounts)
                signals = synth_gen.generate_qualitative_signals(accounts)
                products = synth_gen.generate_products()
                profiles = synth_gen.generate_profiles(accounts)
                customers = synth_gen.generate_customers_csv(customer_id, company_name=company_name, industry=industry)
                
                # Write CSV files using the generator's write functions
                synth_gen.write_accounts_csv(accounts, temp_dir)
                synth_gen.write_kpi_measurements_csv(kpi_measurements, temp_dir)
                synth_gen.write_qualitative_signals_csv(signals, temp_dir)
                synth_gen.write_products_csv(products, temp_dir)
                synth_gen.write_profiles_csv(profiles, temp_dir)
                synth_gen.write_customers_csv(customers, temp_dir)
                
                # Generate and write DEMO_MANIFEST.md (with company_name)
                demo_manifest_content = synth_gen.generate_demo_manifest(accounts, company_name=company_name)
                synth_gen.write_demo_manifest(demo_manifest_content, temp_dir)
                
            finally:
                # Restore original constants
                synth_gen.MONTHS = original_months
            
            # DEMO_MANIFEST.md is already generated by generate_synthetic_customer_data.py
            # Create ZIP file in memory
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add all CSV files
                for file_path in temp_dir.glob('*.csv'):
                    zip_file.write(file_path, file_path.name)
                # Add DEMO_MANIFEST.md if it exists
                manifest_md_path = temp_dir / 'DEMO_MANIFEST.md'
                if manifest_md_path.exists():
                    zip_file.write(manifest_md_path, manifest_md_path.name)
            
            zip_buffer.seek(0)
            
            # Return ZIP file
            from flask import Response
            return Response(
                zip_buffer.getvalue(),
                mimetype='application/zip',
                headers={
                    'Content-Disposition': f'attachment; filename=onboarding_sample_files_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
                }
            )
            
        finally:
            # Cleanup temp directory
            shutil.rmtree(temp_dir, ignore_errors=True)
            
    except ImportError as e:
        current_app.logger.warning(f"Seed data generator not available, using simple generator: {e}")
        # Fallback: Generate simple CSV files
        return _generate_simple_csv_files(data)
    except Exception as e:
        current_app.logger.error(f"Error generating sample files: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Failed to generate sample files: {str(e)}'
        }), 500


def _generate_simple_csv_files(data: dict):
    """Fallback simple CSV generator if main generator is not available"""
    import csv
    import io
    import zipfile
    from io import BytesIO
    
    # Get customer_id from request (required)
    customer_id = data.get('customer_id')
    if not customer_id:
        # Use placeholder if not provided (fallback only)
        customer_id = 99
    
    # Calculate account IDs using formula: customer_id * 1000
    account_id_start = customer_id * 1000
    
    vertical = data.get('vertical', 'dc2_s')
    num_accounts = data.get('num_accounts', 10)
    
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Generate accounts.csv
        accounts_csv = io.StringIO()
        accounts_writer = csv.writer(accounts_csv)
        accounts_writer.writerow(['account_id', 'customer_id', 'account_name', 'revenue', 'industry', 'region', 'account_status'])
        for i in range(1, num_accounts + 1):
            account_id = account_id_start + i
            accounts_writer.writerow([
                account_id, customer_id, f'Demo Account {i}',
                random.randint(100000, 5000000),
                data.get('industry', 'Technology'),
                'US-West', 'active'
            ])
        zip_file.writestr('accounts.csv', accounts_csv.getvalue())
        
        # Generate kpi_measurements.csv
        kpis_csv = io.StringIO()
        kpis_writer = csv.writer(kpis_csv)
        kpis_writer.writerow(['account_id', 'account_name', 'date', 'kpi_code', 'kpi_name', 'pillar', 'value', 'target', 'unit'])
        for i in range(1, num_accounts + 1):
            account_id = account_id_start + i
            for month in range(1, 7):
                date = f'2024-{month:02d}-01'
                kpis_writer.writerow([
                    account_id, f'Demo Account {i}', date,
                    'P1-KPI1', 'System Uptime', 'P1',
                    round(random.uniform(95, 100), 2), 99.9, '%'
                ])
        zip_file.writestr('kpi_measurements.csv', kpis_csv.getvalue())
        
        # Generate qualitative_signals.csv
        signals_csv = io.StringIO()
        signals_writer = csv.writer(signals_csv)
        signals_writer.writerow(['account_id', 'account_name', 'date', 'signal_type', 'from_contact', 'to_contact', 'summary', 'sentiment'])
        for i in range(1, num_accounts + 1):
            account_id = 10000 + i
            for month in range(1, 7):
                date = f'2024-{month:02d}-15'
                signals_writer.writerow([
                    account_id, f'Demo Account {i}', date,
                    'email', 'customer@example.com', 'csm@example.com',
                    f'Monthly check-in for Account {i}', 'positive'
                ])
        zip_file.writestr('qualitative_signals.csv', signals_csv.getvalue())
        
        # Generate products.csv
        products_csv = io.StringIO()
        products_writer = csv.writer(products_csv)
        products_writer.writerow(['id', 'name', 'category'])
        products_data = [
            ['DC-GPU-H100', 'DGX H100 Systems', 'Compute'],
            ['DC-COOL-LQ', 'Liquid Cooling Solutions', 'Infrastructure'],
            ['DC-SEC-ADV', 'Advanced Security Suite', 'Security'],
            ['DC-SUP-PREM', 'Premium Support', 'Support'],
            ['DC-MON-AI', 'AI Monitoring Platform', 'Monitoring']
        ]
        for product in products_data:
            products_writer.writerow(product)
        zip_file.writestr('products.csv', products_csv.getvalue())
        
        # Generate profiles.csv
        profiles_csv = io.StringIO()
        profiles_writer = csv.writer(profiles_csv)
        profiles_writer.writerow(['account_id', 'account_name', 'health_score', 'lifecycle_stage', 'csm_name', 'arr', 'num_licenses', 'deployment_type', 'data_center_size'])
        for i in range(1, num_accounts + 1):
            account_id = account_id_start + i
            profiles_writer.writerow([
                account_id, f'Demo Account {i}',
                round(random.uniform(60, 95), 1),  # health_score
                random.choice(['onboarding', 'adoption', 'expansion', 'renewal']),  # lifecycle_stage
                f'CSM {i}',  # csm_name
                random.randint(100000, 5000000),  # arr
                random.randint(50, 200),  # num_licenses
                random.choice(['On-Premise', 'Hybrid', 'Cloud']),  # deployment_type
                random.choice(['Small', 'Medium', 'Large'])  # data_center_size
            ])
        zip_file.writestr('profiles.csv', profiles_csv.getvalue())
        
        # Generate demo_manifest.json
        manifest = {
            "version": "1.0",
            "customer_name": data.get('company_name', 'Demo Customer'),
            "vertical": vertical,
            "industry": data.get('industry', 'Technology'),
            "generated_at": datetime.now().isoformat(),
            "files": {
                "accounts.csv": {"description": "Account list", "row_count": num_accounts},
                "kpi_measurements.csv": {"description": "KPI measurements", "row_count": num_accounts * 6},
                "qualitative_signals.csv": {"description": "Qualitative signals", "row_count": num_accounts * 6}
            }
        }
        zip_file.writestr('demo_manifest.json', json.dumps(manifest, indent=2))
    
    zip_buffer.seek(0)
    from flask import Response
    return Response(
        zip_buffer.getvalue(),
        mimetype='application/zip',
        headers={
            'Content-Disposition': f'attachment; filename=onboarding_sample_files_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        }
    )


@onboarding_api.route('/api/onboarding/register-journey-api', methods=['POST'])
def register_journey_api():
    """
    Dynamically register journey API blueprint for a customer
    
    Expected JSON body:
    {
        "customer_id": 18
    }
    
    This endpoint:
    1. Finds customer{N}_journey_api.py in customer directory
    2. Dynamically imports and registers the blueprint
    3. Returns registration status
    """
    try:
        data = request.get_json() or {}
        customer_id = data.get('customer_id') or get_current_customer_id()
        
        if not customer_id:
            return jsonify({'status': 'error', 'message': 'customer_id required'}), 400
        
        customer_id = int(customer_id)
        customer_dir = get_customer_directory(customer_id)
        journey_api_path = customer_dir / "journey" / f"customer{customer_id}_journey_api.py"
        
        if not journey_api_path.exists():
            return jsonify({
                'status': 'error',
                'message': f'Journey API adapter not found: {journey_api_path}'
            }), 404
        
        # Try to dynamically import and register
        try:
            # Add journey directory to path
            journey_dir = journey_api_path.parent
            if str(journey_dir) not in sys.path:
                sys.path.insert(0, str(journey_dir))
            
            # Import the module
            module_name = f"customer{customer_id}_journey_api"
            
            # Remove from sys.modules if already loaded (force reload)
            if module_name in sys.modules:
                del sys.modules[module_name]
            
            # Import
            import importlib.util
            spec = importlib.util.spec_from_file_location(module_name, journey_api_path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            
            # Find blueprint (usually named journey_api_c{N} or similar)
            from flask import Blueprint
            blueprint = None
            for attr_name in dir(module):
                attr = getattr(module, attr_name)
                if isinstance(attr, Blueprint):
                    blueprint = attr
                    break
            
            # Fallback: look for common names
            if not blueprint:
                for name in [f'journey_api_c{customer_id}', 'journey_api', f'customer{customer_id}_journey_api']:
                    if hasattr(module, name):
                        blueprint = getattr(module, name)
                        if isinstance(blueprint, Blueprint):
                            break
            
            if not blueprint:
                return jsonify({
                    'status': 'error',
                    'message': 'Could not find Blueprint in journey API module'
                }), 500
            
            # Register blueprint (check if already registered)
            blueprint_name = blueprint.name
            if blueprint_name in current_app.blueprints:
                current_app.logger.info(f"Blueprint {blueprint_name} already registered")
                return jsonify({
                    'status': 'success',
                    'message': f'Journey API already registered',
                    'blueprint_name': blueprint_name,
                    'customer_id': customer_id
                })
            
            # Register
            current_app.register_blueprint(blueprint)
            current_app.logger.info(f"✅ Registered journey API blueprint: {blueprint_name} for customer {customer_id}")
            
            # Log to onboarding logger
            onboarding_logger = get_onboarding_logger(customer_id)
            onboarding_logger.log_journey_api_registration(blueprint_name, True)
            
            return jsonify({
                'status': 'success',
                'message': f'Journey API registered successfully',
                'blueprint_name': blueprint_name,
                'customer_id': customer_id,
                'url_prefix': blueprint.url_prefix if hasattr(blueprint, 'url_prefix') else None,
                'log_file': onboarding_logger.get_log_file_path()
            })
            
        except Exception as e:
            error_msg = f'Failed to register journey API: {str(e)}'
            current_app.logger.error(error_msg, exc_info=True)
            
            # Log to onboarding logger
            onboarding_logger = get_onboarding_logger(customer_id)
            onboarding_logger.log_journey_api_registration('', False, error_msg)
            
            return jsonify({
                'status': 'error',
                'message': error_msg,
                'log_file': onboarding_logger.get_log_file_path()
            }), 500
            
    except Exception as e:
        current_app.logger.error(f"Error in register-journey-api endpoint: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Unexpected error: {str(e)}'
        }), 500
