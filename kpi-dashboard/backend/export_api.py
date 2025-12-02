from flask import Blueprint, request, jsonify, send_file
from auth_middleware import get_current_customer_id, get_current_user_id
from extensions import db
from models import (
    KPI, KPIUpload, Account, Product, CustomerConfig, HealthTrend, 
    KPITimeSeries, KPIReferenceRange, PlaybookTrigger, PlaybookExecution, 
    PlaybookReport, FeatureToggle
)
import pandas as pd
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import json

export_api = Blueprint('export_api', __name__)

@export_api.route('/api/export/<int:upload_id>', methods=['GET'])
def export_excel(upload_id):
    """Export KPI data back to Excel format"""
    try:
        # Get the upload record
        upload = KPIUpload.query.get_or_404(upload_id)
        
        # Get all KPIs for this upload
        kpis = KPI.query.filter_by(upload_id=upload_id).all()
        
        if not kpis:
            return jsonify({'error': 'No KPIs found for this upload'}), 404
        
        # Group KPIs by category
        kpis_by_category = {}
        for kpi in kpis:
            if kpi.category not in kpis_by_category:
                kpis_by_category[kpi.category] = []
            kpis_by_category[kpi.category].append(kpi)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add metadata sheet (first sheet)
        metadata_sheet = wb.create_sheet("Metadata")
        metadata_sheet['A1'] = "KPI Dashboard Export"
        metadata_sheet['A2'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        metadata_sheet['A3'] = f"Original File: {upload.original_filename}"
        metadata_sheet['A4'] = f"Version: {upload.version}"
        metadata_sheet['A5'] = f"Upload Date: {upload.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')}"
        metadata_sheet['A6'] = f"Total KPIs: {len(kpis)}"
        
        # Style metadata
        for row in range(1, 7):
            metadata_sheet[f'A{row}'].font = Font(bold=True)
        
        # Add KPI sheets
        for category, category_kpis in kpis_by_category.items():
            sheet = wb.create_sheet(category)
            
            # Add header row
            headers = [
                'Health Score Component',
                'Weight',
                'Data',
                'Source Review',
                'KPI/Parameter',
                'Impact level',
                'Measurement Frequency'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add KPI data
            for row, kpi in enumerate(category_kpis, 2):
                sheet.cell(row=row, column=1, value=kpi.health_score_component)
                sheet.cell(row=row, column=2, value=kpi.weight)
                sheet.cell(row=row, column=3, value=kpi.data)
                sheet.cell(row=row, column=4, value=kpi.source_review)
                sheet.cell(row=row, column=5, value=kpi.kpi_parameter)
                sheet.cell(row=row, column=6, value=kpi.impact_level)
                sheet.cell(row=row, column=7, value=kpi.measurement_frequency)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet (last sheet)
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = "KPI Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        summary_sheet['A3'] = "Category Distribution:"
        summary_sheet['A3'].font = Font(bold=True)
        
        row = 4
        for category, category_kpis in kpis_by_category.items():
            summary_sheet[f'A{row}'] = f"{category}: {len(category_kpis)} KPIs"
            row += 1
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"KPI_Export_v{upload.version}_{timestamp}.xlsx"
        
        return send_file(
            excel_bytes,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_api.route('/api/export/<int:upload_id>/preview', methods=['GET'])
def export_preview(upload_id):
    """Get preview of export data"""
    try:
        upload = KPIUpload.query.get_or_404(upload_id)
        kpis = KPI.query.filter_by(upload_id=upload_id).all()
        
        # Group KPIs by category
        kpis_by_category = {}
        for kpi in kpis:
            if kpi.category not in kpis_by_category:
                kpis_by_category[kpi.category] = []
            kpis_by_category[kpi.category].append({
                'health_score_component': kpi.health_score_component,
                'weight': kpi.weight,
                'data': kpi.data,
                'source_review': kpi.source_review,
                'kpi_parameter': kpi.kpi_parameter,
                'impact_level': kpi.impact_level,
                'measurement_frequency': kpi.measurement_frequency
            })
        
        return jsonify({
            'upload_id': upload_id,
            'version': upload.version,
            'uploaded_at': upload.uploaded_at.isoformat(),
            'original_filename': upload.original_filename,
            'total_kpis': len(kpis),
            'categories': kpis_by_category
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_api.route('/api/export/all-account-data', methods=['GET'])
def export_all_account_data():
    """Export all account data including KPIs, products, and configuration in Excel format"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        customer_id = get_current_customer_id()
        logger.info(f"Export request received for customer_id: {customer_id}")
        
        # Get all accounts for this customer
        accounts = Account.query.filter_by(customer_id=customer_id).all()
        logger.info(f"Found {len(accounts)} accounts for customer {customer_id}")
        
        if not accounts:
            logger.warning(f"No accounts found for customer {customer_id}")
            return jsonify({'error': 'No accounts found'}), 404
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Accounts Summary (for rehydration)
        accounts_sheet = wb.create_sheet("Accounts Summary")
        accounts_headers = [
            'Account ID', 'Account Name', 'Revenue', 'Industry', 'Region', 
            'Status', 'Health Score', 'Products Used', 'External Account ID'
        ]
        for col, header in enumerate(accounts_headers, 1):
            cell = accounts_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add account data
        for row, account in enumerate(accounts, 2):
            # Get products for this account
            products = Product.query.filter_by(account_id=account.account_id).all()
            products_list = ', '.join([p.product_name for p in products]) if products else ''
            
            # Get health score
            from models import HealthTrend
            from playbook_recommendations_api import calculate_health_score_proxy
            latest_trend = HealthTrend.query.filter_by(
                account_id=account.account_id,
                customer_id=customer_id
            ).order_by(HealthTrend.year.desc(), HealthTrend.month.desc()).first()
            
            health_score = None
            if latest_trend and latest_trend.overall_health_score:
                health_score = float(latest_trend.overall_health_score)
            else:
                health_score = calculate_health_score_proxy(account.account_id)
            
            accounts_sheet.cell(row=row, column=1, value=account.account_id)
            accounts_sheet.cell(row=row, column=2, value=account.account_name)
            accounts_sheet.cell(row=row, column=3, value=float(account.revenue) if account.revenue else 0)
            accounts_sheet.cell(row=row, column=4, value=account.industry or '')
            accounts_sheet.cell(row=row, column=5, value=account.region or '')
            accounts_sheet.cell(row=row, column=6, value=account.account_status or '')
            accounts_sheet.cell(row=row, column=7, value=health_score if health_score else '')
            accounts_sheet.cell(row=row, column=8, value=products_list)
            accounts_sheet.cell(row=row, column=9, value=account.external_account_id or '')
            # Note: Profile Metadata is exported in separate "Customer Profile Data" sheet
        
        # Auto-adjust column widths for accounts sheet
        for column in accounts_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            accounts_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: All KPIs (for rehydration)
        kpis_sheet = wb.create_sheet("All KPIs")
        kpi_headers = [
            'KPI ID', 'Account ID', 'Account Name', 'Product ID', 'Product Name',
            'Category', 'KPI Parameter', 'Data', 'Impact Level', 
            'Measurement Frequency', 'Health Score Component', 'Weight', 'Aggregation Type'
        ]
        for col, header in enumerate(kpi_headers, 1):
            cell = kpis_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Get all KPIs for this customer
        kpis = db.session.query(KPI, Account, Product).join(
            KPIUpload, KPI.upload_id == KPIUpload.upload_id
        ).join(
            Account, KPI.account_id == Account.account_id, isouter=True
        ).outerjoin(
            Product, KPI.product_id == Product.product_id
        ).filter(KPIUpload.customer_id == customer_id).all()
        
        # Add KPI data
        for row, (kpi, account, product) in enumerate(kpis, 2):
            kpis_sheet.cell(row=row, column=1, value=kpi.kpi_id)
            kpis_sheet.cell(row=row, column=2, value=kpi.account_id)
            kpis_sheet.cell(row=row, column=3, value=account.account_name if account else '')
            kpis_sheet.cell(row=row, column=4, value=kpi.product_id if kpi.product_id else '')
            kpis_sheet.cell(row=row, column=5, value=product.product_name if product else '')
            kpis_sheet.cell(row=row, column=6, value=kpi.category or '')
            kpis_sheet.cell(row=row, column=7, value=kpi.kpi_parameter or '')
            kpis_sheet.cell(row=row, column=8, value=kpi.data or '')
            kpis_sheet.cell(row=row, column=9, value=kpi.impact_level or '')
            kpis_sheet.cell(row=row, column=10, value=kpi.measurement_frequency or '')
            kpis_sheet.cell(row=row, column=11, value=kpi.health_score_component or '')
            kpis_sheet.cell(row=row, column=12, value=kpi.weight or '')
            kpis_sheet.cell(row=row, column=13, value=kpi.aggregation_type or '')
        
        # Auto-adjust column widths for KPIs sheet
        for column in kpis_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            kpis_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Products
        products_sheet = wb.create_sheet("Products")
        product_headers = [
            'Product ID', 'Account ID', 'Account Name', 'Product Name', 
            'Product SKU', 'Product Type', 'Revenue', 'Status'
        ]
        for col, header in enumerate(product_headers, 1):
            cell = products_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Get all products for this customer
        all_products = db.session.query(Product, Account).join(
            Account, Product.account_id == Account.account_id
        ).filter(Product.customer_id == customer_id).all()
        
        for row, (product, account) in enumerate(all_products, 2):
            products_sheet.cell(row=row, column=1, value=product.product_id)
            products_sheet.cell(row=row, column=2, value=product.account_id)
            products_sheet.cell(row=row, column=3, value=account.account_name if account else '')
            products_sheet.cell(row=row, column=4, value=product.product_name)
            products_sheet.cell(row=row, column=5, value=product.product_sku or '')
            products_sheet.cell(row=row, column=6, value=product.product_type or '')
            products_sheet.cell(row=row, column=7, value=float(product.revenue) if product.revenue else 0)
            products_sheet.cell(row=row, column=8, value=product.status or '')
        
        # Auto-adjust column widths for products sheet
        for column in products_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            products_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 4: Customer Profile Data (separate tab for account/product cards)
        profile_sheet = wb.create_sheet("Customer Profile Data")
        profile_headers = [
            'Account ID', 'Account Name', 'External Account ID', 'Profile Metadata (JSON)'
        ]
        for col, header in enumerate(profile_headers, 1):
            cell = profile_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Export profile_metadata for each account
        for row, account in enumerate(accounts, 2):
            profile_sheet.cell(row=row, column=1, value=account.account_id)
            profile_sheet.cell(row=row, column=2, value=account.account_name)
            profile_sheet.cell(row=row, column=3, value=account.external_account_id or '')
            # Export profile_metadata as JSON string
            profile_metadata_json = ''
            if account.profile_metadata:
                try:
                    profile_metadata_json = json.dumps(account.profile_metadata) if isinstance(account.profile_metadata, dict) else str(account.profile_metadata)
                except:
                    profile_metadata_json = str(account.profile_metadata)
            profile_sheet.cell(row=row, column=4, value=profile_metadata_json)
        
        # Auto-adjust column widths for profile sheet
        for column in profile_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Wider for JSON
            profile_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 5: Playbook Triggers
        triggers_sheet = wb.create_sheet("Playbook Triggers")
        trigger_headers = [
            'Trigger ID', 'Playbook Type', 'Trigger Config (JSON)', 'Auto Trigger Enabled',
            'Last Evaluated', 'Last Triggered', 'Trigger Count'
        ]
        for col, header in enumerate(trigger_headers, 1):
            cell = triggers_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        triggers = PlaybookTrigger.query.filter_by(customer_id=customer_id).all()
        for row, trigger in enumerate(triggers, 2):
            triggers_sheet.cell(row=row, column=1, value=trigger.trigger_id)
            triggers_sheet.cell(row=row, column=2, value=trigger.playbook_type)
            triggers_sheet.cell(row=row, column=3, value=trigger.trigger_config or '')
            triggers_sheet.cell(row=row, column=4, value='Yes' if trigger.auto_trigger_enabled else 'No')
            triggers_sheet.cell(row=row, column=5, value=trigger.last_evaluated.strftime('%Y-%m-%d %H:%M:%S') if trigger.last_evaluated else '')
            triggers_sheet.cell(row=row, column=6, value=trigger.last_triggered.strftime('%Y-%m-%d %H:%M:%S') if trigger.last_triggered else '')
            triggers_sheet.cell(row=row, column=7, value=trigger.trigger_count or 0)
        
        # Auto-adjust column widths for triggers sheet
        for column in triggers_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            triggers_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 6: Playbook Executions
        executions_sheet = wb.create_sheet("Playbook Executions")
        execution_headers = [
            'Execution ID', 'Account ID', 'Account Name', 'Playbook ID', 'Status',
            'Current Step', 'Execution Data (JSON)', 'Started At', 'Completed At'
        ]
        for col, header in enumerate(execution_headers, 1):
            cell = executions_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        executions = PlaybookExecution.query.filter_by(customer_id=customer_id).all()
        for row, execution in enumerate(executions, 2):
            account_name = ''
            if execution.account_id:
                account = db.session.get(Account, execution.account_id)
                account_name = account.account_name if account else ''
            
            executions_sheet.cell(row=row, column=1, value=execution.execution_id)
            executions_sheet.cell(row=row, column=2, value=execution.account_id or '')
            executions_sheet.cell(row=row, column=3, value=account_name)
            executions_sheet.cell(row=row, column=4, value=execution.playbook_id)
            executions_sheet.cell(row=row, column=5, value=execution.status)
            executions_sheet.cell(row=row, column=6, value=execution.current_step or '')
            # Export execution_data as JSON string
            execution_data_json = ''
            if execution.execution_data:
                try:
                    execution_data_json = json.dumps(execution.execution_data) if isinstance(execution.execution_data, dict) else str(execution.execution_data)
                except:
                    execution_data_json = str(execution.execution_data)
            executions_sheet.cell(row=row, column=7, value=execution_data_json)
            executions_sheet.cell(row=row, column=8, value=execution.started_at.strftime('%Y-%m-%d %H:%M:%S') if execution.started_at else '')
            executions_sheet.cell(row=row, column=9, value=execution.completed_at.strftime('%Y-%m-%d %H:%M:%S') if execution.completed_at else '')
        
        # Auto-adjust column widths for executions sheet
        for column in executions_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            executions_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 7: Playbook Reports
        reports_sheet = wb.create_sheet("Playbook Reports")
        report_headers = [
            'Report ID', 'Execution ID', 'Account ID', 'Account Name', 'Playbook ID',
            'Playbook Name', 'Status', 'Report Data (JSON)', 'Duration', 'Steps Completed',
            'Total Steps', 'Started At', 'Completed At'
        ]
        for col, header in enumerate(report_headers, 1):
            cell = reports_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        reports = PlaybookReport.query.filter_by(customer_id=customer_id).all()
        for row, report in enumerate(reports, 2):
            reports_sheet.cell(row=row, column=1, value=report.report_id)
            reports_sheet.cell(row=row, column=2, value=report.execution_id)
            reports_sheet.cell(row=row, column=3, value=report.account_id or '')
            reports_sheet.cell(row=row, column=4, value=report.account_name or '')
            reports_sheet.cell(row=row, column=5, value=report.playbook_id)
            reports_sheet.cell(row=row, column=6, value=report.playbook_name)
            reports_sheet.cell(row=row, column=7, value=report.status)
            # Export report_data as JSON string
            report_data_json = ''
            if report.report_data:
                try:
                    report_data_json = json.dumps(report.report_data) if isinstance(report.report_data, dict) else str(report.report_data)
                except:
                    report_data_json = str(report.report_data)
            reports_sheet.cell(row=row, column=8, value=report_data_json)
            reports_sheet.cell(row=row, column=9, value=report.duration or '')
            reports_sheet.cell(row=row, column=10, value=report.steps_completed or 0)
            reports_sheet.cell(row=row, column=11, value=report.total_steps or '')
            reports_sheet.cell(row=row, column=12, value=report.started_at.strftime('%Y-%m-%d %H:%M:%S') if report.started_at else '')
            reports_sheet.cell(row=row, column=13, value=report.completed_at.strftime('%Y-%m-%d %H:%M:%S') if report.completed_at else '')
        
        # Auto-adjust column widths for reports sheet
        for column in reports_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            reports_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 8: Customer Config
        config_sheet = wb.create_sheet("Customer Config")
        config_headers = [
            'Config ID', 'KPI Upload Mode', 'Category Weights (JSON)', 'Master File Name'
        ]
        for col, header in enumerate(config_headers, 1):
            cell = config_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        config = CustomerConfig.query.filter_by(customer_id=customer_id).first()
        if config:
            config_sheet.cell(row=2, column=1, value=config.config_id)
            config_sheet.cell(row=2, column=2, value=config.kpi_upload_mode or '')
            config_sheet.cell(row=2, column=3, value=config.category_weights or '')
            config_sheet.cell(row=2, column=4, value=config.master_file_name or '')
        
        # Auto-adjust column widths for config sheet
        for column in config_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            config_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 9: Health Trends
        health_trends_sheet = wb.create_sheet("Health Trends")
        health_headers = [
            'Trend ID', 'Account ID', 'Account Name', 'Month', 'Year',
            'Overall Health Score', 'Product Usage Score', 'Support Score',
            'Customer Sentiment Score', 'Business Outcomes Score', 'Relationship Strength Score',
            'Total KPIs', 'Valid KPIs'
        ]
        for col, header in enumerate(health_headers, 1):
            cell = health_trends_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        health_trends = HealthTrend.query.filter_by(customer_id=customer_id).all()
        for row, trend in enumerate(health_trends, 2):
            account = db.session.get(Account, trend.account_id)
            account_name = account.account_name if account else ''
            
            health_trends_sheet.cell(row=row, column=1, value=trend.trend_id)
            health_trends_sheet.cell(row=row, column=2, value=trend.account_id)
            health_trends_sheet.cell(row=row, column=3, value=account_name)
            health_trends_sheet.cell(row=row, column=4, value=trend.month)
            health_trends_sheet.cell(row=row, column=5, value=trend.year)
            health_trends_sheet.cell(row=row, column=6, value=float(trend.overall_health_score) if trend.overall_health_score else '')
            health_trends_sheet.cell(row=row, column=7, value=float(trend.product_usage_score) if trend.product_usage_score else '')
            health_trends_sheet.cell(row=row, column=8, value=float(trend.support_score) if trend.support_score else '')
            health_trends_sheet.cell(row=row, column=9, value=float(trend.customer_sentiment_score) if trend.customer_sentiment_score else '')
            health_trends_sheet.cell(row=row, column=10, value=float(trend.business_outcomes_score) if trend.business_outcomes_score else '')
            health_trends_sheet.cell(row=row, column=11, value=float(trend.relationship_strength_score) if trend.relationship_strength_score else '')
            health_trends_sheet.cell(row=row, column=12, value=trend.total_kpis or 0)
            health_trends_sheet.cell(row=row, column=13, value=trend.valid_kpis or 0)
        
        # Auto-adjust column widths for health trends sheet
        for column in health_trends_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            health_trends_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 10: KPI Time Series
        time_series_sheet = wb.create_sheet("KPI Time Series")
        ts_headers = [
            'ID', 'KPI ID', 'Account ID', 'Account Name', 'Month', 'Year',
            'Value', 'Health Status', 'Health Score'
        ]
        for col, header in enumerate(ts_headers, 1):
            cell = time_series_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        time_series = KPITimeSeries.query.filter_by(customer_id=customer_id).all()
        for row, ts in enumerate(time_series, 2):
            account = db.session.get(Account, ts.account_id)
            account_name = account.account_name if account else ''
            
            time_series_sheet.cell(row=row, column=1, value=ts.id)
            time_series_sheet.cell(row=row, column=2, value=ts.kpi_id)
            time_series_sheet.cell(row=row, column=3, value=ts.account_id)
            time_series_sheet.cell(row=row, column=4, value=account_name)
            time_series_sheet.cell(row=row, column=5, value=ts.month)
            time_series_sheet.cell(row=row, column=6, value=ts.year)
            time_series_sheet.cell(row=row, column=7, value=float(ts.value) if ts.value else '')
            time_series_sheet.cell(row=row, column=8, value=ts.health_status or '')
            time_series_sheet.cell(row=row, column=9, value=float(ts.health_score) if ts.health_score else '')
        
        # Auto-adjust column widths for time series sheet
        for column in time_series_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            time_series_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 11: KPI Reference Ranges
        ref_ranges_sheet = wb.create_sheet("KPI Reference Ranges")
        ref_headers = [
            'Range ID', 'KPI Name', 'Unit', 'Higher Is Better', 'Critical Min', 'Critical Max',
            'Risk Min', 'Risk Max', 'Healthy Min', 'Healthy Max', 'Critical Range', 'Risk Range',
            'Healthy Range', 'Description'
        ]
        for col, header in enumerate(ref_headers, 1):
            cell = ref_ranges_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        ref_ranges = KPIReferenceRange.query.filter(
            (KPIReferenceRange.customer_id == customer_id) | (KPIReferenceRange.customer_id.is_(None))
        ).all()
        for row, ref_range in enumerate(ref_ranges, 2):
            ref_ranges_sheet.cell(row=row, column=1, value=ref_range.range_id)
            ref_ranges_sheet.cell(row=row, column=2, value=ref_range.kpi_name)
            ref_ranges_sheet.cell(row=row, column=3, value=ref_range.unit)
            ref_ranges_sheet.cell(row=row, column=4, value='Yes' if ref_range.higher_is_better else 'No')
            ref_ranges_sheet.cell(row=row, column=5, value=float(ref_range.critical_min) if ref_range.critical_min else '')
            ref_ranges_sheet.cell(row=row, column=6, value=float(ref_range.critical_max) if ref_range.critical_max else '')
            ref_ranges_sheet.cell(row=row, column=7, value=float(ref_range.risk_min) if ref_range.risk_min else '')
            ref_ranges_sheet.cell(row=row, column=8, value=float(ref_range.risk_max) if ref_range.risk_max else '')
            ref_ranges_sheet.cell(row=row, column=9, value=float(ref_range.healthy_min) if ref_range.healthy_min else '')
            ref_ranges_sheet.cell(row=row, column=10, value=float(ref_range.healthy_max) if ref_range.healthy_max else '')
            ref_ranges_sheet.cell(row=row, column=11, value=ref_range.critical_range or '')
            ref_ranges_sheet.cell(row=row, column=12, value=ref_range.risk_range or '')
            ref_ranges_sheet.cell(row=row, column=13, value=ref_range.healthy_range or '')
            ref_ranges_sheet.cell(row=row, column=14, value=ref_range.description or '')
        
        # Auto-adjust column widths for ref ranges sheet
        for column in ref_ranges_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ref_ranges_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 12: Feature Toggles
        toggles_sheet = wb.create_sheet("Feature Toggles")
        toggle_headers = [
            'ID', 'Feature Name', 'Enabled', 'Config (JSON)', 'Description'
        ]
        for col, header in enumerate(toggle_headers, 1):
            cell = toggles_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        toggles = FeatureToggle.query.filter_by(customer_id=customer_id).all()
        for row, toggle in enumerate(toggles, 2):
            toggles_sheet.cell(row=row, column=1, value=toggle.id)
            toggles_sheet.cell(row=row, column=2, value=toggle.feature_name)
            toggles_sheet.cell(row=row, column=3, value='Yes' if toggle.enabled else 'No')
            # Export config as JSON string
            config_json = ''
            if toggle.config:
                try:
                    config_json = json.dumps(toggle.config) if isinstance(toggle.config, dict) else str(toggle.config)
                except:
                    config_json = str(toggle.config)
            toggles_sheet.cell(row=row, column=4, value=config_json)
            toggles_sheet.cell(row=row, column=5, value=toggle.description or '')
        
        # Auto-adjust column widths for toggles sheet
        for column in toggles_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            toggles_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 13: Export Metadata (updated with all data counts)
        metadata_sheet = wb.create_sheet("Export Metadata")
        export_timestamp = datetime.now()
        export_version = 1  # Version number for this export format
        
        metadata_sheet['A1'] = "KPI Dashboard - All Account Data Export"
        metadata_sheet['A1'].font = Font(bold=True, size=14)
        metadata_sheet['A2'] = f"Export Date: {export_timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
        metadata_sheet['A3'] = f"Export Timestamp: {export_timestamp.isoformat()}"
        metadata_sheet['A4'] = f"Export Version: {export_version}"
        metadata_sheet['A5'] = f"Customer ID: {customer_id}"
        metadata_sheet['A6'] = f"Total Accounts: {len(accounts)}"
        metadata_sheet['A7'] = f"Total KPIs: {len(kpis)}"
        metadata_sheet['A8'] = f"Total Products: {len(all_products)}"
        metadata_sheet['A9'] = f"Total Playbook Triggers: {len(triggers)}"
        metadata_sheet['A10'] = f"Total Playbook Executions: {len(executions)}"
        metadata_sheet['A11'] = f"Total Playbook Reports: {len(reports)}"
        metadata_sheet['A12'] = f"Total Health Trends: {len(health_trends)}"
        metadata_sheet['A13'] = f"Total KPI Time Series: {len(time_series)}"
        metadata_sheet['A14'] = f"Total KPI Reference Ranges: {len(ref_ranges)}"
        metadata_sheet['A15'] = f"Total Feature Toggles: {len(toggles)}"
        metadata_sheet['A16'] = "WARNING: This file contains sensitive customer data. Keep secure."
        
        for row in range(1, 17):
            metadata_sheet[f'A{row}'].font = Font(bold=True)
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Generate filename with version and timestamp for ransomware protection
        timestamp = export_timestamp.strftime('%Y%m%d_%H%M%S')
        filename = f"Account_Data_Export_v{export_version}_{timestamp}.xlsx"
        
        # Return file with proper headers for download
        logger.info(f"Excel file generated successfully: {filename}, size: {excel_bytes.tell()} bytes")
        
        response = send_file(
            excel_bytes,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Add CORS headers if needed
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Sending Excel file response for customer {customer_id}")
        return response
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Export failed for customer {customer_id}: {str(e)}\n{error_trace}")
        traceback.print_exc()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500 