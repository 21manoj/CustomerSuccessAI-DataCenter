#!/usr/bin/env python3
"""
Create Excel file with product dimension KPI analysis
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the JSON data
with open("product_dimension_kpi_analysis.json", "r") as f:
    data = json.load(f)

headers = data["headers"]
rows = data["rows"]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Dimension KPI Analysis"

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
medium_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

yes_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Write data rows
for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Color code impact levels
        if col_idx == 4:  # Product Adoption Impact
            if value == "High":
                cell.fill = high_fill
            elif value == "Medium":
                cell.fill = medium_fill
            elif value == "Low":
                cell.fill = low_fill
        elif col_idx == 5:  # Support Impact
            if value == "High":
                cell.fill = high_fill
            elif value == "Medium":
                cell.fill = medium_fill
            elif value == "Low":
                cell.fill = low_fill
        elif col_idx == 6:  # Cross-sell/Upsell Impact
            if value == "High":
                cell.fill = high_fill
            elif value == "Medium":
                cell.fill = medium_fill
            elif value == "Low":
                cell.fill = low_fill
        elif col_idx == 7:  # Overall Impact
            if value == "High":
                cell.fill = high_fill
                cell.font = Font(bold=True)
            elif value == "Medium":
                cell.fill = medium_fill
            elif value == "Low":
                cell.fill = low_fill
        elif col_idx == 8:  # SKU Required
            if value == "Yes":
                cell.fill = yes_fill
                cell.font = Font(bold=True)

# Set column widths
column_widths = {
    "A": 18,  # Category
    "B": 45,  # KPI Name
    "C": 12,  # Unit
    "D": 20,  # Product Adoption Impact
    "E": 15,  # Support Impact
    "F": 20,  # Cross-sell/Upsell Impact
    "G": 25,  # Overall Product Dimension Impact
    "H": 18,  # SKU Tracking Required
    "I": 30,  # Business Value
    "J": 22,  # Example: Product A Value
    "K": 22,  # Example: Product B Value
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Set row height for header
ws.row_dimensions[1].height = 40

# Freeze first row
ws.freeze_panes = "A2"

# Add summary sheet
ws_summary = wb.create_sheet("Summary", 0)
ws_summary.title = "Summary"

summary_data = [
    ["Product Dimension KPI Impact Analysis", ""],
    ["", ""],
    ["Total KPIs Analyzed", len(rows)],
    ["", ""],
    ["Impact Level Breakdown", ""],
    ["High Impact (SKU Required)", sum(1 for r in rows if r[6] == "High")],
    ["Medium Impact (SKU Recommended)", sum(1 for r in rows if r[6] == "Medium")],
    ["Low Impact (Optional)", sum(1 for r in rows if r[6] == "Low")],
    ["", ""],
    ["SKU Tracking Required", sum(1 for r in rows if r[7] == "Yes")],
    ["", ""],
    ["Category Breakdown", ""],
]

# Category counts
categories = {}
for row in rows:
    category = row[0]
    if category not in categories:
        categories[category] = {"total": 0, "high": 0}
    categories[category]["total"] += 1
    if row[6] == "High":
        categories[category]["high"] += 1

for category, counts in sorted(categories.items()):
    summary_data.append([f"{category}", f"{counts['total']} KPIs ({counts['high']} high impact)"])

# Write summary
for row_idx, row_data in enumerate(summary_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_idx <= 11:
            if col_idx == 1 and row_idx not in [1, 3, 5, 9, 11]:
                cell.font = Font(bold=True)

# Set summary column widths
ws_summary.column_dimensions["A"].width = 35
ws_summary.column_dimensions["B"].width = 30

# Save file
filename = "Product_Dimension_KPI_Analysis.xlsx"
wb.save(filename)
print(f"✅ Excel file created: {filename}")

